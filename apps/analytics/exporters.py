"""
Analytics exporters for exporting analytical data to various formats.
This module provides functionality to export analytics data to Excel format.
"""

from decimal import Decimal
from datetime import date, datetime
from typing import Dict, Any, List, Optional
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


class AnalyticsExporter:
    """
    Exporter for analytics data to Excel format.
    Handles export of all types of analytics with proper formatting and headers.
    """

    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.right_alignment = Alignment(horizontal="right", vertical="center")

    def _format_value(self, value: Any) -> str:
        """
        Format a value for display in Excel.

        Args:
            value: Value to format

        Returns:
            Formatted string representation
        """
        if value is None:
            return ""
        elif isinstance(value, Decimal):
            return f"{value:,.2f}"
        elif isinstance(value, (int, float)):
            return f"{value:,}"
        elif isinstance(value, (date, datetime)):
            return value.strftime("%Y-%m-%d")
        elif isinstance(value, dict):
            return str(value)
        else:
            return str(value)

    def _apply_header_style(self, worksheet, row_num: int, num_columns: int):
        """
        Apply header styling to a row.

        Args:
            worksheet: Excel worksheet
            row_num: Row number to style
            num_columns: Number of columns to style
        """
        for col in range(1, num_columns + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border

    def _apply_data_style(
        self, worksheet, start_row: int, end_row: int, num_columns: int
    ):
        """
        Apply data styling to rows.

        Args:
            worksheet: Excel worksheet
            start_row: Starting row number
            end_row: Ending row number
            num_columns: Number of columns to style
        """
        for row in range(start_row, end_row + 1):
            for col in range(1, num_columns + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = self.border
                # Right-align numeric columns
                if col > 1:  # Assuming first column is text
                    cell.alignment = self.right_alignment

    def _auto_adjust_columns(self, worksheet):
        """
        Auto-adjust column widths based on content.

        Args:
            worksheet: Excel worksheet
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def export_dashboard_metrics(
        self, metrics: Dict[str, Any], applied_filters: Optional[Dict[str, Any]] = None
    ) -> HttpResponse:
        """
        Export dashboard metrics to Excel.

        Args:
            metrics: Dashboard metrics data
            applied_filters: Applied filters information

        Returns:
            HttpResponse with Excel file
        """
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Dashboard Metrics"

            # Add title
            worksheet.cell(row=1, column=1, value="Dashboard Metrics Report")
            worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
            worksheet.merge_cells("A1:B1")

            # Add export date
            worksheet.cell(
                row=2,
                column=1,
                value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            )

            # Add filter information if provided
            current_row = 3
            if applied_filters:
                worksheet.cell(row=current_row, column=1, value="Applied Filters:")
                worksheet.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1

                for filter_name, filter_value in applied_filters.items():
                    if filter_value:
                        worksheet.cell(
                            row=current_row,
                            column=1,
                            value=f"  {filter_name}: {filter_value}",
                        )
                        current_row += 1
                current_row += 1

            # Add metrics data
            worksheet.cell(row=current_row, column=1, value="Metric")
            worksheet.cell(row=current_row, column=2, value="Value")
            self._apply_header_style(worksheet, current_row, 2)
            current_row += 1

            # Key metrics
            metrics_to_export = [
                (
                    "Planned Premium Volume",
                    metrics.get(
                        "planned_premium_volume", metrics.get("total_premium_volume", 0)
                    ),
                ),
                ("Actual Premium Volume", metrics.get("actual_premium_volume", 0)),
                (
                    "Planned Commission Revenue",
                    metrics.get(
                        "planned_commission_revenue",
                        metrics.get("total_commission_revenue", 0),
                    ),
                ),
                (
                    "Actual Commission Revenue",
                    metrics.get("actual_commission_revenue", 0),
                ),
                ("Total Policy Count", metrics.get("total_policy_count", 0)),
                (
                    "Planned Insurance Sum",
                    metrics.get(
                        "planned_insurance_sum", metrics.get("total_insurance_sum", 0)
                    ),
                ),
                ("Actual Insurance Sum", metrics.get("actual_insurance_sum", 0)),
                ("Active Policies Count", metrics.get("active_policies_count", 0)),
                (
                    "Average Commission Rate (%)",
                    metrics.get("average_commission_rate", 0),
                ),
                ("Current Year Growth (%)", metrics.get("current_year_growth", 0)),
            ]

            for metric_name, metric_value in metrics_to_export:
                worksheet.cell(row=current_row, column=1, value=metric_name)
                worksheet.cell(
                    row=current_row, column=2, value=self._format_value(metric_value)
                )
                current_row += 1

            # Apply styling
            self._apply_data_style(
                worksheet, current_row - len(metrics_to_export), current_row - 1, 2
            )
            self._auto_adjust_columns(worksheet)

            # Save to BytesIO
            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response[
                "Content-Disposition"
            ] = f'attachment; filename="dashboard_metrics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

            return response

        except Exception as e:
            logger.error(f"Error exporting dashboard metrics: {e}")
            raise

    def export_branch_analytics(
        self,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ) -> HttpResponse:
        """
        Export branch analytics to Excel.

        Args:
            analytics: Branch analytics data
            applied_filters: Applied filters information

        Returns:
            HttpResponse with Excel file
        """
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Branch Analytics"

            # Add title and metadata
            worksheet.cell(row=1, column=1, value="Branch Analytics Report")
            worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
            worksheet.merge_cells("A1:F1")

            worksheet.cell(
                row=2,
                column=1,
                value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            )

            current_row = 3
            if applied_filters:
                worksheet.cell(row=current_row, column=1, value="Applied Filters:")
                worksheet.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1

                for filter_name, filter_value in applied_filters.items():
                    if filter_value:
                        worksheet.cell(
                            row=current_row,
                            column=1,
                            value=f"  {filter_name}: {filter_value}",
                        )
                        current_row += 1
                current_row += 1

            # Headers
            headers = [
                "Branch Name",
                "Premium Volume",
                "Commission Revenue",
                "Policy Count",
                "Insurance Sum",
                "Market Share (%)",
            ]
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=current_row, column=col, value=header)

            self._apply_header_style(worksheet, current_row, len(headers))
            current_row += 1

            # Data rows
            branch_metrics = analytics.get("branch_metrics", [])
            start_data_row = current_row

            for branch_metric in branch_metrics:
                branch_info = branch_metric.get("branch", {})
                worksheet.cell(
                    row=current_row, column=1, value=branch_info.get("name", "")
                )
                worksheet.cell(
                    row=current_row,
                    column=2,
                    value=self._format_value(branch_metric.get("premium_volume", 0)),
                )
                worksheet.cell(
                    row=current_row,
                    column=3,
                    value=self._format_value(
                        branch_metric.get("commission_revenue", 0)
                    ),
                )
                worksheet.cell(
                    row=current_row,
                    column=4,
                    value=self._format_value(branch_metric.get("policy_count", 0)),
                )
                worksheet.cell(
                    row=current_row,
                    column=5,
                    value=self._format_value(branch_metric.get("insurance_sum", 0)),
                )
                worksheet.cell(
                    row=current_row,
                    column=6,
                    value=self._format_value(branch_metric.get("market_share", 0)),
                )
                current_row += 1

            # Apply styling
            if current_row > start_data_row:
                self._apply_data_style(
                    worksheet, start_data_row, current_row - 1, len(headers)
                )

            self._auto_adjust_columns(worksheet)

            # Save to BytesIO
            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response[
                "Content-Disposition"
            ] = f'attachment; filename="branch_analytics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

            return response

        except Exception as e:
            logger.error(f"Error exporting branch analytics: {e}")
            raise

    def export_insurer_analytics(
        self,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ) -> HttpResponse:
        """
        Export insurer analytics to Excel.

        Args:
            analytics: Insurer analytics data
            applied_filters: Applied filters information

        Returns:
            HttpResponse with Excel file
        """
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Insurer Analytics"

            # Add title and metadata
            worksheet.cell(row=1, column=1, value="Insurer Analytics Report")
            worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
            worksheet.merge_cells("A1:F1")

            worksheet.cell(
                row=2,
                column=1,
                value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            )

            current_row = 3
            if applied_filters:
                worksheet.cell(row=current_row, column=1, value="Applied Filters:")
                worksheet.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1

                for filter_name, filter_value in applied_filters.items():
                    if filter_value:
                        worksheet.cell(
                            row=current_row,
                            column=1,
                            value=f"  {filter_name}: {filter_value}",
                        )
                        current_row += 1
                current_row += 1

            # Headers
            headers = [
                "Insurer Name",
                "Premium Volume",
                "Commission Revenue",
                "Policy Count",
                "Insurance Sum",
                "Market Share (%)",
            ]
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=current_row, column=col, value=header)

            self._apply_header_style(worksheet, current_row, len(headers))
            current_row += 1

            # Data rows
            insurer_metrics = analytics.get("insurer_metrics", [])
            start_data_row = current_row

            for insurer_metric in insurer_metrics:
                insurer_info = insurer_metric.get("insurer", {})
                worksheet.cell(
                    row=current_row, column=1, value=insurer_info.get("name", "")
                )
                worksheet.cell(
                    row=current_row,
                    column=2,
                    value=self._format_value(insurer_metric.get("premium_volume", 0)),
                )
                worksheet.cell(
                    row=current_row,
                    column=3,
                    value=self._format_value(
                        insurer_metric.get("commission_revenue", 0)
                    ),
                )
                worksheet.cell(
                    row=current_row,
                    column=4,
                    value=self._format_value(insurer_metric.get("policy_count", 0)),
                )
                worksheet.cell(
                    row=current_row,
                    column=5,
                    value=self._format_value(insurer_metric.get("insurance_sum", 0)),
                )
                worksheet.cell(
                    row=current_row,
                    column=6,
                    value=self._format_value(insurer_metric.get("market_share", 0)),
                )
                current_row += 1

            # Apply styling
            if current_row > start_data_row:
                self._apply_data_style(
                    worksheet, start_data_row, current_row - 1, len(headers)
                )

            self._auto_adjust_columns(worksheet)

            # Save to BytesIO
            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response[
                "Content-Disposition"
            ] = f'attachment; filename="insurer_analytics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

            return response

        except Exception as e:
            logger.error(f"Error exporting insurer analytics: {e}")
            raise

    def export_client_analytics(
        self,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ) -> HttpResponse:
        """
        Export client analytics to Excel.

        Args:
            analytics: Client analytics data
            applied_filters: Applied filters information

        Returns:
            HttpResponse with Excel file
        """
        try:
            workbook = Workbook()

            # Create multiple sheets for different client rankings

            # Sheet 1: Top Clients by Insurance Sum
            ws1 = workbook.active
            ws1.title = "Top by Insurance Sum"
            self._create_client_sheet(
                ws1,
                "Top Clients by Insurance Sum",
                analytics.get("top_clients_by_insurance_sum", []),
                applied_filters,
            )

            # Sheet 2: Top Clients by Commission
            ws2 = workbook.create_sheet(title="Top by Commission")
            self._create_client_sheet(
                ws2,
                "Top Clients by Commission Revenue",
                analytics.get("top_clients_by_commission", []),
                applied_filters,
            )

            # Sheet 3: Top Clients by Policy Count
            ws3 = workbook.create_sheet(title="Top by Policy Count")
            self._create_client_sheet(
                ws3,
                "Top Clients by Policy Count",
                analytics.get("top_clients_by_policy_count", []),
                applied_filters,
            )

            # Save to BytesIO
            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response[
                "Content-Disposition"
            ] = f'attachment; filename="client_analytics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

            return response

        except Exception as e:
            logger.error(f"Error exporting client analytics: {e}")
            raise

    def _create_client_sheet(
        self,
        worksheet,
        title: str,
        client_metrics: List[Dict],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """
        Create a client analytics sheet.

        Args:
            worksheet: Excel worksheet
            title: Sheet title
            client_metrics: List of client metrics
            applied_filters: Applied filters information
        """
        # Add title and metadata
        worksheet.cell(row=1, column=1, value=title)
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:G1")

        worksheet.cell(
            row=2,
            column=1,
            value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        current_row = 3
        if applied_filters:
            worksheet.cell(row=current_row, column=1, value="Applied Filters:")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for filter_name, filter_value in applied_filters.items():
                if filter_value:
                    worksheet.cell(
                        row=current_row,
                        column=1,
                        value=f"  {filter_name}: {filter_value}",
                    )
                    current_row += 1
            current_row += 1

        # Headers
        headers = [
            "Client Name",
            "INN",
            "Premium Volume",
            "Commission Revenue",
            "Policy Count",
            "Insurance Sum",
            "Avg Policy Value",
        ]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=current_row, column=col, value=header)

        self._apply_header_style(worksheet, current_row, len(headers))
        current_row += 1

        # Data rows
        start_data_row = current_row

        for client_metric in client_metrics:
            client_info = client_metric.get("client", {})
            worksheet.cell(row=current_row, column=1, value=client_info.get("name", ""))
            worksheet.cell(row=current_row, column=2, value=client_info.get("inn", ""))
            worksheet.cell(
                row=current_row,
                column=3,
                value=self._format_value(client_metric.get("premium_volume", 0)),
            )
            worksheet.cell(
                row=current_row,
                column=4,
                value=self._format_value(client_metric.get("commission_revenue", 0)),
            )
            worksheet.cell(
                row=current_row,
                column=5,
                value=self._format_value(client_metric.get("policy_count", 0)),
            )
            worksheet.cell(
                row=current_row,
                column=6,
                value=self._format_value(client_metric.get("insurance_sum", 0)),
            )
            worksheet.cell(
                row=current_row,
                column=7,
                value=self._format_value(client_metric.get("average_policy_value", 0)),
            )
            current_row += 1

        # Apply styling
        if current_row > start_data_row:
            self._apply_data_style(
                worksheet, start_data_row, current_row - 1, len(headers)
            )

        self._auto_adjust_columns(worksheet)

    def export_financial_analytics(
        self,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ) -> HttpResponse:
        """
        Export financial analytics to Excel.

        Args:
            analytics: Financial analytics data
            applied_filters: Applied filters information

        Returns:
            HttpResponse with Excel file
        """
        try:
            workbook = Workbook()

            # Sheet 1: Future forecast summary
            ws1 = workbook.active
            ws1.title = "Сводка прогноза"
            self._create_future_summary_sheet(ws1, analytics, applied_filters)

            # Sheet 2: Future monthly detail
            ws2 = workbook.create_sheet(title="Будущие месяцы")
            self._create_future_monthly_sheet(ws2, analytics, applied_filters)

            # Sheet 3: Current year bridge (actual + forecast)
            ws3 = workbook.create_sheet(title="Годовой мост")
            self._create_current_year_bridge_sheet(ws3, analytics, applied_filters)

            # Sheet 4: Payment contour (control block from the page)
            ws4 = workbook.create_sheet(title="Платежный контур")
            self._create_payment_analysis_sheet(ws4, analytics, applied_filters)

            # Save to BytesIO
            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response[
                "Content-Disposition"
            ] = f'attachment; filename="financial_analytics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

            return response

        except Exception as e:
            logger.error(f"Error exporting financial analytics: {e}")
            raise

    def _create_future_summary_sheet(
        self,
        worksheet,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """Create summary sheet for all future forecast periods."""
        worksheet.cell(row=1, column=1, value="Финансовая аналитика - Сводка прогноза")
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:C1")

        worksheet.cell(
            row=2,
            column=1,
            value=f"Дата выгрузки: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        current_row = 3
        if applied_filters:
            worksheet.cell(row=current_row, column=1, value="Примененные фильтры:")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for filter_name, filter_value in applied_filters.items():
                if filter_value:
                    worksheet.cell(
                        row=current_row,
                        column=1,
                        value=f"  {filter_name}: {filter_value}",
                    )
                    current_row += 1
            current_row += 1

        summary = analytics.get("future_forecast_summary", {})

        worksheet.cell(row=current_row, column=1, value="Показатель")
        worksheet.cell(row=current_row, column=2, value="Значение")
        self._apply_header_style(worksheet, current_row, 2)
        current_row += 1

        current_month = summary.get("current_month")
        next_month = summary.get("next_month")
        current_quarter = summary.get("current_quarter_remaining")
        next_quarter = summary.get("next_quarter")

        metrics = [
            ("Дата среза", self._format_value(summary.get("as_of_date"))),
            ("Горизонт прогноза (мес.)", summary.get("horizon_months", 0)),
            (
                "Всего будущая премия",
                self._format_value(summary.get("total_future_premium", Decimal("0"))),
            ),
            (
                "Всего будущая комиссия",
                self._format_value(
                    summary.get("total_future_commission", Decimal("0"))
                ),
            ),
            (
                "Текущий месяц",
                self._format_value(current_month.get("month"))
                if current_month
                else "N/A",
            ),
            (
                "Текущий месяц: премия",
                self._format_value(
                    current_month.get("forecasted_premium", Decimal("0"))
                )
                if current_month
                else "N/A",
            ),
            (
                "Текущий месяц: комиссия",
                self._format_value(
                    current_month.get("forecasted_commission", Decimal("0"))
                )
                if current_month
                else "N/A",
            ),
            (
                "Следующий месяц",
                self._format_value(next_month.get("month")) if next_month else "N/A",
            ),
            (
                "Следующий месяц: премия",
                self._format_value(next_month.get("forecasted_premium", Decimal("0")))
                if next_month
                else "N/A",
            ),
            (
                "Следующий месяц: комиссия",
                self._format_value(
                    next_month.get("forecasted_commission", Decimal("0"))
                )
                if next_month
                else "N/A",
            ),
            (
                "Остаток текущего квартала",
                current_quarter.get("quarter_label", "N/A")
                if current_quarter
                else "N/A",
            ),
            (
                "Текущий квартал: премия",
                self._format_value(
                    current_quarter.get("forecasted_premium", Decimal("0"))
                )
                if current_quarter
                else "N/A",
            ),
            (
                "Текущий квартал: комиссия",
                self._format_value(
                    current_quarter.get("forecasted_commission", Decimal("0"))
                )
                if current_quarter
                else "N/A",
            ),
            (
                "Следующий квартал",
                next_quarter.get("quarter_label", "N/A") if next_quarter else "N/A",
            ),
            (
                "Следующий квартал: премия",
                self._format_value(next_quarter.get("forecasted_premium", Decimal("0")))
                if next_quarter
                else "N/A",
            ),
            (
                "Следующий квартал: комиссия",
                self._format_value(
                    next_quarter.get("forecasted_commission", Decimal("0"))
                )
                if next_quarter
                else "N/A",
            ),
        ]

        start_data_row = current_row
        for label, value in metrics:
            worksheet.cell(row=current_row, column=1, value=label)
            worksheet.cell(row=current_row, column=2, value=value)
            current_row += 1

        if current_row > start_data_row:
            self._apply_data_style(worksheet, start_data_row, current_row - 1, 2)

        current_row += 2

        quarter_rows = analytics.get("future_quarterly_forecast", [])
        worksheet.cell(row=current_row, column=1, value="Квартал")
        worksheet.cell(row=current_row, column=2, value="Прогноз премии")
        worksheet.cell(row=current_row, column=3, value="Прогноз комиссии")
        worksheet.cell(row=current_row, column=4, value="Месяцев в расчете")
        self._apply_header_style(worksheet, current_row, 4)
        current_row += 1

        start_data_row = current_row
        for quarter in quarter_rows:
            worksheet.cell(
                row=current_row, column=1, value=quarter.get("quarter_label", "")
            )
            worksheet.cell(
                row=current_row,
                column=2,
                value=self._format_value(quarter.get("forecasted_premium", 0)),
            )
            worksheet.cell(
                row=current_row,
                column=3,
                value=self._format_value(quarter.get("forecasted_commission", 0)),
            )
            worksheet.cell(
                row=current_row, column=4, value=quarter.get("months_count", 0)
            )
            current_row += 1

        if current_row > start_data_row:
            self._apply_data_style(worksheet, start_data_row, current_row - 1, 4)

        self._auto_adjust_columns(worksheet)

    def _create_future_monthly_sheet(
        self,
        worksheet,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """Create monthly sheet for all future forecast periods."""
        worksheet.cell(
            row=1, column=1, value="Финансовая аналитика - Будущие месяцы (детализация)"
        )
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:G1")

        worksheet.cell(
            row=2,
            column=1,
            value=f"Дата выгрузки: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        current_row = 3
        if applied_filters:
            worksheet.cell(row=current_row, column=1, value="Примененные фильтры:")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for filter_name, filter_value in applied_filters.items():
                if filter_value:
                    worksheet.cell(
                        row=current_row,
                        column=1,
                        value=f"  {filter_name}: {filter_value}",
                    )
                    current_row += 1
            current_row += 1

        headers = [
            "Месяц",
            "Квартал",
            "Год",
            "Прогноз премии",
            "Прогноз комиссии",
            "Накопительно премия",
            "Накопительно комиссия",
        ]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=current_row, column=col, value=header)

        self._apply_header_style(worksheet, current_row, len(headers))
        current_row += 1

        monthly_rows = analytics.get("future_monthly_forecast") or analytics.get(
            "monthly_premium_forecast", []
        )
        running_premium_total = Decimal("0")
        running_commission_total = Decimal("0")
        start_data_row = current_row

        for row in monthly_rows:
            month_date = row.get("month")
            forecasted_premium = row.get("forecasted_premium", Decimal("0"))
            forecasted_commission = row.get("forecasted_commission", Decimal("0"))
            month_label = row.get("month_label")
            quarter_label = row.get("quarter")
            year_value = row.get("year")
            row_running_premium = row.get("running_premium_total")
            row_running_commission = row.get("running_commission_total")

            if row_running_premium is None:
                running_premium_total += forecasted_premium
            else:
                running_premium_total = row_running_premium

            if row_running_commission is None:
                running_commission_total += forecasted_commission
            else:
                running_commission_total = row_running_commission

            if not quarter_label and month_date:
                quarter_number = ((month_date.month - 1) // 3) + 1
                quarter_label = f"Q{quarter_number}"
            if not year_value and month_date:
                year_value = month_date.year

            worksheet.cell(
                row=current_row,
                column=1,
                value=month_label
                if month_label
                else self._format_value(month_date)
                if month_date
                else "",
            )
            worksheet.cell(
                row=current_row,
                column=2,
                value=quarter_label or "",
            )
            worksheet.cell(row=current_row, column=3, value=year_value or "")
            worksheet.cell(
                row=current_row,
                column=4,
                value=self._format_value(forecasted_premium),
            )
            worksheet.cell(
                row=current_row,
                column=5,
                value=self._format_value(forecasted_commission),
            )
            worksheet.cell(
                row=current_row,
                column=6,
                value=self._format_value(running_premium_total),
            )
            worksheet.cell(
                row=current_row,
                column=7,
                value=self._format_value(running_commission_total),
            )
            current_row += 1

        if current_row > start_data_row:
            self._apply_data_style(
                worksheet, start_data_row, current_row - 1, len(headers)
            )

        self._auto_adjust_columns(worksheet)

    def _create_current_year_bridge_sheet(
        self,
        worksheet,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """Create current year bridge sheet (actual elapsed + forecast remaining)."""
        worksheet.cell(row=1, column=1, value="Финансовая аналитика - Текущий год")
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:D1")

        worksheet.cell(
            row=2,
            column=1,
            value=f"Дата выгрузки: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        current_row = 3
        if applied_filters:
            worksheet.cell(row=current_row, column=1, value="Примененные фильтры:")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for filter_name, filter_value in applied_filters.items():
                if filter_value:
                    worksheet.cell(
                        row=current_row,
                        column=1,
                        value=f"  {filter_name}: {filter_value}",
                    )
                    current_row += 1
            current_row += 1

        outlook = analytics.get("current_year_outlook", {})

        worksheet.cell(row=current_row, column=1, value="Показатель")
        worksheet.cell(row=current_row, column=2, value="Значение")
        self._apply_header_style(worksheet, current_row, 2)
        current_row += 1

        metrics = [
            ("Год", outlook.get("year", datetime.now().year)),
            ("Текущий месяц (номер)", outlook.get("current_month", 0)),
            (
                "Факт с начала года: премия",
                self._format_value(outlook.get("ytd_actual_premium", Decimal("0"))),
            ),
            (
                "Факт с начала года: комиссия",
                self._format_value(outlook.get("ytd_actual_commission", Decimal("0"))),
            ),
            (
                "Прогноз до конца года: премия",
                self._format_value(
                    outlook.get("remaining_forecast_premium", Decimal("0"))
                ),
            ),
            (
                "Прогноз до конца года: комиссия",
                self._format_value(
                    outlook.get("remaining_forecast_commission", Decimal("0"))
                ),
            ),
            (
                "Итог года: премия",
                self._format_value(
                    outlook.get("projected_full_year_premium", Decimal("0"))
                ),
            ),
            (
                "Итог года: комиссия",
                self._format_value(
                    outlook.get("projected_full_year_commission", Decimal("0"))
                ),
            ),
            ("Фактовых месяцев", outlook.get("ytd_months_count", 0)),
            ("Прогнозных месяцев", outlook.get("forecast_months_count", 0)),
        ]

        start_data_row = current_row
        for label, value in metrics:
            worksheet.cell(row=current_row, column=1, value=label)
            worksheet.cell(row=current_row, column=2, value=value)
            current_row += 1

        if current_row > start_data_row:
            self._apply_data_style(worksheet, start_data_row, current_row - 1, 2)

        current_row += 2

        headers = [
            "Месяц",
            "Режим",
            "Премия",
            "Комиссия",
        ]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=current_row, column=col, value=header)
        self._apply_header_style(worksheet, current_row, len(headers))
        current_row += 1

        monthly_breakdown = outlook.get("monthly_breakdown", [])
        start_data_row = current_row
        for month in monthly_breakdown:
            mode = "Факт" if month.get("mode") == "actual" else "Прогноз"
            worksheet.cell(row=current_row, column=1, value=month.get("month_name", ""))
            worksheet.cell(row=current_row, column=2, value=mode)
            worksheet.cell(
                row=current_row,
                column=3,
                value=self._format_value(month.get("premium_value", 0)),
            )
            worksheet.cell(
                row=current_row,
                column=4,
                value=self._format_value(month.get("commission_value", 0)),
            )
            current_row += 1

        if current_row > start_data_row:
            self._apply_data_style(
                worksheet, start_data_row, current_row - 1, len(headers)
            )

        self._auto_adjust_columns(worksheet)

    def _create_forecast_sheet(
        self,
        worksheet,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """Create monthly forecast sheet."""
        # Add title and metadata
        worksheet.cell(row=1, column=1, value="Monthly Financial Forecasts")
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:F1")

        worksheet.cell(
            row=2,
            column=1,
            value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        current_row = 3
        if applied_filters:
            worksheet.cell(row=current_row, column=1, value="Applied Filters:")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for filter_name, filter_value in applied_filters.items():
                if filter_value:
                    worksheet.cell(
                        row=current_row,
                        column=1,
                        value=f"  {filter_name}: {filter_value}",
                    )
                    current_row += 1
            current_row += 1

        # Headers
        headers = [
            "Month",
            "Forecasted Premium",
            "Forecasted Commission",
            "Actual Premium",
            "Actual Commission",
            "Confidence Level (%)",
        ]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=current_row, column=col, value=header)

        self._apply_header_style(worksheet, current_row, len(headers))
        current_row += 1

        # Data rows
        monthly_forecasts = analytics.get("monthly_premium_forecast", [])
        start_data_row = current_row

        for forecast in monthly_forecasts:
            worksheet.cell(
                row=current_row,
                column=1,
                value=self._format_value(forecast.get("month")),
            )
            worksheet.cell(
                row=current_row,
                column=2,
                value=self._format_value(forecast.get("forecasted_premium", 0)),
            )
            worksheet.cell(
                row=current_row,
                column=3,
                value=self._format_value(forecast.get("forecasted_commission", 0)),
            )
            worksheet.cell(
                row=current_row,
                column=4,
                value=self._format_value(forecast.get("actual_premium", 0))
                if forecast.get("actual_premium")
                else "N/A",
            )
            worksheet.cell(
                row=current_row,
                column=5,
                value=self._format_value(forecast.get("actual_commission", 0))
                if forecast.get("actual_commission")
                else "N/A",
            )
            worksheet.cell(
                row=current_row,
                column=6,
                value=self._format_value(forecast.get("confidence_level", 0)),
            )
            current_row += 1

        # Apply styling
        if current_row > start_data_row:
            self._apply_data_style(
                worksheet, start_data_row, current_row - 1, len(headers)
            )

        self._auto_adjust_columns(worksheet)

    def _create_payment_analysis_sheet(
        self,
        worksheet,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """Create payment status analysis sheet."""
        # Add title and metadata
        worksheet.cell(row=1, column=1, value="Платежный контур")
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:C1")

        worksheet.cell(
            row=2,
            column=1,
            value=f"Дата выгрузки: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        current_row = 3
        if applied_filters:
            worksheet.cell(row=current_row, column=1, value="Примененные фильтры:")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for filter_name, filter_value in applied_filters.items():
                if filter_value:
                    worksheet.cell(
                        row=current_row,
                        column=1,
                        value=f"  {filter_name}: {filter_value}",
                    )
                    current_row += 1
            current_row += 1

        # Payment status summary
        payment_analysis = analytics.get("payment_status_analysis", {})

        worksheet.cell(row=current_row, column=1, value="Статус платежа")
        worksheet.cell(row=current_row, column=2, value="Количество")
        worksheet.cell(row=current_row, column=3, value="Сумма")
        self._apply_header_style(worksheet, current_row, 3)
        current_row += 1

        status_data = [
            (
                "Всего платежей",
                payment_analysis.get("total_payments", 0),
                payment_analysis.get("paid_amount", 0)
                + payment_analysis.get("pending_amount", 0)
                + payment_analysis.get("overdue_amount", 0),
            ),
            (
                "Оплачено",
                payment_analysis.get("paid_payments", 0),
                payment_analysis.get("paid_amount", 0),
            ),
            (
                "Ожидает оплаты",
                payment_analysis.get("pending_payments", 0),
                payment_analysis.get("pending_amount", 0),
            ),
            (
                "Просрочено",
                payment_analysis.get("overdue_payments", 0),
                payment_analysis.get("overdue_amount", 0),
            ),
        ]

        start_data_row = current_row
        for status, count, amount in status_data:
            worksheet.cell(row=current_row, column=1, value=status)
            worksheet.cell(row=current_row, column=2, value=self._format_value(count))
            worksheet.cell(row=current_row, column=3, value=self._format_value(amount))
            current_row += 1

        # Apply styling
        self._apply_data_style(worksheet, start_data_row, current_row - 1, 3)

        # Add payment discipline rate
        current_row += 1
        worksheet.cell(row=current_row, column=1, value="Платежная дисциплина (%)")
        worksheet.cell(
            row=current_row,
            column=2,
            value=self._format_value(
                payment_analysis.get("payment_discipline_rate", 0)
            ),
        )
        worksheet.cell(row=current_row, column=1).font = Font(bold=True)

        self._auto_adjust_columns(worksheet)

    def _create_overdue_analysis_sheet(
        self,
        worksheet,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """Create overdue analysis sheet."""
        # Add title and metadata
        worksheet.cell(row=1, column=1, value="Overdue Payments Analysis")
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:B1")

        worksheet.cell(
            row=2,
            column=1,
            value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        current_row = 3
        if applied_filters:
            worksheet.cell(row=current_row, column=1, value="Applied Filters:")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for filter_name, filter_value in applied_filters.items():
                if filter_value:
                    worksheet.cell(
                        row=current_row,
                        column=1,
                        value=f"  {filter_name}: {filter_value}",
                    )
                    current_row += 1
            current_row += 1

        overdue_analysis = analytics.get("overdue_payments_analysis", {})

        # Overdue by days
        worksheet.cell(row=current_row, column=1, value="Overdue by Days")
        worksheet.cell(row=current_row, column=2, value="Amount")
        self._apply_header_style(worksheet, current_row, 2)
        current_row += 1

        overdue_by_days = overdue_analysis.get("overdue_by_days", {})
        start_data_row = current_row
        for days_range, amount in overdue_by_days.items():
            worksheet.cell(row=current_row, column=1, value=days_range)
            worksheet.cell(row=current_row, column=2, value=self._format_value(amount))
            current_row += 1

        if current_row > start_data_row:
            self._apply_data_style(worksheet, start_data_row, current_row - 1, 2)

        # Add summary statistics
        current_row += 1
        worksheet.cell(row=current_row, column=1, value="Total Overdue Amount")
        worksheet.cell(
            row=current_row,
            column=2,
            value=self._format_value(overdue_analysis.get("total_overdue_amount", 0)),
        )
        worksheet.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        worksheet.cell(row=current_row, column=1, value="Average Overdue Days")
        worksheet.cell(
            row=current_row,
            column=2,
            value=self._format_value(overdue_analysis.get("average_overdue_days", 0)),
        )
        worksheet.cell(row=current_row, column=1).font = Font(bold=True)

        self._auto_adjust_columns(worksheet)

    def export_time_series_analytics(
        self,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ) -> HttpResponse:
        """
        Export time series analytics to Excel.

        Args:
            analytics: Time series analytics data
            applied_filters: Applied filters information

        Returns:
            HttpResponse with Excel file
        """
        try:
            workbook = Workbook()

            # Sheet 1: Time Series Data
            ws1 = workbook.active
            ws1.title = "Time Series Data"
            self._create_time_series_sheet(ws1, analytics, applied_filters)

            # Sheet 2: Seasonal Patterns
            ws2 = workbook.create_sheet(title="Seasonal Patterns")
            self._create_seasonal_patterns_sheet(ws2, analytics, applied_filters)

            # Save to BytesIO
            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response[
                "Content-Disposition"
            ] = f'attachment; filename="time_series_analytics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

            return response

        except Exception as e:
            logger.error(f"Error exporting time series analytics: {e}")
            raise

    def _create_time_series_sheet(
        self,
        worksheet,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """Create time series data sheet."""
        # Add title and metadata
        worksheet.cell(row=1, column=1, value="Time Series Analytics")
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:D1")

        worksheet.cell(
            row=2,
            column=1,
            value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        current_row = 3
        if applied_filters:
            worksheet.cell(row=current_row, column=1, value="Applied Filters:")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for filter_name, filter_value in applied_filters.items():
                if filter_value:
                    worksheet.cell(
                        row=current_row,
                        column=1,
                        value=f"  {filter_name}: {filter_value}",
                    )
                    current_row += 1
            current_row += 1

        # Headers
        headers = ["Date", "Policy Count", "Premium Volume", "Commission Revenue"]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=current_row, column=col, value=header)

        self._apply_header_style(worksheet, current_row, len(headers))
        current_row += 1

        # Combine all time series data by date
        policy_dynamics = {
            item["date"]: item["value"]
            for item in analytics.get("policy_count_dynamics", [])
        }
        premium_dynamics = {
            item["date"]: item["value"]
            for item in analytics.get("premium_volume_dynamics", [])
        }
        commission_dynamics = {
            item["date"]: item["value"]
            for item in analytics.get("commission_revenue_dynamics", [])
        }

        # Get all unique dates and sort them
        all_dates = sorted(
            set(
                list(policy_dynamics.keys())
                + list(premium_dynamics.keys())
                + list(commission_dynamics.keys())
            )
        )

        start_data_row = current_row
        for date_val in all_dates:
            worksheet.cell(
                row=current_row, column=1, value=self._format_value(date_val)
            )
            worksheet.cell(
                row=current_row,
                column=2,
                value=self._format_value(policy_dynamics.get(date_val, 0)),
            )
            worksheet.cell(
                row=current_row,
                column=3,
                value=self._format_value(premium_dynamics.get(date_val, 0)),
            )
            worksheet.cell(
                row=current_row,
                column=4,
                value=self._format_value(commission_dynamics.get(date_val, 0)),
            )
            current_row += 1

        # Apply styling
        if current_row > start_data_row:
            self._apply_data_style(
                worksheet, start_data_row, current_row - 1, len(headers)
            )

        self._auto_adjust_columns(worksheet)

    def _create_seasonal_patterns_sheet(
        self,
        worksheet,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """Create seasonal patterns sheet."""
        # Add title and metadata
        worksheet.cell(row=1, column=1, value="Seasonal Patterns Analysis")
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:C1")

        worksheet.cell(
            row=2,
            column=1,
            value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        current_row = 3
        if applied_filters:
            worksheet.cell(row=current_row, column=1, value="Applied Filters:")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for filter_name, filter_value in applied_filters.items():
                if filter_value:
                    worksheet.cell(
                        row=current_row,
                        column=1,
                        value=f"  {filter_name}: {filter_value}",
                    )
                    current_row += 1
            current_row += 1

        seasonal_patterns = analytics.get("seasonal_patterns", {})

        # Monthly averages
        worksheet.cell(row=current_row, column=1, value="Month")
        worksheet.cell(row=current_row, column=2, value="Average Value")
        worksheet.cell(row=current_row, column=3, value="Seasonal Index")
        self._apply_header_style(worksheet, current_row, 3)
        current_row += 1

        monthly_averages = seasonal_patterns.get("monthly_averages", {})
        seasonal_indices = seasonal_patterns.get("seasonal_indices", {})

        start_data_row = current_row
        for month_num in range(1, 13):
            month_name = datetime(2023, month_num, 1).strftime("%B")
            worksheet.cell(row=current_row, column=1, value=month_name)
            worksheet.cell(
                row=current_row,
                column=2,
                value=self._format_value(monthly_averages.get(month_num, 0)),
            )
            worksheet.cell(
                row=current_row,
                column=3,
                value=self._format_value(seasonal_indices.get(month_num, 0)),
            )
            current_row += 1

        # Apply styling
        self._apply_data_style(worksheet, start_data_row, current_row - 1, 3)

        # Add summary statistics
        current_row += 1
        worksheet.cell(row=current_row, column=1, value="Seasonality Strength")
        worksheet.cell(
            row=current_row,
            column=2,
            value=self._format_value(seasonal_patterns.get("seasonality_strength", 0)),
        )
        worksheet.cell(row=current_row, column=1).font = Font(bold=True)

        self._auto_adjust_columns(worksheet)

    def export_financial_history(
        self,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ) -> HttpResponse:
        """
        Export financial history analytics to Excel format.

        Args:
            analytics: Financial history analytics data
            applied_filters: Applied filters information

        Returns:
            HttpResponse with Excel file
        """
        try:
            workbook = Workbook()

            # Remove default sheet
            workbook.remove(workbook.active)

            # Create sheets
            monthly_sheet = workbook.create_sheet("Monthly History")
            summary_sheet = workbook.create_sheet("Summary")
            highlights_sheet = workbook.create_sheet("Highlights")
            problems_sheet = workbook.create_sheet("Problems")

            # Fill sheets with data
            self._create_monthly_history_sheet(
                monthly_sheet, analytics, applied_filters
            )
            self._create_history_summary_sheet(
                summary_sheet, analytics, applied_filters
            )
            self._create_highlights_sheet(highlights_sheet, analytics, applied_filters)
            self._create_problems_sheet(problems_sheet, analytics, applied_filters)

            # Save to BytesIO
            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response[
                "Content-Disposition"
            ] = f'attachment; filename="financial_history_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'

            return response

        except Exception as e:
            logger.error(f"Error exporting financial history: {e}")
            raise

    def _create_monthly_history_sheet(
        self,
        worksheet,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """Create monthly history sheet."""
        # Add title and metadata
        worksheet.cell(row=1, column=1, value="Financial History - Monthly Data")
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:J1")

        worksheet.cell(
            row=2,
            column=1,
            value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        current_row = 3
        if applied_filters:
            worksheet.cell(row=current_row, column=1, value="Applied Filters:")
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for filter_name, filter_value in applied_filters.items():
                if filter_value:
                    worksheet.cell(
                        row=current_row,
                        column=1,
                        value=f"  {filter_name}: {filter_value}",
                    )
                    current_row += 1
            current_row += 1

        # Headers
        headers = [
            "Month",
            "Year",
            "Received Premium",
            "Received Commission",
            "Paid Payments",
            "Total Payments",
            "Payment Discipline %",
            "Overdue Payments",
            "Policies Created",
            "Average Paid Payment",
        ]

        for col, header in enumerate(headers, 1):
            worksheet.cell(row=current_row, column=col, value=header)

        self._apply_header_style(worksheet, current_row, len(headers))
        current_row += 1

        # Data
        monthly_history = analytics.get("monthly_history", [])
        start_data_row = current_row

        for month_data in monthly_history:
            paid_payments_count = month_data.get("paid_payments", 0)
            average_paid_payment = Decimal("0")
            if paid_payments_count:
                average_paid_payment = month_data.get(
                    "actual_premium", Decimal("0")
                ) / Decimal(str(paid_payments_count))

            worksheet.cell(
                row=current_row, column=1, value=month_data.get("month_name", "")
            )
            worksheet.cell(row=current_row, column=2, value=month_data.get("year", ""))
            worksheet.cell(
                row=current_row,
                column=3,
                value=self._format_value(month_data.get("actual_premium", 0)),
            )
            worksheet.cell(
                row=current_row,
                column=4,
                value=self._format_value(month_data.get("actual_commission", 0)),
            )
            worksheet.cell(
                row=current_row,
                column=5,
                value=month_data.get("paid_payments", 0),
            )
            worksheet.cell(
                row=current_row,
                column=6,
                value=month_data.get("total_payments", 0),
            )
            worksheet.cell(
                row=current_row,
                column=7,
                value=self._format_value(month_data.get("payment_discipline", 0)),
            )
            worksheet.cell(
                row=current_row,
                column=8,
                value=month_data.get("overdue_payments", 0),
            )
            worksheet.cell(
                row=current_row, column=9, value=month_data.get("policies_created", 0)
            )
            worksheet.cell(
                row=current_row,
                column=10,
                value=self._format_value(average_paid_payment),
            )
            current_row += 1

        # Apply styling
        if monthly_history:
            self._apply_data_style(
                worksheet, start_data_row, current_row - 1, len(headers)
            )

        self._auto_adjust_columns(worksheet)

    def _create_history_summary_sheet(
        self,
        worksheet,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """Create history summary sheet."""
        # Add title and metadata
        worksheet.cell(row=1, column=1, value="Financial History - Summary")
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:C1")

        worksheet.cell(
            row=2,
            column=1,
            value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        current_row = 4

        # Summary metrics
        summary_metrics = analytics.get("summary_metrics", {})
        performance_trends = analytics.get("performance_trends", {})
        problem_analysis = analytics.get("problem_analysis", {})

        # Summary section
        worksheet.cell(row=current_row, column=1, value="Summary Metrics")
        worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 2

        summary_data = [
            (
                "Total Actual Premium",
                self._format_value(summary_metrics.get("total_actual_premium", 0)),
            ),
            (
                "Total Actual Commission",
                self._format_value(summary_metrics.get("total_actual_commission", 0)),
            ),
            (
                "Paid Payments",
                summary_metrics.get("total_paid_payments", 0),
            ),
            (
                "Total Payments",
                summary_metrics.get("total_payments_count", 0),
            ),
            (
                "Payment Realization Rate",
                f"{self._format_value(summary_metrics.get('payment_realization_rate', 0))}%",
            ),
            (
                "Average Paid Payment",
                self._format_value(summary_metrics.get("average_paid_payment", 0)),
            ),
            (
                "Overdue Payments",
                summary_metrics.get("total_overdue_payments", 0),
            ),
            (
                "Total Overdue Amount",
                self._format_value(problem_analysis.get("total_overdue_amount", 0)),
            ),
            (
                "Total Policies Created",
                summary_metrics.get("total_policies_created", 0),
            ),
            ("Months Analyzed", summary_metrics.get("months_analyzed", 0)),
            (
                "Average Monthly Premium",
                self._format_value(summary_metrics.get("avg_monthly_premium", 0)),
            ),
            (
                "Average Monthly Commission",
                self._format_value(summary_metrics.get("avg_monthly_commission", 0)),
            ),
        ]

        for label, value in summary_data:
            worksheet.cell(row=current_row, column=1, value=label)
            worksheet.cell(row=current_row, column=2, value=value)
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

        current_row += 1

        # Performance Trends section
        worksheet.cell(row=current_row, column=1, value="Performance Trends")
        worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 2

        trends_data = [
            ("Premium Trend", performance_trends.get("premium_trend", "N/A")),
            ("Commission Trend", performance_trends.get("commission_trend", "N/A")),
            ("Policy Trend", performance_trends.get("policy_trend", "N/A")),
            (
                "Growth Rate",
                f"{self._format_value(performance_trends.get('growth_rate', 0))}%",
            ),
            (
                "Volatility",
                f"{self._format_value(performance_trends.get('volatility', 0))}%",
            ),
        ]

        for label, value in trends_data:
            worksheet.cell(row=current_row, column=1, value=label)
            worksheet.cell(row=current_row, column=2, value=value)
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

        self._auto_adjust_columns(worksheet)

    def _create_highlights_sheet(
        self,
        worksheet,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """Create highlights sheet."""
        # Add title and metadata
        worksheet.cell(row=1, column=1, value="Financial History - Monthly Highlights")
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:F1")

        current_row = 3

        # Headers
        headers = [
            "Month",
            "Top Client",
            "Top Client Premium",
            "Top Insurance Type",
            "Insurance Count",
            "Largest Policy Sum",
        ]

        for col, header in enumerate(headers, 1):
            worksheet.cell(row=current_row, column=col, value=header)

        self._apply_header_style(worksheet, current_row, len(headers))
        current_row += 1

        # Data
        monthly_highlights = analytics.get("monthly_highlights", [])
        start_data_row = current_row

        for highlight in monthly_highlights:
            worksheet.cell(
                row=current_row,
                column=1,
                value=f"{highlight.get('month_name', '')} {highlight.get('month', {}).get('year', '')}",
            )
            worksheet.cell(
                row=current_row, column=2, value=highlight.get("top_client", "")
            )
            worksheet.cell(
                row=current_row,
                column=3,
                value=self._format_value(highlight.get("top_client_premium", 0)),
            )
            worksheet.cell(
                row=current_row, column=4, value=highlight.get("top_insurance_type", "")
            )
            worksheet.cell(
                row=current_row, column=5, value=highlight.get("top_insurance_count", 0)
            )
            worksheet.cell(
                row=current_row,
                column=6,
                value=self._format_value(highlight.get("largest_policy_sum", 0)),
            )
            current_row += 1

        # Apply styling
        if monthly_highlights:
            self._apply_data_style(
                worksheet, start_data_row, current_row - 1, len(headers)
            )

        self._auto_adjust_columns(worksheet)

    def _create_problems_sheet(
        self,
        worksheet,
        analytics: Dict[str, Any],
        applied_filters: Optional[Dict[str, Any]] = None,
    ):
        """Create problems analysis sheet."""
        # Add title and metadata
        worksheet.cell(row=1, column=1, value="Financial History - Problem Analysis")
        worksheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        worksheet.merge_cells("A1:D1")

        current_row = 3

        problem_analysis = analytics.get("problem_analysis", {})

        # Summary
        worksheet.cell(row=current_row, column=1, value="Problem Summary")
        worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 2

        summary_data = [
            (
                "Total Overdue Amount",
                self._format_value(problem_analysis.get("total_overdue_amount", 0)),
            ),
            ("Total Overdue Count", problem_analysis.get("total_overdue_count", 0)),
            (
                "Average Monthly Overdue",
                self._format_value(problem_analysis.get("average_monthly_overdue", 0)),
            ),
        ]

        for label, value in summary_data:
            worksheet.cell(row=current_row, column=1, value=label)
            worksheet.cell(row=current_row, column=2, value=value)
            worksheet.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

        current_row += 2

        # Problematic clients
        worksheet.cell(row=current_row, column=1, value="Problematic Clients")
        worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1

        # Headers for problematic clients
        headers = ["Client Name", "Total Overdue", "Overdue Count"]
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=current_row, column=col, value=header)

        self._apply_header_style(worksheet, current_row, len(headers))
        current_row += 1

        # Data for problematic clients
        problematic_clients = problem_analysis.get("problematic_clients", [])
        start_data_row = current_row

        for client in problematic_clients:
            worksheet.cell(
                row=current_row,
                column=1,
                value=client.get("policy__client__client_name", ""),
            )
            worksheet.cell(
                row=current_row,
                column=2,
                value=self._format_value(client.get("total_overdue", 0)),
            )
            worksheet.cell(
                row=current_row, column=3, value=client.get("overdue_count", 0)
            )
            current_row += 1

        # Apply styling
        if problematic_clients:
            self._apply_data_style(
                worksheet, start_data_row, current_row - 1, len(headers)
            )

        self._auto_adjust_columns(worksheet)
