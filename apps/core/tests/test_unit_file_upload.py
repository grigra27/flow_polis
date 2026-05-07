"""
Unit tests for file upload validation.
Validates: Requirements 8.1, 8.2, 8.4
"""
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from apps.core.file_validators import FileUploadValidator


class TestFileUploadValidator:
    """Unit tests for FileUploadValidator class."""

    def create_jpeg_file(self, filename="test.jpg", size=1024):
        """Helper to create a valid JPEG file."""
        # JPEG magic bytes
        content = (
            b"\xFF\xD8\xFF\xE0\x00\x10JFIF\x00\x01\x01\x00\x00\x01\x00\x01\x00\x00"
        )
        content += b"\x00" * (size - len(content))
        return SimpleUploadedFile(filename, content, content_type="image/jpeg")

    def create_png_file(self, filename="test.png", size=1024):
        """Helper to create a valid PNG file."""
        # PNG magic bytes
        content = b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00"
        content += b"\x00" * (size - len(content))
        return SimpleUploadedFile(filename, content, content_type="image/png")

    def create_pdf_file(self, filename="test.pdf", size=1024):
        """Helper to create a valid PDF file."""
        # PDF magic bytes
        content = b"%PDF-1.4\n%\xE2\xE3\xCF\xD3\n"
        content += b"\x00" * (size - len(content))
        return SimpleUploadedFile(filename, content, content_type="application/pdf")

    def create_xlsx_file(self, filename="test.xlsx", size=None):
        """
        Helper to create a real XLSX file (PLAN 9 (a)).

        Раньше fake `PK\\x03\\x04` + нули прошёл бы валидатор — теперь
        мы реально открываем файл через openpyxl, поэтому нужен валидный
        xlsx. `size` параметр игнорируется (формат фиксированный).
        """
        from io import BytesIO

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "test"
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return SimpleUploadedFile(
            filename,
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def create_xlsx_with_uppercase_shared_strings(self, filename="test.xlsx"):
        """Helper for XLSX files exported with xl/SharedStrings.xml."""
        from io import BytesIO
        from zipfile import ZIP_DEFLATED, ZipFile

        buf = BytesIO()
        with ZipFile(buf, "w", ZIP_DEFLATED) as zf:
            zf.writestr(
                "[Content_Types].xml",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/SharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>
</Types>""",
            )
            zf.writestr(
                "_rels/.rels",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>""",
            )
            zf.writestr(
                "xl/workbook.xml",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="TDSheet" sheetId="1" r:id="rId1"/></sheets>
</workbook>""",
            )
            zf.writestr(
                "xl/_rels/workbook.xml.rels",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>""",
            )
            zf.writestr(
                "xl/worksheets/sheet1.xml",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetData>
    <row r="1"><c r="A1" t="s"><v>0</v></c></row>
  </sheetData>
</worksheet>""",
            )
            zf.writestr(
                "xl/SharedStrings.xml",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1">
  <si><t>test</t></si>
</sst>""",
            )

        buf.seek(0)
        return SimpleUploadedFile(
            filename,
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def create_xls_file(self, filename="test.xls"):
        """Helper to create a real old-style XLS file."""
        from io import BytesIO

        import xlwt

        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("TDSheet")
        sheet.write(0, 0, "test")
        buf = BytesIO()
        workbook.save(buf)
        buf.seek(0)
        return SimpleUploadedFile(
            filename,
            buf.read(),
            content_type="application/vnd.ms-excel",
        )

    def create_fake_zip_with_xlsx_extension(self, filename="evil.xlsx"):
        """
        ZIP-архив, который НЕ является xlsx — просто валидный ZIP с произвольными
        файлами. Раньше такой проходил валидацию (filetype определял MIME=zip,
        который был в whitelist для xlsx). PLAN 9 (a): теперь должен быть отклонён
        потому что openpyxl не сможет его открыть как Workbook.
        """
        import zipfile
        from io import BytesIO

        buf = BytesIO()
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("malicious.txt", "not an xlsx, just a zip")
        buf.seek(0)
        return SimpleUploadedFile(
            filename,
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # Test: Загрузка разрешенного типа файла
    def test_allowed_file_type_jpg(self):
        """
        Test that JPG files are accepted.
        Validates: Requirement 8.1
        """
        uploaded_file = self.create_jpeg_file("logo.jpg")
        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert is_valid, f"JPG file should be accepted, error: {error}"
        assert error is None
        assert safe_filename is not None
        assert safe_filename.endswith(".jpg")

    def test_allowed_file_type_png(self):
        """
        Test that PNG files are accepted.
        Validates: Requirement 8.1
        """
        uploaded_file = self.create_png_file("icon.png")
        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert is_valid, f"PNG file should be accepted, error: {error}"
        assert error is None
        assert safe_filename is not None
        assert safe_filename.endswith(".png")

    def test_allowed_file_type_pdf(self):
        """
        Test that PDF files are accepted.
        Validates: Requirement 8.1
        """
        uploaded_file = self.create_pdf_file("document.pdf")
        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert is_valid, f"PDF file should be accepted, error: {error}"
        assert error is None
        assert safe_filename is not None
        assert safe_filename.endswith(".pdf")

    def test_allowed_file_type_xlsx(self):
        """
        Test that XLSX files are accepted.
        Validates: Requirement 8.1
        """
        uploaded_file = self.create_xlsx_file("spreadsheet.xlsx")
        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert is_valid, f"XLSX file should be accepted, error: {error}"
        assert error is None
        assert safe_filename is not None
        assert safe_filename.endswith(".xlsx")

    def test_allowed_xlsx_with_uppercase_shared_strings(self):
        uploaded_file = self.create_xlsx_with_uppercase_shared_strings(
            "spreadsheet.xlsx"
        )
        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert is_valid, f"XLSX file should be accepted, error: {error}"
        assert error is None
        assert safe_filename is not None
        assert safe_filename.endswith(".xlsx")

    def test_allowed_file_type_xls(self):
        """
        Test that old-style XLS files are accepted.
        """
        uploaded_file = self.create_xls_file("accept.xls")
        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert is_valid, f"XLS file should be accepted, error: {error}"
        assert error is None
        assert safe_filename is not None
        assert safe_filename.endswith(".xls")

    # Test: Отклонение запрещенного типа
    def test_disallowed_file_type_exe(self):
        """
        Test that EXE files are rejected.
        Validates: Requirement 8.1
        """
        content = b"MZ\x90\x00"  # EXE magic bytes
        uploaded_file = SimpleUploadedFile(
            "malware.exe", content, content_type="application/x-msdownload"
        )

        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert not is_valid, "EXE file should be rejected"
        assert error is not None
        assert "Недопустимый тип файла" in error or "Разрешены" in error
        assert safe_filename is None

    def test_disallowed_file_type_sh(self):
        """
        Test that shell script files are rejected.
        Validates: Requirement 8.1
        """
        content = b'#!/bin/bash\necho "test"'
        uploaded_file = SimpleUploadedFile(
            "script.sh", content, content_type="application/x-sh"
        )

        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert not is_valid, "Shell script file should be rejected"
        assert error is not None
        assert "Недопустимый тип файла" in error or "Разрешены" in error
        assert safe_filename is None

    def test_disallowed_file_type_php(self):
        """
        Test that PHP files are rejected.
        Validates: Requirement 8.1
        """
        content = b'<?php echo "test"; ?>'
        uploaded_file = SimpleUploadedFile(
            "webshell.php", content, content_type="application/x-php"
        )

        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert not is_valid, "PHP file should be rejected"
        assert error is not None
        assert "Недопустимый тип файла" in error or "Разрешены" in error
        assert safe_filename is None

    def test_disallowed_file_type_html(self):
        """
        Test that HTML files are rejected.
        Validates: Requirement 8.1
        """
        content = b"<html><body>test</body></html>"
        uploaded_file = SimpleUploadedFile(
            "page.html", content, content_type="text/html"
        )

        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert not is_valid, "HTML file should be rejected"
        assert error is not None
        assert "Недопустимый тип файла" in error or "Разрешены" in error
        assert safe_filename is None

    # Test: Отклонение файла с несоответствующим MIME
    def test_mime_type_mismatch_jpg_with_png_content(self):
        """
        Test that a file with .jpg extension but PNG content is rejected.
        Validates: Requirement 8.2
        """
        # PNG content with JPG extension
        content = b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00"
        content += b"\x00" * 100
        uploaded_file = SimpleUploadedFile(
            "fake.jpg", content, content_type="image/jpeg"
        )

        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert not is_valid, "File with MIME mismatch should be rejected"
        assert error is not None
        assert "не соответствует" in error
        assert safe_filename is None

    def test_mime_type_mismatch_png_with_pdf_content(self):
        """
        Test that a file with .png extension but PDF content is rejected.
        Validates: Requirement 8.2
        """
        # PDF content with PNG extension
        content = b"%PDF-1.4\n%\xE2\xE3\xCF\xD3\n"
        content += b"\x00" * 100
        uploaded_file = SimpleUploadedFile(
            "fake.png", content, content_type="image/png"
        )

        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert not is_valid, "File with MIME mismatch should be rejected"
        assert error is not None
        assert "не соответствует" in error
        assert safe_filename is None

    def test_mime_type_mismatch_pdf_with_jpg_content(self):
        """
        Test that a file with .pdf extension but JPEG content is rejected.
        Validates: Requirement 8.2
        """
        # JPEG content with PDF extension
        content = (
            b"\xFF\xD8\xFF\xE0\x00\x10JFIF\x00\x01\x01\x00\x00\x01\x00\x01\x00\x00"
        )
        content += b"\x00" * 100
        uploaded_file = SimpleUploadedFile(
            "fake.pdf", content, content_type="application/pdf"
        )

        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert not is_valid, "File with MIME mismatch should be rejected"
        assert error is not None
        assert "не соответствует" in error
        assert safe_filename is None

    # PLAN 9 (a): фейковый ZIP под видом xlsx должен быть отклонён.
    def test_xlsx_rejects_arbitrary_zip(self):
        """
        Раньше любой ZIP-архив с расширением .xlsx проходил валидацию
        (filetype определял MIME=application/zip, который был в whitelist
        для xlsx). Теперь openpyxl реально открывает файл — если внутри
        не xlsx, отклоняем.
        """
        uploaded_file = self.create_fake_zip_with_xlsx_extension("evil.xlsx")
        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert not is_valid, "Произвольный ZIP под видом .xlsx должен быть отклонён"
        assert error is not None
        assert "Excel" in error or "xlsx" in error.lower()
        assert safe_filename is None

    # Test: Отклонение файла превышающего лимит
    def test_file_size_exceeds_limit(self):
        """
        Test that files exceeding 10MB are rejected.
        Validates: Requirement 8.4
        """
        # Create a file larger than 10MB
        size = 11 * 1024 * 1024  # 11MB
        uploaded_file = self.create_jpeg_file("large.jpg", size=size)

        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert not is_valid, "File exceeding size limit should be rejected"
        assert error is not None
        assert "превышает" in error or "10" in error
        assert safe_filename is None

    def test_file_size_at_limit(self):
        """
        Test that files at exactly 10MB are accepted.
        Validates: Requirement 8.4
        """
        # Create a file at exactly 10MB
        size = 10 * 1024 * 1024  # 10MB
        uploaded_file = self.create_jpeg_file("at_limit.jpg", size=size)

        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert is_valid, f"File at size limit should be accepted, error: {error}"
        assert error is None
        assert safe_filename is not None

    def test_empty_file_rejected(self):
        """
        Test that empty files are rejected.
        Validates: Requirement 8.4
        """
        uploaded_file = SimpleUploadedFile("empty.jpg", b"", content_type="image/jpeg")

        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert not is_valid, "Empty file should be rejected"
        assert error is not None
        assert "пустым" in error
        assert safe_filename is None

    # Additional tests
    def test_file_without_extension(self):
        """
        Test that files without extension are rejected.
        """
        content = (
            b"\xFF\xD8\xFF\xE0\x00\x10JFIF\x00\x01\x01\x00\x00\x01\x00\x01\x00\x00"
        )
        uploaded_file = SimpleUploadedFile(
            "noextension", content, content_type="image/jpeg"
        )

        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert not is_valid, "File without extension should be rejected"
        assert error is not None

    def test_safe_filename_generation(self):
        """
        Test that safe filenames are generated correctly.
        """
        uploaded_file = self.create_jpeg_file("original_name.jpg")
        is_valid, error, safe_filename = FileUploadValidator.validate_file(
            uploaded_file
        )

        assert is_valid
        assert safe_filename != "original_name.jpg"
        assert safe_filename.endswith(".jpg")
        # Check that filename is UUID-like (32 hex chars + extension)
        base_name = safe_filename.rsplit(".", 1)[0]
        assert len(base_name) == 32
        assert all(c in "0123456789abcdef" for c in base_name.lower())
