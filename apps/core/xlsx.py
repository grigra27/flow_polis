from io import BytesIO
from zipfile import ZIP_DEFLATED, BadZipFile, ZipFile


CANONICAL_SHARED_STRINGS = "xl/sharedStrings.xml"


def load_workbook_compat(file_obj, **kwargs):
    """
    Load XLSX files with a small compatibility fix for exporters that write
    xl/SharedStrings.xml instead of the canonical xl/sharedStrings.xml.
    """
    from openpyxl import load_workbook

    current_pos = _safe_tell(file_obj)
    try:
        _safe_seek(file_obj, 0)
        return load_workbook(file_obj, **kwargs)
    except KeyError as exc:
        if not _is_missing_shared_strings_error(exc):
            raise

        normalized = _normalize_shared_strings_name(file_obj)
        if normalized is None:
            raise

        return load_workbook(normalized, **kwargs)
    finally:
        if current_pos is not None:
            _safe_seek(file_obj, current_pos)


def _normalize_shared_strings_name(file_obj):
    current_pos = _safe_tell(file_obj)
    try:
        _safe_seek(file_obj, 0)
        content = file_obj.read()
    finally:
        if current_pos is not None:
            _safe_seek(file_obj, current_pos)

    try:
        source = ZipFile(BytesIO(content))
    except BadZipFile:
        return None

    names_by_lowercase = {
        info.filename.lower(): info.filename for info in source.infolist()
    }
    source_name = names_by_lowercase.get(CANONICAL_SHARED_STRINGS.lower())
    if not source_name or source_name == CANONICAL_SHARED_STRINGS:
        source.close()
        return None

    normalized = BytesIO()
    with source:
        with ZipFile(normalized, "w", ZIP_DEFLATED) as target:
            for info in source.infolist():
                name = (
                    CANONICAL_SHARED_STRINGS
                    if info.filename == source_name
                    else info.filename
                )
                target.writestr(name, source.read(info.filename))

    normalized.seek(0)
    return normalized


def _is_missing_shared_strings_error(exc):
    return CANONICAL_SHARED_STRINGS in str(exc)


def _safe_tell(file_obj):
    try:
        return file_obj.tell()
    except Exception:
        return None


def _safe_seek(file_obj, position):
    try:
        file_obj.seek(position)
    except Exception:
        pass
