import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.exceptions import ValidationError


class AcceptParseError(ValidationError):
    """Raised when an insurance accept file cannot be parsed."""


@dataclass
class ParsedAcceptData:
    source_filename: str = ""
    client_short: str = ""
    client_full: str = ""
    client_legal: str = ""
    client_inn: str = ""
    legal_address: str = ""
    mail_address: str = ""
    contact: str = ""
    phone: str = ""
    email: str = ""
    lessee: str = ""
    dfa_number: str = ""
    dfa_date: date | None = None
    policyholder_text: str = ""
    bank: str = ""
    beneficiary_loss: str = ""
    beneficiary_damage: str = ""
    property_description: str = ""
    purchase_price: Decimal | None = None
    start_date: date | None = None
    end_date: date | None = None
    term_text: str = ""
    insurance_type_name: str = ""
    osago: str = ""
    model: str = ""
    vin_number: str = ""
    asset_type: str = ""
    vehicle_category: str = ""
    property_year: int | None = None
    pts: str = ""
    usage_year: int | None = None
    registration_place: str = ""
    plate_number: str = ""

    def to_session_dict(self):
        data = asdict(self)
        for key in ("dfa_date", "start_date", "end_date"):
            if data[key]:
                data[key] = data[key].isoformat()
        for key in ("purchase_price",):
            if data[key] is not None:
                data[key] = str(data[key])
        return data


@dataclass
class AcceptParseResult:
    data: ParsedAcceptData
    warnings: list[str] = field(default_factory=list)
    raw_fields: dict[str, str] = field(default_factory=dict)


FIELD_LABELS = {
    "client_short": "Наименование сокращенное",
    "client_full": "Наименование полное",
    "client_legal": "Наименование по учредительным документам",
    "client_inn": "ИНН",
    "legal_address": "Адрес места нахождения (юридический)",
    "mail_address": "Почтовый адрес",
    "contact": "Контактное лицо по страхованию",
    "phone": "Телефон",
    "email": "e-mail",
    "lessee": "Лизингополучатель",
    "dfa_number": "Номер ДФА",
    "dfa_date": "Дата ДФА",
    "policyholder_text": "Страхователь",
    "bank": "Банк-кредитор",
    "beneficiary_loss": "Выгодоприобретатель (риск утраты)",
    "beneficiary_damage": "Выгодоприобретатель (риск ущерба)",
    "property_description": "наименование страхуемого имущества",
    "purchase_price": "стоимость по ДКП, валюта",
    "start_date": "Дата поставки / дата начала страхования",
    "end_date": "Дата окончания срока лизинга",
    "term_text": "Срок страхования",
    "insurance_type_name": "Необходимый вид страхования",
    "osago": "ОСАГО",
    "model": "Марка, модель предмета лизинга",
    "vin_number": "VIN (или иной ID)",
    "asset_type": "Наименование, тип имущества (ТС, спецтехники, оборудования)",
    "vehicle_category": "Категория ТС",
    "property_year": "Год выпуска",
    "pts": "ПТС, дата выдачи (для ТС)",
    "usage_year": "Год начала эксплуатации (лизинга)",
    "registration_place": "Место регистрации",
    "plate_number": "Государственный регистрационный номер (для ТС)",
}

REQUIRED_FIELDS = {
    "client_inn": "ИНН",
    "dfa_number": "Номер ДФА",
    "property_description": "Описание имущества",
    "end_date": "Дата окончания срока лизинга",
    "insurance_type_name": "Вид страхования",
}


def parse_accept_file(uploaded_file) -> AcceptParseResult:
    parser = AcceptParser()
    return parser.parse(uploaded_file)


class AcceptParser:
    def parse(self, uploaded_file) -> AcceptParseResult:
        filename = getattr(uploaded_file, "name", "")
        suffix = Path(filename).suffix.lower()
        if suffix == ".xls":
            rows, sheet_warnings = self._read_xls_rows(uploaded_file)
        elif suffix == ".xlsx":
            rows, sheet_warnings = self._read_xlsx_rows(uploaded_file)
        else:
            raise AcceptParseError("Поддерживаются только файлы .xls и .xlsx")

        label_values = self._build_label_values(rows)
        raw_fields = {
            label: self._stringify(values[0]) for label, values in label_values.items()
        }

        values = {}
        for field_name, label in FIELD_LABELS.items():
            raw_value = self._first_label_value(label_values, label)
            values[field_name] = raw_value

        data = ParsedAcceptData(
            source_filename=filename,
            client_short=self._clean_text(values["client_short"]),
            client_full=self._clean_text(values["client_full"]),
            client_legal=self._clean_text(values["client_legal"]),
            client_inn=self._digits_only(values["client_inn"]),
            legal_address=self._clean_text(values["legal_address"]),
            mail_address=self._clean_text(values["mail_address"]),
            contact=self._clean_text(values["contact"]),
            phone=self._clean_text(values["phone"]),
            email=self._clean_text(values["email"]),
            lessee=self._clean_text(values["lessee"]),
            dfa_number=self._clean_text(values["dfa_number"]),
            dfa_date=self._parse_date(values["dfa_date"]),
            policyholder_text=self._clean_text(values["policyholder_text"]),
            bank=self._clean_text(values["bank"]),
            beneficiary_loss=self._clean_text(values["beneficiary_loss"]),
            beneficiary_damage=self._clean_text(values["beneficiary_damage"]),
            property_description=self._clean_text(values["property_description"]),
            purchase_price=self._parse_money(values["purchase_price"]),
            start_date=self._parse_date(values["start_date"]),
            end_date=self._parse_date(values["end_date"]),
            term_text=self._clean_text(values["term_text"]),
            insurance_type_name=self._clean_text(values["insurance_type_name"]),
            osago=self._clean_text(values["osago"]),
            model=self._clean_text(values["model"]),
            vin_number=self._clean_text(values["vin_number"]).upper(),
            asset_type=self._clean_text(values["asset_type"]),
            vehicle_category=self._clean_text(values["vehicle_category"]),
            property_year=self._parse_int(values["property_year"]),
            pts=self._clean_text(values["pts"]),
            usage_year=self._parse_int(values["usage_year"]),
            registration_place=self._clean_text(values["registration_place"]),
            plate_number=self._clean_text(values["plate_number"]),
        )

        warnings = list(sheet_warnings)
        warnings.extend(self._build_missing_field_warnings(data))
        if data.purchase_price is None:
            warnings.append("Стоимость по ДКП не распознана как число.")
        if not data.start_date:
            warnings.append("Дата начала страхования в акцепте не заполнена.")

        return AcceptParseResult(data=data, warnings=warnings, raw_fields=raw_fields)

    def _read_xls_rows(self, uploaded_file):
        import xlrd
        from xlrd.xldate import xldate_as_datetime

        current_pos = uploaded_file.tell()
        try:
            uploaded_file.seek(0)
            # formatting_info=True is required to read cell number formats
            # (needed for restoring leading zeros in numeric INN cells).
            book = xlrd.open_workbook(
                file_contents=uploaded_file.read(),
                formatting_info=True,
            )
        except Exception as exc:
            raise AcceptParseError(f"Не удалось открыть .xls файл: {exc}") from exc
        finally:
            uploaded_file.seek(current_pos)

        sheet, warnings = self._select_sheet(book.sheet_names(), book.sheet_by_index)

        rows = []
        for row_idx in range(sheet.nrows):
            key = self._xls_cell_value(book, sheet.cell(row_idx, 0), xldate_as_datetime)
            if sheet.ncols > 1:
                value_cell = sheet.cell(row_idx, 1)
                value = self._xls_cell_value(book, value_cell, xldate_as_datetime)
                value = self._restore_inn_from_xls_cell(book, key, value_cell, value)
            else:
                value = ""
            rows.append((key, value))
        return rows, warnings

    def _read_xlsx_rows(self, uploaded_file):
        from openpyxl import load_workbook

        current_pos = uploaded_file.tell()
        try:
            uploaded_file.seek(0)
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        except Exception as exc:
            raise AcceptParseError(f"Не удалось открыть .xlsx файл: {exc}") from exc
        finally:
            uploaded_file.seek(current_pos)

        sheet_name, warnings = self._select_sheet(
            workbook.sheetnames, lambda index: workbook.worksheets[index]
        )
        sheet = workbook[sheet_name] if isinstance(sheet_name, str) else sheet_name

        rows = []
        for row in sheet.iter_rows(min_row=1, max_col=2, values_only=False):
            key_cell = row[0] if len(row) > 0 else None
            value_cell = row[1] if len(row) > 1 else None
            key = key_cell.value if key_cell is not None else ""
            value = value_cell.value if value_cell is not None else ""
            value = self._restore_inn_from_xlsx_cell(key, value_cell, value)
            rows.append((key, value))
        workbook.close()
        return rows, warnings

    def _select_sheet(self, sheet_names, sheet_getter):
        warnings = []
        if not sheet_names:
            raise AcceptParseError("Excel-файл не содержит листов.")
        if "TDSheet" in sheet_names:
            index = sheet_names.index("TDSheet")
        else:
            index = 0
            warnings.append(
                "Лист TDSheet не найден, использован первый лист Excel-файла."
            )
        return sheet_getter(index), warnings

    def _xls_cell_value(self, book, cell, date_converter):
        import xlrd

        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return date_converter(cell.value, book.datemode).date()
            except Exception:
                return cell.value
        return cell.value

    def _build_label_values(self, rows):
        label_values: dict[str, list[Any]] = {}
        for key, value in rows:
            label = self._normalize_label(key)
            if not label:
                continue
            label_values.setdefault(label, []).append(value)
        return label_values

    def _first_label_value(self, label_values, label):
        values = label_values.get(self._normalize_label(label), [])
        if not values:
            return ""
        return values[0]

    def _build_missing_field_warnings(self, data):
        warnings = []
        session_data = data.to_session_dict()
        for field_name, label in REQUIRED_FIELDS.items():
            if not session_data.get(field_name):
                warnings.append(f"Обязательное поле акцепта не найдено: {label}.")
        return warnings

    def _parse_date(self, value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = self._clean_text(value)
        if not text:
            return None
        for date_format in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, date_format).date()
            except ValueError:
                continue
        return None

    def _parse_money(self, value):
        text = self._clean_text(value)
        if not text:
            return None
        match = re.search(r"\d[\d\s]*(?:[,.]\d+)?", text)
        if not match:
            return None
        normalized = match.group(0).replace(" ", "").replace(",", ".")
        try:
            return Decimal(normalized)
        except InvalidOperation:
            return None

    def _parse_int(self, value):
        if value in ("", None):
            return None
        if isinstance(value, float) and value.is_integer():
            return int(value)
        if isinstance(value, int):
            return value
        match = re.search(r"\d+", self._clean_text(value))
        return int(match.group(0)) if match else None

    def _digits_only(self, value):
        return re.sub(r"\D", "", self._clean_text(value))

    def _clean_text(self, value):
        if value in ("", None):
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return re.sub(r"\s+", " ", str(value).strip())

    def _stringify(self, value):
        if isinstance(value, date):
            return value.isoformat()
        return self._clean_text(value)

    def _normalize_label(self, value):
        return re.sub(r"\s+", " ", str(value or "").strip().lower())

    def _restore_inn_from_xls_cell(self, book, key, cell, value):
        if not self._is_inn_label(key):
            return value
        format_str = self._get_xls_number_format(book, cell)
        return self._restore_padded_numeric_value(value, format_str)

    def _restore_inn_from_xlsx_cell(self, key, cell, value):
        if not self._is_inn_label(key):
            return value
        number_format = getattr(cell, "number_format", "") if cell is not None else ""
        return self._restore_padded_numeric_value(value, number_format)

    def _is_inn_label(self, label):
        return self._normalize_label(label) == self._normalize_label(
            FIELD_LABELS["client_inn"]
        )

    def _get_xls_number_format(self, book, cell):
        xf_index = getattr(cell, "xf_index", None)
        if xf_index is None:
            return ""
        try:
            xf = book.xf_list[xf_index]
            fmt = book.format_map.get(xf.format_key)
            return fmt.format_str if fmt else ""
        except Exception:
            return ""

    def _restore_padded_numeric_value(self, value, number_format):
        width = self._extract_zero_mask_width(number_format)
        if width <= 0:
            return value
        digits = self._numeric_digits(value)
        if not digits:
            return value
        if len(digits) > width:
            return value
        return digits.zfill(width)

    def _extract_zero_mask_width(self, number_format):
        if not number_format:
            return 0
        # Use only the positive-number section of Excel format.
        section = str(number_format).split(";", 1)[0].strip()
        if re.fullmatch(r"0{2,20}", section):
            return len(section)
        return 0

    def _numeric_digits(self, value):
        if value in ("", None):
            return ""
        if isinstance(value, bool):
            return ""
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            if not value.is_integer():
                return ""
            return str(int(value))
        return ""
