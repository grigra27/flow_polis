from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse

from apps.policies.models import Policy
from apps.policies.services.accept_parser import parse_accept_file
from apps.policies.services.accept_resolver import (
    ACCEPT_IMPORT_SESSION_KEY,
    resolve_accept_data,
)


def make_accept_xls(**overrides):
    import xlwt

    values = {
        "Наименование сокращенное": "АЛТЫН ЯР ООО, ООО",
        "Наименование полное": 'ООО "АЛТЫН ЯР"',
        "Наименование по учредительным документам": 'ООО "АЛТЫН ЯР"',
        "ИНН": "1686035567",
        "Лизингополучатель": "АЛТЫН ЯР ООО",
        "Номер ДФА": "20212-ГА-КЗ",
        "Дата ДФА": "15.08.2025",
        "Страхователь": "Лизингополучатель",
        "Банк-кредитор": "Собственные средства",
        "Выгодоприобретатель (риск утраты)": "Лизингодатель",
        "Выгодоприобретатель (риск ущерба)": "Лизингодатель",
        "наименование страхуемого имущества": "специальный кран СПМ Авто 732457",
        "стоимость по ДКП, валюта": "12 750 000,00 руб",
        "Дата поставки / дата начала страхования": "26.08.2025",
        "Дата окончания срока лизинга": "20.08.2028",
        "Срок страхования": "на весь срок лизинга",
        "Необходимый вид страхования": "КАСКО",
        "ОСАГО": "да",
        "Марка, модель предмета лизинга": "специальный кран СПМ Авто 732457",
        "VIN (или иной ID)": "XDC732457S9003298",
        "Год выпуска": "2025",
        "Место регистрации": "город Казань",
    }
    values.update(overrides)

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("TDSheet")
    for row, (label, value) in enumerate(values.items()):
        sheet.write(row, 0, label)
        sheet.write(row, 1, value)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile(
        "accept.xls",
        buffer.read(),
        content_type="application/vnd.ms-excel",
    )


def make_accept_xls_with_numeric_inn():
    import xlwt

    values = {
        "Наименование сокращенное": "АЛТЫН ЯР ООО, ООО",
        "Наименование полное": 'ООО "АЛТЫН ЯР"',
        "Наименование по учредительным документам": 'ООО "АЛТЫН ЯР"',
        "ИНН": 123456789,
        "Лизингополучатель": "АЛТЫН ЯР ООО",
        "Номер ДФА": "20212-ГА-КЗ",
        "Дата ДФА": "15.08.2025",
        "Страхователь": "Лизингополучатель",
        "Банк-кредитор": "Собственные средства",
        "Выгодоприобретатель (риск утраты)": "Лизингодатель",
        "Выгодоприобретатель (риск ущерба)": "Лизингодатель",
        "наименование страхуемого имущества": "специальный кран СПМ Авто 732457",
        "стоимость по ДКП, валюта": "12 750 000,00 руб",
        "Дата поставки / дата начала страхования": "26.08.2025",
        "Дата окончания срока лизинга": "20.08.2028",
        "Срок страхования": "на весь срок лизинга",
        "Необходимый вид страхования": "КАСКО",
        "ОСАГО": "да",
        "Марка, модель предмета лизинга": "специальный кран СПМ Авто 732457",
        "VIN (или иной ID)": "XDC732457S9003298",
        "Год выпуска": "2025",
        "Место регистрации": "город Казань",
    }

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("TDSheet")
    inn_style = xlwt.easyxf(num_format_str="0000000000")

    for row, (label, value) in enumerate(values.items()):
        sheet.write(row, 0, label)
        if label == "ИНН":
            sheet.write(row, 1, value, inn_style)
        else:
            sheet.write(row, 1, value)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile(
        "accept.xls",
        buffer.read(),
        content_type="application/vnd.ms-excel",
    )


def make_accept_xlsx_with_numeric_inn():
    from openpyxl import Workbook

    values = {
        "Наименование сокращенное": "АЛТЫН ЯР ООО, ООО",
        "Наименование полное": 'ООО "АЛТЫН ЯР"',
        "Наименование по учредительным документам": 'ООО "АЛТЫН ЯР"',
        "ИНН": 123456789,
        "Лизингополучатель": "АЛТЫН ЯР ООО",
        "Номер ДФА": "20212-ГА-КЗ",
        "Дата ДФА": "15.08.2025",
        "Страхователь": "Лизингополучатель",
        "Банк-кредитор": "Собственные средства",
        "Выгодоприобретатель (риск утраты)": "Лизингодатель",
        "Выгодоприобретатель (риск ущерба)": "Лизингодатель",
        "наименование страхуемого имущества": "специальный кран СПМ Авто 732457",
        "стоимость по ДКП, валюта": "12 750 000,00 руб",
        "Дата поставки / дата начала страхования": "26.08.2025",
        "Дата окончания срока лизинга": "20.08.2028",
        "Срок страхования": "на весь срок лизинга",
        "Необходимый вид страхования": "КАСКО",
        "ОСАГО": "да",
        "Марка, модель предмета лизинга": "специальный кран СПМ Авто 732457",
        "VIN (или иной ID)": "XDC732457S9003298",
        "Год выпуска": "2025",
        "Место регистрации": "город Казань",
    }

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TDSheet"

    for row_idx, (label, value) in enumerate(values.items(), start=1):
        sheet.cell(row=row_idx, column=1, value=label)
        value_cell = sheet.cell(row=row_idx, column=2, value=value)
        if label == "ИНН":
            value_cell.number_format = "0000000000"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile(
        "accept.xlsx",
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@pytest.mark.django_db
class TestAcceptParser:
    def test_parses_core_accept_fields(self):
        result = parse_accept_file(make_accept_xls())

        assert result.data.dfa_number == "20212-ГА-КЗ"
        assert result.data.client_inn == "1686035567"
        assert result.data.client_full == 'ООО "АЛТЫН ЯР"'
        assert result.data.purchase_price == Decimal("12750000.00")
        assert result.data.start_date == date(2025, 8, 26)
        assert result.data.end_date == date(2028, 8, 20)
        assert result.data.insurance_type_name == "КАСКО"
        assert result.data.vin_number == "XDC732457S9003298"

    def test_missing_start_date_becomes_warning(self):
        result = parse_accept_file(
            make_accept_xls(**{"Дата поставки / дата начала страхования": ""})
        )

        assert result.data.start_date is None
        assert "Дата начала страхования" in " ".join(result.warnings)

    def test_preserves_leading_zero_inn_from_xls_number_format(self):
        result = parse_accept_file(make_accept_xls_with_numeric_inn())

        assert result.data.client_inn == "0123456789"

    def test_preserves_leading_zero_inn_from_xlsx_number_format(self):
        result = parse_accept_file(make_accept_xlsx_with_numeric_inn())

        assert result.data.client_inn == "0123456789"


@pytest.mark.django_db
class TestAcceptResolver:
    def test_resolves_references_and_payment_initial(
        self, client_factory, branch_factory, insurance_type_factory
    ):
        client = client_factory(
            client_name='ООО "АЛТЫН ЯР"',
            client_inn="1686035567",
        )
        branch = branch_factory(branch_name="Казань")
        insurance_type = insurance_type_factory(name="КАСКО")

        parsed = parse_accept_file(make_accept_xls())
        resolved = resolve_accept_data(parsed.data, parsed.warnings)

        assert resolved.policy_initial["client"] == client.pk
        assert resolved.policy_initial["policyholder"] == client.pk
        assert resolved.policy_initial["branch"] == branch.pk
        assert resolved.policy_initial["insurance_type"] == insurance_type.pk
        assert resolved.policy_initial["vin_number"] == "XDC732457S9003298"
        assert resolved.payment_initial["insurance_sum"] == "12750000.00"

    def test_invalid_vin_is_not_prefilled(
        self, client_factory, branch_factory, insurance_type_factory
    ):
        client_factory(client_inn="1686035567")
        branch_factory(branch_name="Казань")
        insurance_type_factory(name="КАСКО")

        parsed = parse_accept_file(
            make_accept_xls(**{"VIN (или иной ID)": "TLB935-0089"})
        )
        resolved = resolve_accept_data(parsed.data, parsed.warnings)

        assert "vin_number" not in resolved.policy_initial
        assert any("не прошел валидацию" in warning for warning in resolved.warnings)


@pytest.mark.django_db
class TestAcceptAdminFlow:
    def test_preview_does_not_create_policy_and_add_form_is_prefilled(
        self,
        client,
        admin_user,
        client_factory,
        branch_factory,
        insurance_type_factory,
    ):
        lessee = client_factory(
            client_name='ООО "АЛТЫН ЯР"',
            client_inn="1686035567",
        )
        branch_factory(branch_name="Казань")
        insurance_type_factory(name="КАСКО")

        client.force_login(admin_user)
        upload_url = reverse("admin:policies_policy_import_accept")

        response = client.post(upload_url, {"accept_file": make_accept_xls()})

        assert response.status_code == 200
        assert Policy.objects.count() == 0
        assert "Открыть форму создания полиса" in response.content.decode()

        session_payloads = client.session[ACCEPT_IMPORT_SESSION_KEY]
        token = next(iter(session_payloads))
        add_url = f"{reverse('admin:policies_policy_add')}?accept_token={token}"

        add_response = client.get(add_url)

        assert add_response.status_code == 200
        content = add_response.content.decode()
        assert 'value="20212-ГА-КЗ"' in content
        assert "XDC732457S9003298" in content
        assert "12750000.00" in content
        assert str(lessee.pk) in content
