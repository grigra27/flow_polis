from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from datetime import date
from decimal import Decimal
import logging

logger = logging.getLogger(__name__)


class BaseExporter:
    """Базовый класс для всех экспортеров"""

    def __init__(self, queryset, fields):
        self.queryset = queryset
        self.fields = fields

    def export(self):
        """Генерирует Excel файл и возвращает HttpResponse"""
        wb = Workbook()
        ws = wb.active

        # Заголовки
        headers = self.get_headers()
        ws.append(headers)

        # Данные
        for obj in self.queryset:
            row = self.get_row_data(obj)
            ws.append(row)

        # Форматирование
        self.apply_formatting(ws)

        return self.create_response(wb)

    def get_headers(self):
        """Возвращает список заголовков"""
        raise NotImplementedError("Subclasses must implement get_headers()")

    def get_row_data(self, obj):
        """Возвращает данные строки для объекта"""
        raise NotImplementedError("Subclasses must implement get_row_data()")

    def apply_formatting(self, ws):
        """Применяет форматирование к листу"""
        # Базовое форматирование можно добавить здесь
        pass

    def create_response(self, wb):
        """Создает HttpResponse с Excel файлом"""
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.get_filename()}_{timestamp}.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"
        wb.save(response)
        return response

    def get_filename(self):
        """Возвращает базовое имя файла"""
        return "report"

    def format_value(self, value):
        """Форматирует значение для Excel"""
        if isinstance(value, date):
            return value.strftime("%d.%m.%Y")
        elif isinstance(value, Decimal):
            return float(value)
        elif isinstance(value, bool):
            return "Да" if value else "Нет"
        elif value is None:
            return ""
        return value


class CustomExporter(BaseExporter):
    """Экспортер для кастомного экспорта с динамическими полями"""

    FIELD_LABELS = {
        "policies": {
            "policy_number": "Номер полиса",
            "dfa_number": "Номер ДФА",
            "client__client_name": "Клиент",
            "insurer__insurer_name": "Страховщик",
            "insurance_type__name": "Вид страхования",
            "branch__branch_name": "Филиал",
            "start_date": "Дата начала",
            "end_date": "Дата окончания",
            "property_value": "Стоимость имущества",
            "premium_total": "Общая премия",
            "franchise": "Франшиза",
            "policy_active": "Статус полиса",
            "termination_date": "Дата расторжения",
            "dfa_active": "Статус ДФА",
            "broker_participation": "Участие брокера",
        },
        "payments": {
            "policy__policy_number": "Номер полиса",
            "policy__dfa_number": "Номер ДФА",
            "policy__client__client_name": "Клиент",
            "policy__insurer__insurer_name": "Страховщик",
            "policy__branch__branch_name": "Филиал",
            "year_number": "Год",
            "installment_number": "Платеж №",
            "due_date": "Дата платежа",
            "amount": "Сумма",
            "insurance_sum": "Страховая сумма",
            "commission_rate__kv_percent": "КВ %",
            "kv_rub": "КВ руб",
            "paid_date": "Дата оплаты",
            "insurer_date": "Дата согласования СК",
        },
        "clients": {
            "client_name": "Название клиента",
            "client_inn": "ИНН",
            "client_address": "Адрес",
            "client_phone": "Телефон",
            "client_email": "Email",
        },
        "insurers": {
            "insurer_name": "Название страховщика",
            "insurer_inn": "ИНН",
            "insurer_address": "Адрес",
            "insurer_phone": "Телефон",
            "insurer_email": "Email",
        },
    }

    def __init__(self, queryset, fields, data_source):
        super().__init__(queryset, fields)
        self.data_source = data_source

    def get_headers(self):
        """Возвращает список заголовков на основе выбранных полей"""
        return [self.FIELD_LABELS[self.data_source].get(f, f) for f in self.fields]

    def get_row_data(self, obj):
        """Возвращает данные строки для объекта"""
        row = []
        for field in self.fields:
            value = self.get_field_value(obj, field)
            row.append(value)
        return row

    def get_field_value(self, obj, field):
        """Получает значение поля, поддерживая вложенные поля (через __)"""
        parts = field.split("__")
        value = obj
        for part in parts:
            if value is None:
                break
            value = getattr(value, part, None)
        return self.format_value(value)

    def get_filename(self):
        """Возвращает базовое имя файла на основе источника данных"""
        return f"custom_export_{self.data_source}"


class PolicyExporter(BaseExporter):
    """Экспортер для готового экспорта по полисам"""

    def get_filename(self):
        """Возвращает базовое имя файла"""
        return "policies"

    def get_headers(self):
        """Возвращает список заголовков"""
        return [
            "Номер полиса",
            "Номер ДФА",
            "Клиент",
            "Страховщик",
            "Вид страхования",
            "Филиал",
            "Дата начала",
            "Дата окончания",
            "Общая премия",
            "Франшиза",
            "Статус полиса",
            "Дата расторжения",
            "Статус ДФА",
            "Полис подгружен",
        ]

    def get_row_data(self, policy):
        """Возвращает данные строки для полиса"""
        return [
            policy.policy_number,
            policy.dfa_number,
            policy.client.client_name,
            policy.insurer.insurer_name,
            policy.insurance_type.name,
            policy.branch.branch_name,
            self.format_value(policy.start_date),
            self.format_value(policy.end_date),
            self.format_value(policy.premium_total),
            self.format_value(policy.franchise),
            "Активен" if policy.policy_active else "Расторгнут",
            self.format_value(policy.termination_date),
            "Активен" if policy.dfa_active else "Закрыт",
            self.format_value(policy.policy_uploaded),
        ]


# PaymentExporter был удален - используйте ScheduledPaymentsExporter вместо него
# Создаем алиас для обратной совместимости с тестами
class PaymentExporter(BaseExporter):
    """Алиас для ScheduledPaymentsExporter для обратной совместимости с тестами"""

    def get_filename(self):
        """Возвращает базовое имя файла"""
        return "payments"

    def get_headers(self):
        """Возвращает список заголовков"""
        return [
            "Номер полиса",
            "Клиент",
            "Год",
            "Платеж №",
            "Дата платежа",
            "Сумма",
            "Страховая сумма",
            "КВ %",
            "КВ руб",
            "Дата оплаты",
            "Дата согласования СК",
            "Статус",
        ]

    def get_row_data(self, payment):
        """Возвращает данные строки для платежа"""
        policy = payment.policy

        # Определяем статус платежа
        if payment.paid_date:
            status = "Оплачен"
        elif not policy.policy_active and policy.termination_date:
            status = "Отменен"
        else:
            status = "Ожидается"

        # Получаем КВ %
        kv_percent = 0
        if payment.commission_rate:
            kv_percent = float(payment.commission_rate.kv_percent)

        return [
            policy.policy_number,
            policy.client.client_name,
            payment.year_number,
            payment.installment_number,
            self.format_value(payment.due_date),
            self.format_value(payment.amount),
            self.format_value(payment.insurance_sum),
            kv_percent,
            self.format_value(payment.kv_rub),
            self.format_value(payment.paid_date),
            self.format_value(payment.insurer_date),
            status,
        ]


class ScheduledPaymentsExporter(BaseExporter):
    """Экспортер для очередных взносов с форматированием как в четверговом отчете"""

    def __init__(self, queryset, fields, date_from=None, date_to=None):
        """
        Инициализация экспортера

        Args:
            queryset: QuerySet платежей
            fields: Список полей (не используется, для совместимости)
            date_from: Дата начала периода
            date_to: Дата окончания периода
        """
        super().__init__(queryset, fields)
        self.date_from = date_from
        self.date_to = date_to

    def get_filename(self):
        """Возвращает базовое имя файла"""
        return "scheduled_payments"

    def get_headers(self):
        """Возвращает список заголовков"""
        return [
            "Номер полиса",
            "Номер ДФА",
            "Лизингополучатель",
            "Страховщик",
            "Страхователь",
            "Дата начала страхования",
            "Дата оконч. страхования",
            "Объект страхования",
            "Страховая сумма",
            "Очередной взнос",
            "Статус рассрочки",
            "Этот взнос",
            "Дата платежа по договору",
            "Участие брокера",
            "Филиал",
            "Контактное лицо",
            "Крайняя дата запроса счета",
            "Примечание",
        ]

    def export(self):
        """Генерирует Excel файл с форматированием"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active

        # Добавляем заголовок отчета с диапазоном дат
        if self.date_from and self.date_to:
            date_from_str = self.date_from.strftime("%d.%m.%Y")
            date_to_str = self.date_to.strftime("%d.%m.%Y")
            report_title = f"ОЧЕРЕДНЫЕ ВЗНОСЫ С {date_from_str} ПО {date_to_str}"
        else:
            from datetime import date

            current_date = date.today().strftime("%d.%m.%Y")
            report_title = f"ОЧЕРЕДНЫЕ ВЗНОСЫ НА {current_date}"

        ws.append([report_title])

        # Форматирование заголовка отчета
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14, color="000000")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Объединяем ячейки для заголовка отчета
        num_columns = len(self.get_headers())
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)

        # Пустая строка после заголовка отчета
        ws.append([""])

        # Заголовки столбцов
        headers = self.get_headers()
        ws.append(headers)

        # Форматирование заголовков столбцов (теперь в строке 3)
        # Используем зеленый цвет вместо синего
        header_fill = PatternFill(
            start_color="2F855A", end_color="2F855A", fill_type="solid"
        )  # Темно-зеленый
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,  # Перенос текста в заголовках
            )

        # Увеличенная высота строки заголовков столбцов
        ws.row_dimensions[3].height = 30

        # Пустая строка после заголовков столбцов
        ws.append([""] * len(headers))

        # Группируем платежи по филиалам
        from collections import defaultdict

        payments_by_branch = defaultdict(list)
        for payment in self.queryset:
            branch_name = (
                payment.policy.branch.branch_name
                if payment.policy.branch
                else "Без филиала"
            )
            payments_by_branch[branch_name].append(payment)

        # Сортируем филиалы по алфавиту
        sorted_branches = sorted(payments_by_branch.keys())

        # Добавляем данные по филиалам
        for branch_name in sorted_branches:
            # Добавляем заголовок филиала
            ws.append([branch_name])
            branch_header_row = ws.max_row

            # Форматирование заголовка филиала (зеленая цветовая схема)
            branch_cell = ws.cell(row=branch_header_row, column=1)
            branch_cell.font = Font(bold=True, size=12, color="000000")
            branch_cell.fill = PatternFill(
                start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"
            )  # Светло-зеленый фон
            branch_cell.alignment = Alignment(horizontal="left", vertical="center")

            # Добавляем границы для заголовка филиала
            thin_border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
            branch_cell.border = thin_border

            ws.row_dimensions[branch_header_row].height = 25

            # Объединяем ячейки для заголовка филиала
            ws.merge_cells(
                start_row=branch_header_row,
                start_column=1,
                end_row=branch_header_row,
                end_column=num_columns,
            )

            # Применяем границы ко всем объединенным ячейкам
            for col in range(1, num_columns + 1):
                cell = ws.cell(row=branch_header_row, column=col)
                cell.border = thin_border

            # Добавляем платежи филиала
            for payment in payments_by_branch[branch_name]:
                row = self.get_row_data(payment)
                ws.append(row)

            # Пустая строка после блока филиала
            ws.append([""] * len(headers))

        # Форматирование
        self.apply_formatting(ws)

        return self.create_response(wb)

    def get_installment_status(self, payment, payments_in_year):
        """Определяет статус рассрочки для платежа"""
        # Если номер платежа не равен 1, то это рассрочка
        if payment.installment_number != 1:
            return "рассрочка"

        # Если номер платежа равен 1, проверяем есть ли другие платежи в том же году
        if payments_in_year and payments_in_year > 1:
            return "рассрочка"
        else:
            return "годовой"

    def get_row_data(self, payment):
        """Возвращает данные строки для платежа"""
        policy = payment.policy

        # Формируем текст "X из Y" для столбца "Этот взнос"
        payments_in_year = getattr(payment, "payments_in_year", None)
        if payments_in_year:
            payment_position = f"{payment.installment_number} из {payments_in_year}"
        else:
            # Fallback на случай, если аннотация не сработала
            payment_position = str(payment.installment_number)

        # Определяем статус рассрочки
        installment_status = self.get_installment_status(payment, payments_in_year)

        # Получаем фамилию менеджера лизинговой компании
        leasing_manager_name = ""
        if policy.leasing_manager:
            # Используем поле name (Фамилия менеджера)
            leasing_manager_name = policy.leasing_manager.name

        # Рассчитываем крайнюю дату запроса счета (дата платежа минус 2 недели)
        from datetime import timedelta

        invoice_request_deadline = None
        if payment.due_date:
            invoice_request_deadline = payment.due_date - timedelta(weeks=2)

        return [
            policy.policy_number,
            policy.dfa_number,
            policy.client.client_name,
            policy.insurer.insurer_name,
            policy.policyholder.client_name if policy.policyholder else "",
            self.format_value(policy.start_date),
            self.format_value(policy.end_date),
            policy.property_description,
            self.format_value(payment.insurance_sum),  # Страховая сумма из платежа
            self.format_value(
                payment.amount
            ),  # Сумма конкретного платежа (очередной взнос)
            installment_status,  # Статус рассрочки
            payment_position,  # "1 из 2", "2 из 3" и т.д.
            self.format_value(payment.due_date),
            self.format_value(policy.broker_participation),  # Участие брокера
            policy.branch.branch_name if policy.branch else "",
            leasing_manager_name,  # Контактное лицо (менеджер лизинговой компании)
            self.format_value(invoice_request_deadline),  # Крайняя дата запроса счета
            "",  # Примечание (пустое)
        ]

    def apply_formatting(self, ws):
        """Применяет расширенное форматирование к листу"""
        from openpyxl.styles import Alignment

        # 1. Настройка ширины столбцов
        # Фиксированная ширина для столбцов с предсказуемым форматом
        column_widths = {
            "A": 15,  # Номер полиса
            "B": 15,  # Номер ДФА
            "C": None,  # Лизингополучатель - автоподгонка
            "D": None,  # Страховщик - автоподгонка
            "E": None,  # Страхователь - автоподгонка
            "F": 13,  # Дата начала страхования
            "G": 13,  # Дата окончания страхования
            "H": None,  # Объект страхования - автоподгонка с большим лимитом
            "I": 16,  # Страховая сумма
            "J": 16,  # Очередной взнос
            "K": 12,  # Статус рассрочки
            "L": 12,  # Этот взнос ("1 из 2")
            "M": 13,  # Дата платежа по договору
            "N": 12,  # Участие брокера
            "O": None,  # Филиал - автоподгонка
            "P": None,  # Контактное лицо - автоподгонка
            "Q": 18,  # Крайняя дата запроса счета
            "R": 15,  # Примечание
        }

        from openpyxl.utils import get_column_letter

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            column_letter = get_column_letter(col_idx)

            # Если задана фиксированная ширина, используем её
            if (
                column_letter in column_widths
                and column_widths[column_letter] is not None
            ):
                ws.column_dimensions[column_letter].width = column_widths[column_letter]
            else:
                # Иначе автоподгонка
                max_length = 0
                for cell in column_cells:
                    # Пропускаем объединенные ячейки
                    if hasattr(cell, "value") and cell.value:
                        # Пропускаем заголовок отчета и заголовки филиалов
                        if isinstance(cell.value, str) and (
                            "ОЧЕРЕДНЫЕ ВЗНОСЫ" in cell.value
                            or (
                                col_idx == 1
                                and ws.cell(row=cell.row, column=2).value is None
                                and cell.row > 3
                            )
                        ):
                            continue
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = max_length + 2

                if column_letter == "H":  # Объект страхования (теперь столбец H)
                    adjusted_width = min(max(adjusted_width, 12), 60)
                else:
                    adjusted_width = min(max(adjusted_width, 10), 50)

                ws.column_dimensions[column_letter].width = adjusted_width

        # 2. Перенос текста для столбца "Объект страхования" (H)
        for cell in ws["H"]:
            if (
                cell.row > 3
            ):  # Пропускаем заголовок отчета, пустую строку и заголовки столбцов
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 3. Выравнивание и форматирование для разных типов данных
        from openpyxl.styles import PatternFill

        for row in ws.iter_rows(
            min_row=5
        ):  # Начинаем с 5 строки (первая строка данных)
            # Проверяем, является ли строка заголовком филиала
            first_cell_value = row[0].value
            is_branch_header = (
                first_cell_value
                and isinstance(first_cell_value, str)
                and ws.cell(row=row[0].row, column=2).value is None
                and row[0].row > 3
            )

            # Пропускаем форматирование заголовков филиалов (они уже отформатированы)
            if is_branch_header:
                continue

            # Проверяем значение в столбце "Участие брокера" (столбец N, индекс 14)
            broker_participation_cell = row[13] if len(row) > 13 else None
            is_no_broker_participation = (
                broker_participation_cell and broker_participation_cell.value == "Нет"
            )

            # Темно-серый фон для строк где участие брокера = "Нет"
            gray_fill = (
                PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                if is_no_broker_participation
                else None
            )

            for idx, cell in enumerate(row, start=1):
                # Применяем серый фон если участие брокера = "Нет"
                if gray_fill:
                    cell.fill = gray_fill

                # Столбцы с финансовыми данными (I - Страховая сумма, J - Очередной взнос)
                if idx in [9, 10]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    # Применяем числовой формат с разделителями тысяч и двумя десятичными знаками
                    # Без символа валюты для совместимости с Excel на Mac
                    cell.number_format = "#,##0.00"
                # Столбцы с датами (F, G, M, Q) - по центру
                elif idx in [6, 7, 13, 17]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                # Столбцы "Статус рассрочки" (K), "Этот взнос" (L) и "Участие брокера" (N) - по центру
                elif idx in [11, 12, 14]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                # Остальные - слева
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # 4. Закрепление заголовков (заголовок отчета + пустая строка + заголовки столбцов + пустая строка)
        ws.freeze_panes = "A5"  # Закрепляем первые 4 строки

        # 5. Автофильтры на заголовки столбцов (строка 3) с данными
        # Устанавливаем автофильтр на весь диапазон данных, включая заголовки
        if ws.max_row > 3:  # Проверяем, что есть данные
            ws.auto_filter.ref = (
                f"A3:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
            )

        # 6. Высота строк данных (пропускаем заголовки)
        for row in range(5, ws.max_row + 1):
            ws.row_dimensions[row].height = 20


class ThursdayReportExporter(BaseExporter):
    """Экспортер для четвергового отчета - полисы с платежами"""

    def __init__(self, queryset, fields, payment_date=None):
        """
        Инициализация экспортера

        Args:
            queryset: QuerySet полисов для раздела 1
            fields: Список полей (не используется, для совместимости)
            payment_date: Дата для фильтрации платежей в разделе 2
        """
        super().__init__(queryset, fields)
        self.payment_date = payment_date

    def get_filename(self):
        """Возвращает базовое имя файла"""
        return "thursday_report"

    def get_headers(self):
        """Возвращает список заголовков"""
        return [
            "Номер полиса",
            "Номер ДФА",
            "Филиал",
            "Лизингополучатель",
            "Страховщик",
            "Страхователь",
            "Дата начала страхования",
            "Дата оконч. страхования",
            "Объект страхования",
            "Страховая премия",
            "Дата платежа по договору",
            "Дата факт. оплаты",
            "Причина",
        ]

    def export(self):
        """Генерирует Excel файл с визуальным разделением по разделам"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import date

        wb = Workbook()
        ws = wb.active

        # Добавляем заголовок отчета с текущей датой
        current_date = date.today().strftime("%d.%m.%Y")
        report_title = f"ПЕРЕЧЕНЬ ДФА, ПО КОТОРЫМ НЕ ХВАТАЕТ ДОКУМЕНТОВ ПО СТРАХОВАНИЮ НА {current_date}"
        ws.append([report_title])

        # Форматирование заголовка отчета
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14, color="000000")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Объединяем ячейки для заголовка отчета
        num_columns = len(self.get_headers())
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)

        # Пустая строка после заголовка отчета
        ws.append([""])

        # Заголовки столбцов
        headers = self.get_headers()
        ws.append(headers)

        # Форматирование заголовков столбцов (теперь в строке 3)
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,  # Перенос текста в заголовках
            )

        # Увеличенная высота строки заголовков столбцов
        ws.row_dimensions[3].height = 30

        # Пустая строка после заголовков столбцов
        ws.append([""] * len(headers))

        # РАЗДЕЛ 1: Полисы без документов
        self._add_section_header(ws, "ПОЛИСЫ БЕЗ ДОКУМЕНТОВ", len(headers))

        for policy in self.queryset:
            row = self.get_row_data(policy)
            ws.append(row)

        # Пустая строка между разделами
        ws.append([])

        # РАЗДЕЛ 2: Платежи без данных об оплате
        self._add_section_header(ws, "НЕТ ДАННЫХ ОБ ОПЛАТЕ", len(headers))

        from apps.policies.models import PaymentSchedule

        # Получаем все платежи без даты оплаты (для активных полисов)
        # Фильтруем по дате: дата платежа <= payment_date
        unpaid_payments_query = PaymentSchedule.objects.filter(
            paid_date__isnull=True, policy__policy_active=True
        )

        # Применяем фильтр по дате, если указана
        if self.payment_date:
            unpaid_payments_query = unpaid_payments_query.filter(
                due_date__lte=self.payment_date
            )

        unpaid_payments = unpaid_payments_query.select_related(
            "policy", "policy__client", "policy__insurer", "policy__policyholder"
        ).order_by("due_date", "policy__policy_number")

        for payment in unpaid_payments:
            row = self.get_payment_row_data(payment)
            ws.append(row)

        # Форматирование
        self.apply_formatting(ws)

        return self.create_response(wb)

    def _add_section_header(self, ws, title, num_columns):
        """Добавляет заголовок раздела с форматированием"""
        from openpyxl.styles import Font, PatternFill, Alignment

        # Добавляем строку с заголовком раздела
        ws.append([title] + [""] * (num_columns - 1))
        current_row = ws.max_row

        # Объединяем ячейки для заголовка
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=num_columns,
        )

        # Форматирование заголовка раздела
        section_cell = ws.cell(row=current_row, column=1)
        section_cell.fill = PatternFill(
            start_color="FFC000", end_color="FFC000", fill_type="solid"
        )
        section_cell.font = Font(bold=True, size=13, color="000000")
        section_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Увеличенная высота строки заголовка раздела
        ws.row_dimensions[current_row].height = 25

    def get_row_data(self, policy):
        """Возвращает данные строки для полиса (раздел 1)"""
        from apps.policies.models import PaymentSchedule

        # Получаем первый платеж для полиса (если есть)
        first_payment = (
            PaymentSchedule.objects.filter(policy=policy)
            .order_by("year_number", "installment_number")
            .first()
        )

        # Определяем причину только по полису (не по платежу)
        # Для раздела 1 причина всегда связана с полисом
        reason = self.get_reason(policy, payment=None)

        return [
            policy.policy_number,
            policy.dfa_number,
            policy.branch.branch_name if policy.branch else "",
            policy.client.client_name,
            policy.insurer.insurer_name,
            policy.policyholder.client_name if policy.policyholder else "",
            self.format_value(policy.start_date),
            self.format_value(policy.end_date),
            policy.property_description,
            self.format_value(policy.premium_total),
            self.format_value(first_payment.due_date) if first_payment else "",
            self.format_value(first_payment.paid_date) if first_payment else "",
            reason,
        ]

    def get_payment_row_data(self, payment):
        """Возвращает данные строки для платежа (раздел 2)"""
        policy = payment.policy

        # Для раздела 2 причина всегда "нет данных об оплате"
        reason = "нет данных об оплате"

        return [
            policy.policy_number,
            policy.dfa_number,
            policy.branch.branch_name if policy.branch else "",
            policy.client.client_name,
            policy.insurer.insurer_name,
            policy.policyholder.client_name if policy.policyholder else "",
            self.format_value(policy.start_date),
            self.format_value(policy.end_date),
            policy.property_description,
            self.format_value(policy.premium_total),
            self.format_value(payment.due_date),
            self.format_value(payment.paid_date),
            reason,
        ]

    def get_reason(self, policy, payment=None):
        """
        Определяет причину попадания записи в отчет

        Используется только для раздела 1 (полисы без документов).
        Для раздела 2 причина устанавливается напрямую в get_payment_row_data().
        """
        # Проверяем полис
        if not policy.policy_uploaded:
            return "не подгружены документы"

        # Для будущих разделов можно добавить другие проверки:
        # if payment:
        #     if not payment.is_paid:
        #         if payment.is_overdue:
        #             return 'просрочен платеж'
        #         else:
        #             return 'не оплачен платеж'
        #     if payment.is_paid and not payment.is_approved:
        #         return 'ожидается согласование СК'

        return "другая причина"

    def apply_formatting(self, ws):
        """Применяет расширенное форматирование к листу"""
        from openpyxl.styles import Alignment

        # 1. Настройка ширины столбцов
        # Фиксированная ширина для столбцов с предсказуемым форматом
        column_widths = {
            "A": 15,  # Номер полиса
            "B": 15,  # Номер ДФА
            "C": None,  # Филиал - автоподгонка
            "D": None,  # Лизингополучатель - автоподгонка
            "E": None,  # Страховщик - автоподгонка
            "F": None,  # Страхователь - автоподгонка
            "G": 13,  # Дата начала страхования
            "H": 13,  # Дата окончания страхования
            "I": None,  # Объект страхования - автоподгонка с большим лимитом
            "J": 16,  # Страховая премия
            "K": 13,  # Дата платежа по договору
            "L": 13,  # Дата фактической оплаты
            "M": 22,  # Причина
        }

        from openpyxl.utils import get_column_letter

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            column_letter = get_column_letter(col_idx)

            # Если задана фиксированная ширина, используем её
            if (
                column_letter in column_widths
                and column_widths[column_letter] is not None
            ):
                ws.column_dimensions[column_letter].width = column_widths[column_letter]
            else:
                # Иначе автоподгонка
                max_length = 0
                for cell in column_cells:
                    # Пропускаем объединенные ячейки
                    if hasattr(cell, "value") and cell.value:
                        # Пропускаем заголовок отчета и заголовки разделов
                        if isinstance(cell.value, str) and (
                            "ПЕРЕЧЕНЬ ДФА" in cell.value
                            or "ПОЛИСЫ БЕЗ ДОКУМЕНТОВ" in cell.value
                            or "НЕТ ДАННЫХ ОБ ОПЛАТЕ" in cell.value
                        ):
                            continue
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = max_length + 2

                if column_letter == "I":  # Объект страхования (теперь столбец I)
                    adjusted_width = min(max(adjusted_width, 12), 60)
                else:
                    adjusted_width = min(max(adjusted_width, 10), 50)

                ws.column_dimensions[column_letter].width = adjusted_width

        # 2. Перенос текста для столбца "Объект страхования" (I)
        for cell in ws["I"]:
            if (
                cell.row > 3
            ):  # Пропускаем заголовок отчета, пустую строку и заголовки столбцов
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 3. Выравнивание для разных типов данных
        for row in ws.iter_rows(
            min_row=4
        ):  # Пропускаем заголовок отчета, пустую строку и заголовки столбцов
            for idx, cell in enumerate(row, start=1):
                if (
                    cell.value
                    and not isinstance(cell.value, str)
                    or (
                        isinstance(cell.value, str)
                        and cell.value.replace(".", "").replace(",", "").isdigit()
                    )
                ):
                    # Пропускаем заголовки разделов
                    if isinstance(cell.value, str) and (
                        "ПОЛИСЫ БЕЗ ДОКУМЕНТОВ" in cell.value
                        or "НЕТ ДАННЫХ ОБ ОПЛАТЕ" in cell.value
                    ):
                        continue

                    # Столбцы с датами (G, H, K, L) - по центру
                    if idx in [7, 8, 11, 12]:
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    # Столбцы с суммами (J) - справа
                    elif idx == 10:
                        cell.alignment = Alignment(
                            horizontal="right", vertical="center"
                        )
                    # Остальные текстовые - слева
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    # Текстовые поля - слева
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # 4. Закрепление заголовков (заголовок отчета + пустая строка + заголовки столбцов + пустая строка)
        ws.freeze_panes = "A5"  # Закрепляем первые 4 строки

        # 5. Автофильтры на заголовки столбцов (строка 3) с данными
        # Устанавливаем автофильтр на весь диапазон данных, включая заголовки
        if ws.max_row > 3:  # Проверяем, что есть данные
            ws.auto_filter.ref = (
                f"A3:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
            )

        # 6. Высота строк данных (пропускаем заголовки)
        for row in range(5, ws.max_row + 1):
            # Пропускаем заголовки разделов (они уже имеют высоту 25)
            cell_value = ws.cell(row=row, column=1).value
            if (
                cell_value
                and isinstance(cell_value, str)
                and (
                    "ПОЛИСЫ БЕЗ ДОКУМЕНТОВ" in cell_value
                    or "НЕТ ДАННЫХ ОБ ОПЛАТЕ" in cell_value
                )
            ):
                continue
            ws.row_dimensions[row].height = 20


class PolicyExpirationExporter(BaseExporter):
    """Экспортер для полисов с окончанием страхования в заданном периоде"""

    def __init__(self, queryset, fields, date_from=None, date_to=None):
        """
        Инициализация экспортера

        Args:
            queryset: QuerySet полисов
            fields: Список полей (не используется, для совместимости)
            date_from: Дата начала периода
            date_to: Дата окончания периода
        """
        super().__init__(queryset, fields)
        self.date_from = date_from
        self.date_to = date_to

    def get_filename(self):
        """Возвращает базовое имя файла"""
        return "policy_expiration"

    def get_headers(self):
        """Возвращает список заголовков"""
        return [
            # Основные данные (заполняются автоматически)
            "Номер полиса",
            "Номер ДФА",
            "Лизингополучатель",
            "Страховщик",
            "Страхователь",
            "Дата окончания страхования",
            "Объект страхования",
            # Дополнительные столбцы для ручного заполнения
            "Страхователь на новый период",
            "Выгодоприобретатель",
            "№ и дата кредитного договора / кредитной линии, банк-кредитор",
            "№ и дата договора залога",
            "Идентификационный признак имущества (VIN, заводской №)",
            "ГРН для транспортных средств и спецтехники",
            "Срок окончания ДФА (досрочного выкупа)",
            "Необходимый срок страхования",
            "Место нахождения имущества*",
            "Контактные данные для осмотра*",
            "Страховщик на новый срок",
            "Страховая сумма на новый срок",
            "Страховая премия на новый срок",
            "Условия страхования на новый срок",
            "Необходимость ПСО",
            "Дата отправки письма ЛП с предложением",
            "Ответ ЛП (дата и решение)",
            "Дата заключения договора на новый срок",
            "Примечания",
        ]

    def export(self):
        """Генерирует Excel файл с форматированием"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active

        # Добавляем заголовок отчета с диапазоном дат
        if self.date_from and self.date_to:
            date_from_str = self.date_from.strftime("%d.%m.%Y")
            date_to_str = self.date_to.strftime("%d.%m.%Y")
            report_title = f"ДОГОВОРА СТРАХОВАНИЯ С ОКОНЧАНИЕМ СРОКА СТРАХОВАНИЯ С {date_from_str} ПО {date_to_str}"
        else:
            from datetime import date

            current_date = date.today().strftime("%d.%m.%Y")
            report_title = (
                f"ДОГОВОРА СТРАХОВАНИЯ С ОКОНЧАНИЕМ СРОКА СТРАХОВАНИЯ НА {current_date}"
            )

        ws.append([report_title])

        # Форматирование заголовка отчета
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14, color="000000")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Объединяем ячейки для заголовка отчета
        num_columns = len(self.get_headers())
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)

        # Пустая строка после заголовка отчета
        ws.append([""])

        # Заголовки столбцов
        headers = self.get_headers()
        ws.append(headers)

        # Форматирование заголовков столбцов (теперь в строке 3)
        # Используем синий цвет как в четверговом отчете
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=9)  # Уменьшили с 11 до 9
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,  # Перенос текста в заголовках
            )

        # Увеличиваем высоту строки заголовков для длинных названий столбцов
        ws.row_dimensions[3].height = 60  # Увеличили с 45 до 60 для лучшего переноса

        # Добавляем строку 4 с информацией о заполняющих
        responsibility_row = [
            "заполняет страховой брокер",  # A
            "",  # B (будет объединена с A)
            "",  # C (будет объединена с A)
            "",  # D (будет объединена с A)
            "",  # E (будет объединена с A)
            "",  # F (будет объединена с A)
            "",  # G (будет объединена с A)
            "зап. УБТиСЛО",  # H
            "заполняет финансовая дирекция",  # I
            "",  # J (будет объединена с I)
            "",  # K (будет объединена с I)
            "заполняет УБТиСЛО",  # L
            "",  # M (будет объединена с L)
            "",  # N (будет объединена с L)
            "",  # O (будет объединена с L)
            "",  # P (будет объединена с L)
            "",  # Q (будет объединена с L)
            "заполняет страховой брокер",  # R
            "",  # S (будет объединена с R)
            "",  # T (будет объединена с R)
            "",  # U (будет объединена с R)
            "",  # V (будет объединена с R)
            "заполняет УБТиСЛО",  # W
            "",  # X (будет объединена с W)
            "заполн. стр. брокер",  # Y
            "заполняет УБТиСЛО",  # Z
        ]
        ws.append(responsibility_row)

        # Форматирование строки ответственности (строка 4)
        # Цвета для разных заполняющих (спокойные, но контрастные к синему #366092)
        broker_fill = PatternFill(
            start_color="D4F1D4", end_color="D4F1D4", fill_type="solid"
        )  # Мягкий зеленый
        ubt_fill = PatternFill(
            start_color="F0E6FF", end_color="F0E6FF", fill_type="solid"
        )  # Мягкий лавандовый
        finance_fill = PatternFill(
            start_color="FFF2E6", end_color="FFF2E6", fill_type="solid"
        )  # Мягкий кремовый

        responsibility_font = Font(bold=True, size=9, color="000000")

        # Объединяем ячейки и применяем форматирование
        # A-G: страховой брокер (зеленый)
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
        for col in range(1, 8):
            cell = ws.cell(row=4, column=col)
            cell.fill = broker_fill
            cell.font = responsibility_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # H: УБТиСЛО (розовый)
        cell_h = ws.cell(row=4, column=8)
        cell_h.fill = ubt_fill
        cell_h.font = responsibility_font
        cell_h.alignment = Alignment(horizontal="center", vertical="center")

        # I-K: финансовая дирекция (персиковый)
        ws.merge_cells(start_row=4, start_column=9, end_row=4, end_column=11)
        for col in range(9, 12):
            cell = ws.cell(row=4, column=col)
            cell.fill = finance_fill
            cell.font = responsibility_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # L-Q: УБТиСЛО (розовый)
        ws.merge_cells(start_row=4, start_column=12, end_row=4, end_column=17)
        for col in range(12, 18):
            cell = ws.cell(row=4, column=col)
            cell.fill = ubt_fill
            cell.font = responsibility_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # R-V: страховой брокер (зеленый)
        ws.merge_cells(start_row=4, start_column=18, end_row=4, end_column=22)
        for col in range(18, 23):
            cell = ws.cell(row=4, column=col)
            cell.fill = broker_fill
            cell.font = responsibility_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # W-X: УБТиСЛО (розовый)
        ws.merge_cells(start_row=4, start_column=23, end_row=4, end_column=24)
        for col in range(23, 25):
            cell = ws.cell(row=4, column=col)
            cell.fill = ubt_fill
            cell.font = responsibility_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Y: страховой брокер (зеленый)
        cell_y = ws.cell(row=4, column=25)
        cell_y.fill = broker_fill
        cell_y.font = responsibility_font
        cell_y.alignment = Alignment(horizontal="center", vertical="center")

        # Z: УБТиСЛО (розовый)
        cell_z = ws.cell(row=4, column=26)
        cell_z.fill = ubt_fill
        cell_z.font = responsibility_font
        cell_z.alignment = Alignment(horizontal="center", vertical="center")

        # Высота строки ответственности
        ws.row_dimensions[4].height = 25

        # Пустая строка после строки ответственности
        ws.append([""] * len(headers))

        # Группируем полисы по филиалам
        from collections import defaultdict

        policies_by_branch = defaultdict(list)
        for policy in self.queryset:
            branch_name = policy.branch.branch_name if policy.branch else "Без филиала"
            policies_by_branch[branch_name].append(policy)

        # Сортируем филиалы по алфавиту
        sorted_branches = sorted(policies_by_branch.keys())

        # Добавляем данные по филиалам
        for branch_name in sorted_branches:
            # Добавляем заголовок филиала
            ws.append([branch_name])
            branch_header_row = ws.max_row

            # Форматирование заголовка филиала
            branch_cell = ws.cell(row=branch_header_row, column=1)
            branch_cell.font = Font(bold=True, size=12, color="000000")
            branch_cell.fill = PatternFill(
                start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"
            )  # Светло-голубой фон
            branch_cell.alignment = Alignment(horizontal="left", vertical="center")

            # Добавляем границы для заголовка филиала
            thin_border = Border(
                left=Side(style="thin", color="000000"),
                right=Side(style="thin", color="000000"),
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
            branch_cell.border = thin_border

            ws.row_dimensions[branch_header_row].height = 25

            # Объединяем ячейки для заголовка филиала
            ws.merge_cells(
                start_row=branch_header_row,
                start_column=1,
                end_row=branch_header_row,
                end_column=num_columns,
            )

            # Применяем границы ко всем объединенным ячейкам
            for col in range(1, num_columns + 1):
                cell = ws.cell(row=branch_header_row, column=col)
                cell.border = thin_border

            # Добавляем полисы филиала
            for policy in policies_by_branch[branch_name]:
                row = self.get_row_data(policy)
                ws.append(row)

            # Пустая строка после блока филиала
            ws.append([""] * len(headers))

        # Форматирование
        self.apply_formatting(ws)

        return self.create_response(wb)

    def get_row_data(self, policy):
        """Возвращает данные строки для полиса"""
        # Определяем примечание для полисов без участия брокера
        notes = "без брокера" if not policy.broker_participation else ""

        return [
            # Основные данные (заполняются автоматически)
            policy.policy_number,
            policy.dfa_number,
            policy.client.client_name,
            policy.insurer.insurer_name,
            policy.policyholder.client_name if policy.policyholder else "",
            self.format_value(policy.end_date),
            policy.property_description,
            # Дополнительные столбцы для ручного заполнения (пустые)
            "",  # Страхователь на новый страховой период
            "",  # Выгодоприобретатель
            "",  # № и дата кредитного договора/кредитной линии, банк-кредитор
            "",  # № и дата договора залога
            "",  # Идентификационный признак имущества (VIN, заводской №)
            "",  # ГРН для транспортных средств и спецтехники
            "",  # Срок окончания ДФА (досрочного выкупа)
            "",  # Необходимый срок страхования
            "",  # Место нахождения имущества*
            "",  # Контактные данные для организации осмотра*
            "",  # Страховщик на новый срок
            "",  # страховая сумма на новый срок
            "",  # страховая премия на новый срок
            "",  # условия страхования на новый срок
            "",  # Необходимость ПСО
            "",  # Дата отправки письма ЛП-лю с предложением
            "",  # Ответ ЛП-ля (дата и решение)
            "",  # Дата заключения договора страхования на новый срок
            notes,  # Примечания (заполняется автоматически)
        ]

    def apply_formatting(self, ws):
        """Применяет расширенное форматирование к листу"""
        from openpyxl.styles import Alignment, Font, PatternFill

        # 1. Настройка ширины столбцов (оптимизировано для Times New Roman 9pt)
        column_widths = {
            # Основные столбцы
            "A": 12,  # Номер полиса
            "B": 12,  # Номер ДФА
            "C": None,  # Лизингополучатель - автоподгонка
            "D": None,  # Страховщик - автоподгонка
            "E": None,  # Страхователь - автоподгонка
            "F": 11,  # Дата окончания страхования
            "G": None,  # Объект страхования - автоподгонка
            # Дополнительные столбцы для ручного заполнения
            "H": 20,  # Страхователь на новый страховой период
            "I": 18,  # Выгодоприобретатель
            "J": 25,  # № и дата кредитного договора/кредитной линии, банк-кредитор
            "K": 20,  # № и дата договора залога
            "L": 18,  # Идентификационный признак имущества (VIN, заводской №)
            "M": 15,  # ГРН для транспортных средств и спецтехники
            "N": 15,  # Срок окончания ДФА (досрочного выкупа)
            "O": 15,  # Необходимый срок страхования
            "P": 20,  # Место нахождения имущества*
            "Q": 25,  # Контактные данные для организации осмотра*
            "R": 15,  # Страховщик на новый срок
            "S": 15,  # страховая сумма на новый срок
            "T": 15,  # страховая премия на новый срок
            "U": 20,  # условия страхования на новый срок
            "V": 12,  # Необходимость ПСО
            "W": 18,  # Дата отправки письма ЛП-лю с предложением
            "X": 20,  # Ответ ЛП-ля (дата и решение)
            "Y": 18,  # Дата заключения договора страхования на новый срок
            "Z": 15,  # Примечания
        }

        from openpyxl.utils import get_column_letter

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            column_letter = get_column_letter(col_idx)

            # Если задана фиксированная ширина, используем её
            if (
                column_letter in column_widths
                and column_widths[column_letter] is not None
            ):
                ws.column_dimensions[column_letter].width = column_widths[column_letter]
            else:
                # Иначе автоподгонка
                max_length = 0
                for cell in column_cells:
                    # Пропускаем объединенные ячейки
                    if hasattr(cell, "value") and cell.value:
                        # Пропускаем заголовок отчета и заголовки филиалов
                        if isinstance(cell.value, str) and (
                            "ДОГОВОРА СТРАХОВАНИЯ С ОКОНЧАНИЕМ СРОКА СТРАХОВАНИЯ"
                            in cell.value
                            or (
                                col_idx == 1
                                and ws.cell(row=cell.row, column=2).value is None
                                and cell.row > 4
                            )
                        ):
                            continue
                        max_length = max(max_length, len(str(cell.value)))

                # Уменьшаем отступ для Times New Roman 9pt
                adjusted_width = max_length + 1

                if column_letter == "G":  # Объект страхования
                    adjusted_width = min(
                        max(adjusted_width, 12), 45
                    )  # Минимум 12 для описания объектов
                elif column_letter in ["C", "E"]:  # Лизингополучатель, Страхователь
                    adjusted_width = min(
                        max(adjusted_width, 15), 40
                    )  # Минимум 15 для названий компаний
                else:  # Страховщик
                    adjusted_width = min(
                        max(adjusted_width, 10), 30
                    )  # Минимум 10 для страховщиков

                ws.column_dimensions[column_letter].width = adjusted_width

        # 2. Применяем шрифт Times New Roman 9pt ко всем ячейкам кроме заголовков
        times_font = Font(name="Times New Roman", size=9)
        gray_fill = PatternFill(
            start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
        )

        # Применяем шрифт Times New Roman 9pt ко всем строкам данных
        for row_idx in range(1, ws.max_row + 1):
            # Строка 1 - заголовок отчета (оставляем существующее форматирование)
            # Строка 3 - заголовки столбцов (оставляем существующее форматирование)
            # Строка 4 - строка ответственности (оставляем существующее форматирование)
            if row_idx in [1, 3, 4]:
                continue

            # Проверяем, является ли строка заголовком филиала
            first_cell = ws.cell(row=row_idx, column=1)
            is_branch_header = (
                first_cell.value
                and isinstance(first_cell.value, str)
                and ws.cell(row=row_idx, column=2).value
                is None  # Вторая ячейка пустая (объединенная)
                and row_idx > 4
            )

            # Заголовки филиалов - оставляем существующее форматирование
            if is_branch_header:
                continue

            # Проверяем пустые строки-разделители
            row_cells = [
                ws.cell(row=row_idx, column=col) for col in range(1, ws.max_column + 1)
            ]
            is_empty_row = all(
                cell.value is None or cell.value == "" for cell in row_cells
            )
            if is_empty_row:
                continue

            # Для всех остальных строк (строки данных) применяем Times New Roman 9pt
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = times_font

        # 3. Выделяем серым фоном строки где статус участия брокера = "Нет"
        # Нужно получить данные о broker_participation для каждого полиса
        current_row = 6  # Начинаем с первой строки данных (после добавления строки ответственности)

        # Группируем полисы по филиалам (как в основном методе export)
        from collections import defaultdict

        policies_by_branch = defaultdict(list)
        for policy in self.queryset:
            branch_name = policy.branch.branch_name if policy.branch else "Без филиала"
            policies_by_branch[branch_name].append(policy)

        # Сортируем филиалы по алфавиту
        sorted_branches = sorted(policies_by_branch.keys())

        # Проходим по данным и выделяем строки без участия брокера
        for branch_name in sorted_branches:
            # Пропускаем заголовок филиала
            current_row += 1

            # Обрабатываем полисы филиала
            for policy in policies_by_branch[branch_name]:
                # Проверяем статус участия брокера
                if not policy.broker_participation:
                    # Выделяем всю строку серым фоном
                    for col_idx in range(1, len(self.get_headers()) + 1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.fill = gray_fill
                        # Применяем шрифт Times New Roman 9pt (перезаписываем)
                        cell.font = times_font

                current_row += 1

            # Пропускаем пустую строку после блока филиала
            current_row += 1

        # 4. Перенос текста для столбцов с длинными заголовками и содержимым
        wrap_text_columns = [
            "G",
            "J",
            "P",
            "Q",
            "U",
        ]  # Объект страхования, кредитный договор, место нахождения, контакты, условия
        for column_letter in wrap_text_columns:
            for cell in ws[column_letter]:
                if cell.row > 4:  # Пропускаем заголовки и строку ответственности
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Перенос текста уже применен ко всем заголовкам выше

        # 5. Выравнивание для разных типов данных
        for row in ws.iter_rows(min_row=6):
            # Проверяем, является ли строка заголовком филиала
            first_cell_value = row[0].value
            is_branch_header = (
                first_cell_value
                and isinstance(first_cell_value, str)
                and len(row) > 1
                and row[1].value is None  # Вторая ячейка пустая (объединенная)
            )

            # Пропускаем пустые строки-разделители
            is_empty_row = all(cell.value is None or cell.value == "" for cell in row)

            if is_branch_header or is_empty_row:
                continue

            for idx, cell in enumerate(row, start=1):
                # Столбец с датой (F) - по центру
                if idx == 6:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                # Остальные - слева
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # 6. Закрепление заголовков
        ws.freeze_panes = "A6"  # Закрепляем первые 5 строк (заголовок, пустая, столбцы, ответственность, пустая)

        # 7. Автофильтры на заголовки столбцов (строка 3) с данными
        if ws.max_row > 3:  # Проверяем, что есть данные
            ws.auto_filter.ref = (
                f"A3:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
            )

        # 8. Высота строк данных
        for row in range(6, ws.max_row + 1):
            # Проверяем, является ли строка заголовком филиала (уже имеет высоту 25)
            cell_value = ws.cell(row=row, column=1).value
            is_branch_header = (
                cell_value
                and isinstance(cell_value, str)
                and ws.cell(row=row, column=2).value is None
            )

            if not is_branch_header:
                ws.row_dimensions[row].height = 20


class CommissionReportExporter(BaseExporter):
    """Экспортер для отчета по КВ - платежи оплаченные но не согласованные СК"""

    def __init__(self, queryset, fields, insurer_name=None):
        """
        Инициализация экспортера

        Args:
            queryset: QuerySet платежей
            fields: Список полей (не используется, для совместимости)
            insurer_name: Название страховой компании для заголовка
        """
        super().__init__(queryset, fields)
        self.insurer_name = insurer_name

    def get_filename(self):
        """Возвращает базовое имя файла"""
        return "commission_report_kv"

    def get_headers(self):
        """Возвращает список заголовков"""
        return [
            "Номер полиса",
            "Страхователь",
            "Дата начала страхования",
            "Дата оконч. страхования",
            "Объект страхования",
            "Страховая сумма",
            "Очередной взнос",
            "КВ %",
            "КВ руб",
            "Дата платежа по договору",
            "Дата факт. оплаты",
            "Филиал",
        ]

    def export(self):
        """Генерирует Excel файл с форматированием"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import date

        wb = Workbook()
        ws = wb.active

        # Добавляем заголовок отчета с названием страховой компании и датой
        current_date = date.today().strftime("%d.%m.%Y")
        if self.insurer_name:
            report_title = (
                f"ОТЧЕТ ПО КВ - {self.insurer_name.upper()} - НА {current_date}"
            )
        else:
            report_title = f"ОТЧЕТ ПО КВ - НА {current_date}"

        ws.append([report_title])

        # Форматирование заголовка отчета
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14, color="000000")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Объединяем ячейки для заголовка отчета
        num_columns = len(self.get_headers())
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)

        # Пустая строка после заголовка отчета
        ws.append([""])

        # Заголовки столбцов
        headers = self.get_headers()
        ws.append(headers)

        # Форматирование заголовков столбцов (теперь в строке 3)
        # Используем темно-коричневый цвет
        header_fill = PatternFill(
            start_color="5D4037", end_color="5D4037", fill_type="solid"
        )  # Темно-коричневый (Brown 800)
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,  # Перенос текста в заголовках
            )

        # Увеличенная высота строки заголовков столбцов
        ws.row_dimensions[3].height = 30

        # Пустая строка после заголовков столбцов
        ws.append([""] * len(headers))

        # Данные платежей
        for payment in self.queryset:
            row = self.get_row_data(payment)
            ws.append(row)

        # Форматирование
        self.apply_formatting(ws)

        return self.create_response(wb)

    def get_row_data(self, payment):
        """Возвращает данные строки для платежа"""
        policy = payment.policy

        # Получаем фактический КВ % (рассчитанный из КВ руб и премии)
        kv_percent = int(round(float(payment.kv_percent_actual)))

        return [
            policy.policy_number,
            policy.policyholder.client_name if policy.policyholder else "",
            self.format_value(policy.start_date),
            self.format_value(policy.end_date),
            policy.property_description,
            self.format_value(payment.insurance_sum),  # Страховая сумма из платежа
            self.format_value(
                payment.amount
            ),  # Сумма конкретного платежа (очередной взнос)
            kv_percent,  # КВ %
            self.format_value(payment.kv_rub),  # КВ руб
            self.format_value(payment.due_date),  # Дата платежа по договору
            self.format_value(payment.paid_date),  # Дата фактической оплаты
            policy.branch.branch_name if policy.branch else "",
        ]

    def apply_formatting(self, ws):
        """Применяет расширенное форматирование к листу"""
        from openpyxl.styles import Alignment

        # 1. Настройка ширины столбцов
        # Фиксированная ширина для столбцов с предсказуемым форматом
        column_widths = {
            "A": 15,  # Номер полиса
            "B": None,  # Страхователь - автоподгонка
            "C": 13,  # Дата начала страхования
            "D": 13,  # Дата окончания страхования
            "E": None,  # Объект страхования - автоподгонка с большим лимитом
            "F": 16,  # Страховая сумма
            "G": 16,  # Очередной взнос
            "H": 10,  # КВ %
            "I": 16,  # КВ руб
            "J": 13,  # Дата платежа по договору
            "K": 13,  # Дата факт. оплаты
            "L": None,  # Филиал - автоподгонка
        }

        from openpyxl.utils import get_column_letter

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            column_letter = get_column_letter(col_idx)

            # Если задана фиксированная ширина, используем её
            if (
                column_letter in column_widths
                and column_widths[column_letter] is not None
            ):
                ws.column_dimensions[column_letter].width = column_widths[column_letter]
            else:
                # Иначе автоподгонка
                max_length = 0
                for cell in column_cells:
                    # Пропускаем объединенные ячейки
                    if hasattr(cell, "value") and cell.value:
                        # Пропускаем заголовок отчета
                        if isinstance(cell.value, str) and "ОТЧЕТ ПО КВ" in cell.value:
                            continue
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = max_length + 2

                if column_letter == "E":  # Объект страхования (теперь столбец E)
                    adjusted_width = min(max(adjusted_width, 12), 60)
                else:
                    adjusted_width = min(max(adjusted_width, 10), 50)

                ws.column_dimensions[column_letter].width = adjusted_width

        # 2. Перенос текста для столбца "Объект страхования" (E)
        for cell in ws["E"]:
            if (
                cell.row > 3
            ):  # Пропускаем заголовок отчета, пустую строку и заголовки столбцов
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 3. Выравнивание и форматирование для разных типов данных
        for row in ws.iter_rows(
            min_row=5
        ):  # Начинаем с 5 строки (первая строка данных)
            for idx, cell in enumerate(row, start=1):
                # Столбцы с финансовыми данными (F - Страховая сумма, G - Очередной взнос, I - КВ руб)
                if idx in [6, 7, 9]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    # Применяем числовой формат с разделителями тысяч и двумя десятичными знаками
                    # Без символа валюты для совместимости с Excel на Mac
                    cell.number_format = "#,##0.00"
                # Столбец КВ % (H) - по центру
                elif idx == 8:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                # Столбцы с датами (C, D, J, K) - по центру
                elif idx in [3, 4, 10, 11]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                # Остальные - слева
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # 4. Закрепление заголовков (заголовок отчета + пустая строка + заголовки столбцов + пустая строка)
        ws.freeze_panes = "A5"  # Закрепляем первые 4 строки

        # 5. Автофильтры на заголовки столбцов (строка 3) с данными
        # Устанавливаем автофильтр на весь диапазон данных, включая заголовки
        if ws.max_row > 3:  # Проверяем, что есть данные
            ws.auto_filter.ref = (
                f"A3:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
            )

        # 6. Высота строк данных (пропускаем заголовки)
        for row in range(5, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
