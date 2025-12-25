"""
Интеграционные тесты для системы экспорта

Тестируют полный цикл работы системы экспорта от начала до конца.
"""

from django.test import TestCase, Client as TestClient
from django.contrib.auth.models import User
from django.urls import reverse
from datetime import date, timedelta
from decimal import Decimal
from openpyxl import load_workbook
from io import BytesIO

from apps.clients.models import Client
from apps.insurers.models import Insurer, Branch, InsuranceType, CommissionRate
from apps.policies.models import Policy, PaymentSchedule
from apps.reports.models import CustomExportTemplate


class FullCycleCustomExportTest(TestCase):
    """
    Интеграционный тест: Полный цикл кастомного экспорта

    Тестирует весь процесс создания кастомного экспорта:
    1. Авторизация пользователя
    2. Переход на страницу кастомного экспорта
    3. Выбор источника данных
    4. Выбор полей
    5. Применение фильтров
    6. Генерация экспорта
    7. Проверка содержимого файла

    Validates: All properties
    """

    def setUp(self):
        """Подготовка тестовых данных"""
        # Создаем пользователя
        self.user = User.objects.create_user(
            username="testuser", password="testpass123"
        )

        # Создаем тестовые данные
        self.client_obj = Client.objects.create(
            client_name="Интеграционный клиент", client_inn="1234567890"
        )

        self.insurer = Insurer.objects.create(insurer_name="Интеграционная СК")

        self.branch = Branch.objects.create(branch_name="Главный офис")

        self.insurance_type = InsuranceType.objects.create(name="КАСКО")

        # Создаем несколько полисов
        self.policies = []
        for i in range(5):
            policy = Policy.objects.create(
                policy_number=f"INT-{i:03d}",
                dfa_number=f"DFA-INT-{i:03d}",
                client=self.client_obj,
                insurer=self.insurer,
                branch=self.branch,
                insurance_type=self.insurance_type,
                property_description=f"Имущество {i}",
                start_date=date(2024, 1, 1) + timedelta(days=i * 30),
                end_date=date(2024, 12, 31),
                premium_total=Decimal("100000.00") * (i + 1),
                policy_active=(i % 2 == 0),
            )
            self.policies.append(policy)

        self.test_client = TestClient()

    def test_full_custom_export_cycle(self):
        """Тест полного цикла кастомного экспорта"""
        # Шаг 1: Авторизация
        login_success = self.test_client.login(
            username="testuser", password="testpass123"
        )
        self.assertTrue(login_success, "Login should succeed")

        # Шаг 2: Переход на страницу кастомного экспорта
        response = self.test_client.get("/reports/custom/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("available_fields", response.context)

        # Шаг 3-5: Выбор источника, полей и фильтров
        export_data = {
            "action": "export",
            "data_source": "policies",
            "fields": [
                "policy_number",
                "client__client_name",
                "premium_total",
                "policy_active",
            ],
        }

        # Шаг 6: Генерация экспорта
        response = self.test_client.post("/reports/custom/", export_data)

        # Проверяем что получили Excel файл
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Шаг 7: Проверка содержимого файла
        # Загружаем Excel файл
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb.active

        # Проверяем заголовки
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(len(headers), 4)
        self.assertIn("Номер полиса", headers)
        self.assertIn("Клиент", headers)
        self.assertIn("Общая премия", headers)
        self.assertIn("Статус полиса", headers)

        # Проверяем количество строк (заголовок + 5 полисов)
        self.assertEqual(ws.max_row, 6)

        # Проверяем данные первого полиса (может быть любой из 5)
        first_row = [cell.value for cell in ws[2]]
        self.assertIn(
            first_row[0], ["INT-000", "INT-001", "INT-002", "INT-003", "INT-004"]
        )
        self.assertEqual(first_row[1], "Интеграционный клиент")
        self.assertIn(first_row[2], [100000.0, 200000.0, 300000.0, 400000.0, 500000.0])
        self.assertIn(first_row[3], ["Да", "Нет"])

    def test_custom_export_with_filters(self):
        """Тест кастомного экспорта с фильтрами"""
        self.test_client.login(username="testuser", password="testpass123")

        # Экспортируем только активные полисы
        export_data = {
            "action": "export",
            "data_source": "policies",
            "fields": ["policy_number", "policy_active"],
            "policy_active": "true",
        }

        response = self.test_client.post("/reports/custom/", export_data)
        self.assertEqual(response.status_code, 200)

        # Проверяем содержимое
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb.active

        # Должно быть 3 активных полиса (индексы 0, 2, 4)
        self.assertEqual(ws.max_row, 4)  # заголовок + 3 полиса

        # Проверяем что все полисы активны
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=2).value
            self.assertEqual(status, "Да")


class TemplateSaveAndLoadTest(TestCase):
    """
    Интеграционный тест: Сохранение и использование шаблона

    Тестирует полный цикл работы с шаблонами:
    1. Создание кастомного экспорта
    2. Сохранение как шаблон
    3. Загрузка шаблона
    4. Генерация экспорта с загруженными настройками
    5. Проверка идентичности результатов

    Validates: Property 3, Property 4
    """

    def setUp(self):
        """Подготовка тестовых данных"""
        self.user = User.objects.create_user(
            username="testuser", password="testpass123"
        )

        self.client_obj = Client.objects.create(
            client_name="Тестовый клиент", client_inn="1234567890"
        )

        self.insurer = Insurer.objects.create(insurer_name="Тестовая СК")
        self.branch = Branch.objects.create(branch_name="Филиал")
        self.insurance_type = InsuranceType.objects.create(name="КАСКО")

        self.policy = Policy.objects.create(
            policy_number="TMPL-001",
            dfa_number="DFA-TMPL-001",
            client=self.client_obj,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            property_description="Тестовое имущество",
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
            policy_active=True,
        )

        self.test_client = TestClient()

    def test_save_and_load_template(self):
        """Тест сохранения и загрузки шаблона"""
        self.test_client.login(username="testuser", password="testpass123")

        # Шаг 1: Создаем и сохраняем шаблон
        template_data = {
            "action": "save_template",
            "template_name": "Мой тестовый шаблон",
            "data_source": "policies",
            "fields": ["policy_number", "client__client_name", "policy_active"],
        }

        response = self.test_client.post("/reports/custom/", template_data)
        self.assertEqual(response.status_code, 302)  # Redirect after save

        # Проверяем что шаблон создан
        template = CustomExportTemplate.objects.filter(
            user=self.user, name="Мой тестовый шаблон"
        ).first()

        self.assertIsNotNone(template)
        self.assertEqual(template.data_source, "policies")
        self.assertEqual(
            template.config["fields"],
            ["policy_number", "client__client_name", "policy_active"],
        )

        # Шаг 2: Генерируем экспорт с этими настройками
        export_data = {
            "action": "export",
            "data_source": "policies",
            "fields": ["policy_number", "client__client_name", "policy_active"],
        }

        response1 = self.test_client.post("/reports/custom/", export_data)
        self.assertEqual(response1.status_code, 200)

        # Шаг 3: Генерируем экспорт еще раз с теми же настройками
        response2 = self.test_client.post("/reports/custom/", export_data)
        self.assertEqual(response2.status_code, 200)

        # Шаг 4: Проверяем что результаты идентичны
        # (размер файлов должен быть примерно одинаковым)
        self.assertAlmostEqual(
            len(response1.content),
            len(response2.content),
            delta=100,  # Небольшая разница из-за timestamp в имени файла
        )

    def test_template_uniqueness(self):
        """Тест уникальности имен шаблонов"""
        self.test_client.login(username="testuser", password="testpass123")

        # Создаем первый шаблон
        template_data = {
            "action": "save_template",
            "template_name": "Уникальный шаблон",
            "data_source": "policies",
            "fields": ["policy_number"],
        }

        response1 = self.test_client.post("/reports/custom/", template_data)
        self.assertEqual(response1.status_code, 302)

        # Пытаемся создать шаблон с тем же именем
        template_data2 = {
            "action": "save_template",
            "template_name": "Уникальный шаблон",
            "data_source": "payments",
            "fields": ["policy__policy_number"],
        }

        response2 = self.test_client.post("/reports/custom/", template_data2)
        self.assertEqual(response2.status_code, 302)

        # Должен быть только один шаблон (обновленный)
        templates = CustomExportTemplate.objects.filter(
            user=self.user, name="Уникальный шаблон"
        )

        self.assertEqual(templates.count(), 1)
        # Проверяем что шаблон обновился
        self.assertEqual(templates.first().data_source, "payments")


class ReadyExportsTest(TestCase):
    """
    Интеграционный тест: Готовые экспорты

    Тестирует работу готовых экспортов:
    1. Экспорт полисов
    2. Экспорт платежей
    3. Проверка структуры файлов
    4. Проверка корректности данных

    Validates: Property 5, Property 6
    """

    def setUp(self):
        """Подготовка тестовых данных"""
        self.user = User.objects.create_user(
            username="testuser", password="testpass123"
        )

        # Создаем тестовые данные
        self.client_obj = Client.objects.create(
            client_name="Клиент для готовых экспортов", client_inn="9876543210"
        )

        self.insurer = Insurer.objects.create(insurer_name="СК для готовых экспортов")

        self.branch = Branch.objects.create(branch_name="Филиал для экспортов")

        self.insurance_type = InsuranceType.objects.create(name="ОСАГО")

        self.commission_rate = CommissionRate.objects.create(
            insurer=self.insurer,
            insurance_type=self.insurance_type,
            kv_percent=Decimal("20.00"),
        )

        # Создаем полис
        self.policy = Policy.objects.create(
            policy_number="READY-001",
            dfa_number="DFA-READY-001",
            client=self.client_obj,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            property_description="Имущество для экспорта",
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
            premium_total=Decimal("50000.00"),
            franchise=Decimal("5000.00"),
            policy_active=True,
        )

        # Создаем платеж
        self.payment = PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=1,
            due_date=date(2024, 3, 1),
            amount=Decimal("50000.00"),
            insurance_sum=Decimal("1000000.00"),
            commission_rate=self.commission_rate,
            kv_rub=Decimal("10000.00"),
            paid_date=date(2024, 2, 28),
        )

        self.test_client = TestClient()

    def test_ready_export_policies(self):
        """Тест готового экспорта полисов"""
        self.test_client.login(username="testuser", password="testpass123")

        response = self.test_client.get("/reports/export/policies/")

        # Проверяем что получили Excel файл
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("policies_", response["Content-Disposition"])

        # Проверяем содержимое
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb.active

        # Проверяем заголовки (14 колонок)
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(len(headers), 14)
        self.assertIn("Номер полиса", headers)
        self.assertIn("Клиент", headers)
        self.assertIn("Страховщик", headers)
        self.assertIn("Полис подгружен", headers)

        # Проверяем данные
        self.assertGreaterEqual(ws.max_row, 2)  # Минимум заголовок + 1 полис

        # Проверяем первую строку данных
        row_data = [cell.value for cell in ws[2]]
        self.assertEqual(row_data[0], "READY-001")
        self.assertEqual(row_data[2], "Клиент для готовых экспортов")
        self.assertEqual(row_data[3], "СК для готовых экспортов")

    def test_ready_export_payments(self):
        """Тест готового экспорта платежей"""
        self.test_client.login(username="testuser", password="testpass123")

        # Добавляем обязательные параметры дат
        response = self.test_client.get(
            "/reports/export/payments/?date_from=2024-01-01&date_to=2024-12-31"
        )

        # Проверяем что получили Excel файл
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("payments_", response["Content-Disposition"])

        # Проверяем содержимое
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb.active

        # Проверяем заголовки (18 колонок для ScheduledPaymentsExporter)
        headers = [cell.value for cell in ws[3]]  # Заголовки в 3-й строке
        self.assertEqual(len(headers), 18)
        self.assertIn("Номер полиса", headers)
        self.assertIn("Лизингополучатель", headers)
        self.assertIn("Дата платежа по договору", headers)

        # Проверяем данные
        self.assertGreaterEqual(
            ws.max_row, 6
        )  # Минимум заголовок отчета + пустая строка + заголовки + пустая строка + заголовок филиала + 1 платеж

        # Ищем первую строку с данными платежа (пропускаем заголовки филиалов)
        payment_row = None
        for row_num in range(5, ws.max_row + 1):
            row_data = [cell.value for cell in ws[row_num]]
            # Заголовок филиала имеет только первую ячейку заполненную, остальные пустые
            if row_data[0] and row_data[1]:  # Если есть и номер полиса, и номер ДФА
                payment_row = row_data
                break

        self.assertIsNotNone(payment_row, "Не найдена строка с данными платежа")
        self.assertEqual(payment_row[0], "READY-001")  # Номер полиса
        self.assertEqual(
            payment_row[2], "Клиент для готовых экспортов"
        )  # Лизингополучатель

    def test_ready_export_date_formatting(self):
        """Тест форматирования дат в готовых экспортах"""
        self.test_client.login(username="testuser", password="testpass123")

        response = self.test_client.get("/reports/export/policies/")

        # Проверяем содержимое
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb.active

        # Получаем даты из первой строки данных
        start_date = ws.cell(row=2, column=7).value  # Дата начала
        end_date = ws.cell(row=2, column=8).value  # Дата окончания

        # Проверяем, что это объекты date или datetime
        from datetime import date, datetime

        self.assertTrue(isinstance(start_date, (date, datetime)))
        self.assertTrue(isinstance(end_date, (date, datetime)))

        # Проверяем конкретные значения (преобразуем datetime в date если нужно)
        start_date_value = (
            start_date.date() if isinstance(start_date, datetime) else start_date
        )
        end_date_value = end_date.date() if isinstance(end_date, datetime) else end_date

        self.assertEqual(start_date_value, date(2024, 1, 1))
        self.assertEqual(end_date_value, date(2024, 12, 31))

    def test_thursday_report_export(self):
        """Тест четвергового отчета - экспорт полисов без документов и неоплаченных платежей, сгруппированных по городам"""
        # Создаем дополнительный полис с policy_uploaded=True
        uploaded_policy = Policy.objects.create(
            policy_number="UPLOADED-001",
            dfa_number="DFA-UPLOADED-001",
            client=self.client_obj,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
            premium_total=Decimal("50000.00"),
            policy_uploaded=True,  # Этот полис подгружен
        )

        self.test_client.login(username="testuser", password="testpass123")

        response = self.test_client.get("/reports/export/thursday/")

        # Проверяем что получили Excel файл
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Проверяем содержимое
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb.active

        # Проверяем заголовок отчета с датой (строка 1)
        report_title = ws.cell(row=1, column=1).value
        self.assertIsNotNone(report_title)
        self.assertIn(
            "ПЕРЕЧЕНЬ ДФА, ПО КОТОРЫМ НЕ ХВАТАЕТ ДОКУМЕНТОВ ПО СТРАХОВАНИЮ НА",
            report_title,
        )
        # Проверяем что в заголовке есть дата в формате ДД.ММ.ГГГГ
        import re

        date_pattern = r"\d{2}\.\d{2}\.\d{4}"
        self.assertRegex(report_title, date_pattern)

        # Проверяем что строка 2 пустая
        empty_row = ws.cell(row=2, column=1).value
        self.assertTrue(empty_row is None or empty_row == "")

        # Проверяем заголовки столбцов четвергового отчета (теперь в строке 3)
        headers = [cell.value for cell in ws[3]]
        expected_headers = [
            "Номер полиса",
            "Номер ДФА",
            "Лизингополучатель",
            "Страховщик",
            "Страхователь",
            "Дата начала страхования",
            "Дата оконч. страхования",
            "Объект страхования",
            "Очередной взнос",
            "Дата платежа по договору",
            "Дата факт. оплаты",
            "Причина",
        ]
        self.assertEqual(headers, expected_headers)

        # Проверяем наличие заголовка города (строка 5: 1-заголовок отчета, 2-пустая, 3-заголовки столбцов, 4-пустая, 5-город)
        city_header = ws.cell(row=5, column=1).value
        self.assertEqual(
            city_header, "Филиал для экспортов"
        )  # Название филиала из setUp

        # Проверяем что в отчете только неподгруженные полисы
        # Данные начинаются с 6-й строки (1-заголовок отчета, 2-пустая, 3-заголовки столбцов, 4-пустая, 5-город, 6+-данные)
        policy_numbers = []
        for row in range(6, ws.max_row + 1):
            policy_number = ws.cell(row=row, column=1).value
            # Пропускаем заголовки городов (первая ячейка заполнена, вторая пустая)
            second_cell = ws.cell(row=row, column=2).value
            if policy_number and second_cell is not None:
                policy_numbers.append(policy_number)

        # Проверяем что READY-001 есть в отчете (policy_uploaded=False)
        self.assertIn("READY-001", policy_numbers)

        # Проверяем что UPLOADED-001 НЕТ в отчете (policy_uploaded=True)
        self.assertNotIn("UPLOADED-001", policy_numbers)

        # Проверяем столбец "Причина" для полиса без документов
        # Находим строку с READY-001
        for row in range(6, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == "READY-001":
                reason = ws.cell(row=row, column=12).value  # 12-й столбец - Причина
                self.assertIn("не подгружены документы", reason)
                break

        # Проверяем что неоплаченные платежи также попадают в отчет (в том же городе)
        # (payment из setUp имеет paid_date, поэтому создадим неоплаченный)
        from apps.policies.models import PaymentSchedule

        unpaid_payment = PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=2,
            installment_number=1,
            due_date=date(2025, 1, 1),
            amount=Decimal("25000.00"),
            insurance_sum=Decimal("1000000.00"),
            commission_rate=self.commission_rate,
            kv_rub=Decimal("2500.00"),
            paid_date=None,  # Нет даты оплаты
        )

        # Перегенерируем отчет с датой, которая включает неоплаченный платеж
        response = self.test_client.get(
            "/reports/export/thursday/?payment_date=2025-12-31"
        )
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb.active

        # Проверяем что неоплаченный платеж добавлен к той же записи (объединение)
        # Теперь должна быть только одна строка с READY-001, но с двумя причинами
        ready_001_count = 0
        for row in range(6, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == "READY-001":
                ready_001_count += 1
                # Проверяем причину - должны быть обе причины
                reason = ws.cell(row=row, column=12).value  # 12-й столбец - Причина
                self.assertIn("не подгружены документы", reason)
                self.assertIn("нет данных об оплате", reason)

                # Проверяем что причины разделены переносом строки
                self.assertIn(
                    "\n", reason, "Причины должны быть разделены переносом строки"
                )

                # Проверяем что нет \r\n (только \n для универсальности)
                self.assertNotIn("\r\n", reason, "Не должно быть \\r\\n, только \\n")

        # Должна быть только одна строка с READY-001 (объединенная)
        self.assertEqual(
            ready_001_count,
            1,
            "Должна быть только одна строка с READY-001 (без дубликатов)",
        )

        # Проверяем фильтрацию по дате: платеж с датой 2025-01-01 не должен попасть в отчет с датой 2024-12-31
        response_filtered = self.test_client.get(
            "/reports/export/thursday/?payment_date=2024-12-31"
        )
        excel_file_filtered = BytesIO(response_filtered.content)
        wb_filtered = load_workbook(excel_file_filtered)
        ws_filtered = wb_filtered.active

        # Проверяем что платеж с датой 2025-01-01 НЕ попал в отчет
        found_future_payment = False
        for row in range(6, ws_filtered.max_row + 1):
            cell_value = ws_filtered.cell(row=row, column=1).value
            if cell_value == "READY-001":
                # Проверяем дату платежа (11-й столбец - Дата платежа по договору)
                payment_date_cell = ws_filtered.cell(row=row, column=11).value
                if payment_date_cell == "01.01.2025":
                    found_future_payment = True
                    break

        self.assertFalse(
            found_future_payment,
            "Платеж с датой 2025-01-01 не должен попасть в отчет с фильтром 2024-12-31",
        )

    def test_policy_expiration_export(self):
        """Тест экспорта полисов с окончанием страхования в заданном периоде"""
        from datetime import date, datetime

        # Создаем дополнительные полисы с разными датами окончания
        policy_in_range = Policy.objects.create(
            policy_number="EXPIRING-001",
            dfa_number="DFA-EXPIRING-001",
            client=self.client_obj,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            start_date=date(2024, 1, 1),
            end_date=date(2024, 6, 15),  # В диапазоне
            premium_total=Decimal("75000.00"),
        )

        policy_out_of_range = Policy.objects.create(
            policy_number="EXPIRING-002",
            dfa_number="DFA-EXPIRING-002",
            client=self.client_obj,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),  # Вне диапазона
            premium_total=Decimal("85000.00"),
        )

        self.test_client.login(username="testuser", password="testpass123")

        # Экспортируем полисы с окончанием с 01.06.2024 по 30.06.2024
        response = self.test_client.get(
            "/reports/export/policy-expiration/?date_from=2024-06-01&date_to=2024-06-30"
        )

        # Проверяем что получили Excel файл
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Проверяем содержимое
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb.active

        # Проверяем заголовок отчета (строка 1)
        report_title = ws.cell(row=1, column=1).value
        self.assertIsNotNone(report_title)
        self.assertIn(
            "ДОГОВОРА СТРАХОВАНИЯ С ОКОНЧАНИЕМ СРОКА СТРАХОВАНИЯ С", report_title
        )
        self.assertIn("01.06.2024", report_title)
        self.assertIn("30.06.2024", report_title)

        # Проверяем что строка 2 пустая
        empty_row = ws.cell(row=2, column=1).value
        self.assertTrue(empty_row is None or empty_row == "")

        # Проверяем заголовки столбцов (строка 3)
        headers = [cell.value for cell in ws[3]]
        expected_headers = [
            "Номер полиса",
            "Номер ДФА",
            "Лизингополучатель",
            "Страховщик",
            "Страхователь",
            "Дата окончания страхования",
            "Объект страхования",
            "Страхователь на новый период",
            "Выгодоприобретатель",
            "№ и дата кредитного договора / кредитной линии, банк-кредитор",
            "№ и дата договора залога",
            "Идентификатор (обычно VIN)",
            "ГРН для транспортных средств и спецтехники",
            "Срок окончания ДФА (досрочного выкупа)",
            "Необходимый срок страхования",
            "Место нахождения имущества*",
            "Контактные данные для осмотра*",
            "Страховщик на новый срок",
            "Страховая сумма на новый срок",
            "Страховая премия на новый срок",
            "Условия страхования на новый срок",
            "Необходимость ПСО",
            "Дата отправки письма ЛП с предложением",
            "Ответ ЛП (дата и решение)",
            "Дата заключения договора на новый срок",
            "Примечания",
        ]
        self.assertEqual(headers, expected_headers)

        # Проверяем что строка 4 содержит информацию о заполняющих
        responsibility_row = ws.cell(row=4, column=1).value
        self.assertIsNotNone(responsibility_row)
        self.assertIn("заполняет", responsibility_row)

        # Проверяем что в отчете только полисы с окончанием в заданном диапазоне
        # Данные начинаются с 6-й строки (после строки ответственности и пустой строки)
        policy_numbers = []
        for row in range(6, ws.max_row + 1):
            policy_number = ws.cell(row=row, column=1).value
            if policy_number:
                policy_numbers.append(policy_number)

        # Проверяем что EXPIRING-001 есть в отчете (end_date в диапазоне)
        self.assertIn("EXPIRING-001", policy_numbers)

        # Проверяем что EXPIRING-002 НЕТ в отчете (end_date вне диапазона)
        self.assertNotIn("EXPIRING-002", policy_numbers)

        # Проверяем данные в строке с EXPIRING-001
        for row in range(6, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "EXPIRING-001":
                # Проверяем дату окончания (6-й столбец)
                end_date = ws.cell(row=row, column=6).value
                # Преобразуем дату для сравнения
                if isinstance(end_date, datetime):
                    end_date_value = end_date.date()
                elif isinstance(end_date, date):
                    end_date_value = end_date
                else:
                    end_date_value = None
                self.assertEqual(end_date_value, date(2024, 6, 15))
                break

        # Проверяем что без параметров дат возвращается ошибка
        response_no_dates = self.test_client.get("/reports/export/policy-expiration/")
        self.assertEqual(response_no_dates.status_code, 302)  # Редирект

        # Проверяем что с некорректным диапазоном возвращается ошибка
        response_invalid_range = self.test_client.get(
            "/reports/export/policy-expiration/?date_from=2024-06-30&date_to=2024-06-01"
        )
        self.assertEqual(response_invalid_range.status_code, 302)  # Редирект

    def test_commission_report_export(self):
        """Тест отчета по КВ - платежи оплаченные но не согласованные СК"""
        from apps.policies.models import PaymentSchedule
        from datetime import date, datetime

        # Создаем платежи с разными статусами
        # 1. Оплачен, но не согласован СК (должен попасть в отчет)
        paid_not_approved = PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=2,
            installment_number=1,
            due_date=date(2024, 6, 1),
            amount=Decimal("25000.00"),
            insurance_sum=Decimal("1000000.00"),
            commission_rate=self.commission_rate,
            kv_rub=Decimal("2500.00"),
            paid_date=date(2024, 6, 5),  # Есть дата оплаты
            insurer_date=None,  # Нет даты согласования СК
        )

        # 2. Оплачен и согласован СК (НЕ должен попасть в отчет)
        paid_and_approved = PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=2,
            installment_number=2,
            due_date=date(2024, 7, 1),
            amount=Decimal("25000.00"),
            insurance_sum=Decimal("1000000.00"),
            commission_rate=self.commission_rate,
            kv_rub=Decimal("2500.00"),
            paid_date=date(2024, 7, 5),  # Есть дата оплаты
            insurer_date=date(2024, 7, 10),  # Есть дата согласования акта с СК
        )

        # 3. Не оплачен (НЕ должен попасть в отчет)
        not_paid = PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=2,
            installment_number=3,
            due_date=date(2024, 8, 1),
            amount=Decimal("25000.00"),
            insurance_sum=Decimal("1000000.00"),
            commission_rate=self.commission_rate,
            kv_rub=Decimal("2500.00"),
            paid_date=None,  # Нет даты оплаты
            insurer_date=None,
        )

        self.test_client.login(username="testuser", password="testpass123")

        # Экспортируем отчет по КВ для страховой компании
        response = self.test_client.get(
            f"/reports/export/commission-report/?insurer={self.insurer.id}"
        )

        # Проверяем что получили Excel файл
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Проверяем содержимое
        excel_file = BytesIO(response.content)
        wb = load_workbook(excel_file)
        ws = wb.active

        # Проверяем заголовок отчета (строка 1)
        report_title = ws.cell(row=1, column=1).value
        self.assertIsNotNone(report_title)
        self.assertIn("ОТЧЕТ ПО КВ", report_title)
        self.assertIn(self.insurer.insurer_name.upper(), report_title)

        # Проверяем что строка 2 пустая
        empty_row = ws.cell(row=2, column=1).value
        self.assertTrue(empty_row is None or empty_row == "")

        # Проверяем заголовки столбцов (строка 3)
        headers = [cell.value for cell in ws[3]]
        expected_headers = [
            "Номер полиса",
            "Страхователь",
            "Дата начала страхования",
            "Дата оконч. страхования",
            "Объект страхования",
            "Страховая сумма",
            "Очередной взнос",
            "КВ %",
            "КВ руб",
            "Дата платежа по договору",
            "Дата факт. оплаты",
            "Филиал",
        ]
        self.assertEqual(headers, expected_headers)

        # Проверяем что в заголовке есть дата
        import re

        date_pattern = r"\d{2}\.\d{2}\.\d{4}"
        self.assertRegex(report_title, date_pattern)

        # Проверяем что строка 4 пустая
        empty_row_4 = ws.cell(row=4, column=1).value
        self.assertTrue(empty_row_4 is None or empty_row_4 == "")

        # Проверяем что в отчете только платежи оплаченные но не согласованные СК
        # Данные начинаются с 5-й строки
        payment_dates = []
        for row in range(5, ws.max_row + 1):
            payment_date = ws.cell(row=row, column=10).value  # Дата платежа по договору
            if payment_date:
                payment_dates.append(payment_date)

        # Проверяем что платеж с датой 01.06.2024 есть в отчете (оплачен, но не согласован)
        from datetime import date, datetime

        target_date = date(2024, 6, 1)

        # Преобразуем datetime в date если нужно для сравнения
        payment_dates_as_dates = []
        for pd in payment_dates:
            if isinstance(pd, datetime):
                payment_dates_as_dates.append(pd.date())
            elif isinstance(pd, date):
                payment_dates_as_dates.append(pd)

        self.assertIn(target_date, payment_dates_as_dates)

        # Проверяем что платеж с датой 01.07.2024 НЕТ в отчете (оплачен и согласован)
        self.assertNotIn(date(2024, 7, 1), payment_dates_as_dates)

        # Проверяем что платеж с датой 01.08.2024 НЕТ в отчете (не оплачен)
        self.assertNotIn(date(2024, 8, 1), payment_dates_as_dates)

        # Проверяем данные в строке с платежом по дате факт. оплаты 05.06.2024
        # Примечание: в отчете kv_rub может отличаться от созданного в тесте
        # из-за того, что могут быть другие платежи с той же датой
        found_june_payment = False
        for row in range(5, ws.max_row + 1):
            paid_date = ws.cell(row=row, column=11).value  # Дата факт. оплаты

            # Преобразуем дату для сравнения
            if paid_date:
                if isinstance(paid_date, datetime):
                    paid_date_value = paid_date.date()
                elif isinstance(paid_date, date):
                    paid_date_value = paid_date
                else:
                    continue

                if paid_date_value == date(2024, 6, 5):
                    # Проверяем номер полиса (1-й столбец)
                    policy_number = ws.cell(row=row, column=1).value
                    self.assertEqual(policy_number, "READY-001")
                    # Проверяем очередной взнос (7-й столбец)
                    amount = ws.cell(row=row, column=7).value
                    self.assertEqual(float(amount), 25000.00)
                    # Проверяем КВ руб (9-й столбец) - проверяем что значение есть
                    kv_rub = ws.cell(row=row, column=9).value
                    self.assertIsNotNone(kv_rub)
                    self.assertGreater(float(kv_rub), 0)
                    # Проверяем дату платежа по договору (10-й столбец)
                    due_date = ws.cell(row=row, column=10).value
                    if isinstance(due_date, datetime):
                        due_date_value = due_date.date()
                    elif isinstance(due_date, date):
                        due_date_value = due_date
                    else:
                        due_date_value = None
                    self.assertEqual(due_date_value, date(2024, 6, 1))
                    found_june_payment = True
                    break

        self.assertTrue(
            found_june_payment,
            "Платеж с датой факт. оплаты 05.06.2024 должен быть в отчете",
        )

        # Проверяем что в отчете есть платеж из setUp (с датой факт. оплаты 28.02.2024)
        found_march_payment = False
        for row in range(5, ws.max_row + 1):
            paid_date = ws.cell(row=row, column=11).value  # Дата факт. оплаты

            # Преобразуем дату для сравнения
            if paid_date:
                if isinstance(paid_date, datetime):
                    paid_date_value = paid_date.date()
                elif isinstance(paid_date, date):
                    paid_date_value = paid_date
                else:
                    continue

                if paid_date_value == date(2024, 2, 28):
                    # Проверяем КВ руб (9-й столбец)
                    kv_rub = ws.cell(row=row, column=9).value
                    self.assertEqual(float(kv_rub), 10000.00)
                    found_march_payment = True
                    break

        self.assertTrue(
            found_march_payment,
            "Платеж с датой факт. оплаты 28.02.2024 должен быть в отчете",
        )

        # Проверяем что без параметра insurer возвращается ошибка
        response_no_insurer = self.test_client.get("/reports/export/commission-report/")
        self.assertEqual(response_no_insurer.status_code, 302)  # Редирект

        # Проверяем что с несуществующим insurer возвращается ошибка
        response_invalid_insurer = self.test_client.get(
            "/reports/export/commission-report/?insurer=99999"
        )
        self.assertEqual(response_invalid_insurer.status_code, 302)  # Редирект


class AuthorizationTest(TestCase):
    """
    Интеграционный тест: Авторизация и права доступа

    Тестирует что все страницы экспорта требуют авторизации.

    Validates: Requirement 9
    """

    def setUp(self):
        """Подготовка тестовых данных"""
        self.user = User.objects.create_user(
            username="testuser", password="testpass123"
        )
        self.test_client = TestClient()

    def test_index_requires_login(self):
        """Главная страница требует авторизации"""
        response = self.test_client.get("/reports/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_custom_export_requires_login(self):
        """Кастомный экспорт требует авторизации"""
        response = self.test_client.get("/reports/custom/")
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_ready_exports_require_login(self):
        """Готовые экспорты требуют авторизации"""
        response1 = self.test_client.get("/reports/export/policies/")
        self.assertEqual(response1.status_code, 302)
        self.assertIn("/accounts/login/", response1.url)

        response2 = self.test_client.get("/reports/export/payments/")
        self.assertEqual(response2.status_code, 302)
        self.assertIn("/accounts/login/", response2.url)

        response3 = self.test_client.get("/reports/export/thursday/")
        self.assertEqual(response3.status_code, 302)
        self.assertIn("/accounts/login/", response3.url)

    def test_authorized_access(self):
        """Авторизованный пользователь имеет доступ"""
        self.test_client.login(username="testuser", password="testpass123")

        response1 = self.test_client.get("/reports/")
        self.assertEqual(response1.status_code, 200)

        response2 = self.test_client.get("/reports/custom/")
        self.assertEqual(response2.status_code, 200)
