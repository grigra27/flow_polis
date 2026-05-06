from django.test import TestCase, override_settings
from django.contrib.auth.models import User
from django.urls import reverse
from pathlib import Path
from tempfile import TemporaryDirectory
from .models import CustomExportTemplate


class CustomExportTemplateModelTest(TestCase):
    """Тесты модели CustomExportTemplate"""

    def setUp(self):
        self.user = User.objects.create_user(
            username="testuser", password="testpass123"
        )

    def test_create_template(self):
        """Тест создания шаблона экспорта"""
        template = CustomExportTemplate.objects.create(
            user=self.user,
            name="Тестовый шаблон",
            data_source="policies",
            config={
                "fields": ["policy_number", "client__client_name"],
                "filters": {"branch": 1},
            },
        )

        self.assertEqual(template.name, "Тестовый шаблон")
        self.assertEqual(template.data_source, "policies")
        self.assertEqual(template.user, self.user)
        self.assertIsNotNone(template.created_at)
        self.assertIsNotNone(template.updated_at)

    def test_template_str(self):
        """Тест строкового представления шаблона"""
        template = CustomExportTemplate.objects.create(
            user=self.user, name="Мой шаблон", data_source="payments", config={}
        )

        self.assertEqual(str(template), "Мой шаблон (Платежи)")

    def test_unique_name_per_user(self):
        """Тест уникальности имени шаблона для пользователя"""
        CustomExportTemplate.objects.create(
            user=self.user, name="Уникальный шаблон", data_source="policies", config={}
        )

        # Попытка создать шаблон с тем же именем должна вызвать ошибку
        with self.assertRaises(Exception):
            CustomExportTemplate.objects.create(
                user=self.user,
                name="Уникальный шаблон",
                data_source="payments",
                config={},
            )

    def test_different_users_same_name(self):
        """Тест: разные пользователи могут иметь шаблоны с одинаковым именем"""
        user2 = User.objects.create_user(username="testuser2", password="testpass123")

        template1 = CustomExportTemplate.objects.create(
            user=self.user, name="Общий шаблон", data_source="policies", config={}
        )

        template2 = CustomExportTemplate.objects.create(
            user=user2, name="Общий шаблон", data_source="payments", config={}
        )

        self.assertNotEqual(template1.id, template2.id)
        self.assertEqual(template1.name, template2.name)


class BaseExporterTest(TestCase):
    """Тесты базового экспортера"""

    def setUp(self):
        from .exporters import BaseExporter

        # Создаем простой экспортер для тестирования
        class TestExporter(BaseExporter):
            def get_headers(self):
                return ["Заголовок 1", "Заголовок 2"]

            def get_row_data(self, obj):
                return [obj.get("field1"), obj.get("field2")]

            def get_filename(self):
                return "test_report"

        self.TestExporter = TestExporter

    def test_format_value_date(self):
        """Тест форматирования даты"""
        from datetime import date

        exporter = self.TestExporter([], [])

        test_date = date(2024, 1, 15)
        formatted = exporter.format_value(test_date)

        self.assertEqual(formatted, test_date)

    def test_format_value_decimal(self):
        """Тест форматирования Decimal"""
        from decimal import Decimal

        exporter = self.TestExporter([], [])

        test_decimal = Decimal("123.45")
        formatted = exporter.format_value(test_decimal)

        self.assertEqual(formatted, 123.45)
        self.assertIsInstance(formatted, float)

    def test_format_value_bool(self):
        """Тест форматирования boolean"""
        exporter = self.TestExporter([], [])

        self.assertEqual(exporter.format_value(True), "Да")
        self.assertEqual(exporter.format_value(False), "Нет")

    def test_format_value_none(self):
        """Тест форматирования None"""
        exporter = self.TestExporter([], [])

        self.assertEqual(exporter.format_value(None), "")

    def test_export_creates_response(self):
        """Тест создания HTTP ответа с Excel файлом"""
        test_data = [
            {"field1": "Значение 1", "field2": "Значение 2"},
            {"field1": "Значение 3", "field2": "Значение 4"},
        ]

        exporter = self.TestExporter(test_data, [])
        response = exporter.export()

        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("test_report_", response["Content-Disposition"])
        self.assertIn(".xlsx", response["Content-Disposition"])


class FilterTest(TestCase):
    """Тесты фильтров экспорта"""

    def setUp(self):
        from apps.clients.models import Client
        from apps.insurers.models import Insurer, Branch, InsuranceType
        from apps.policies.models import Policy
        from datetime import date

        # Создаем тестовые данные
        self.client1 = Client.objects.create(
            client_name="Тестовый клиент 1", client_inn="1234567890"
        )
        self.client2 = Client.objects.create(
            client_name="Другой клиент", client_inn="0987654321"
        )

        self.insurer = Insurer.objects.create(insurer_name="Тестовая СК")
        self.branch = Branch.objects.create(branch_name="Тестовый филиал")
        self.insurance_type = InsuranceType.objects.create(name="КАСКО")

        self.policy1 = Policy.objects.create(
            policy_number="POL-001",
            dfa_number="DFA-001",
            client=self.client1,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            property_description="Тестовое имущество",
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
            policy_active=True,
        )

        self.policy2 = Policy.objects.create(
            policy_number="POL-002",
            dfa_number="DFA-002",
            client=self.client2,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            property_description="Другое имущество",
            start_date=date(2024, 6, 1),
            end_date=date(2025, 5, 31),
            policy_active=False,
        )

    def test_policy_filter_by_number(self):
        """Тест фильтрации полисов по номеру"""
        from .filters import PolicyExportFilter
        from apps.policies.models import Policy

        f = PolicyExportFilter(
            {"policy_number": "POL-001"}, queryset=Policy.objects.all()
        )

        self.assertEqual(f.qs.count(), 1)
        self.assertEqual(f.qs.first(), self.policy1)

    def test_policy_filter_by_client_name(self):
        """Тест фильтрации полисов по имени клиента"""
        from .filters import PolicyExportFilter
        from apps.policies.models import Policy

        f = PolicyExportFilter(
            {"client__client_name": "Тестовый"}, queryset=Policy.objects.all()
        )

        self.assertEqual(f.qs.count(), 1)
        self.assertEqual(f.qs.first().client, self.client1)

    def test_policy_filter_by_active(self):
        """Тест фильтрации полисов по статусу активности"""
        from .filters import PolicyExportFilter
        from apps.policies.models import Policy

        f = PolicyExportFilter({"policy_active": True}, queryset=Policy.objects.all())

        self.assertEqual(f.qs.count(), 1)
        self.assertEqual(f.qs.first(), self.policy1)

    def test_policy_filter_by_date_range(self):
        """Тест фильтрации полисов по диапазону дат"""
        from .filters import PolicyExportFilter
        from apps.policies.models import Policy
        from datetime import date

        f = PolicyExportFilter(
            {"start_date_from": date(2024, 1, 1), "start_date_to": date(2024, 3, 31)},
            queryset=Policy.objects.all(),
        )

        self.assertEqual(f.qs.count(), 1)
        self.assertEqual(f.qs.first(), self.policy1)

    def test_client_filter_by_name(self):
        """Тест фильтрации клиентов по имени"""
        from .filters import ClientExportFilter
        from apps.clients.models import Client

        f = ClientExportFilter(
            {"client_name": "Тестовый"}, queryset=Client.objects.all()
        )

        self.assertEqual(f.qs.count(), 1)
        self.assertEqual(f.qs.first(), self.client1)

    def test_client_filter_by_inn(self):
        """Тест фильтрации клиентов по ИНН"""
        from .filters import ClientExportFilter
        from apps.clients.models import Client

        f = ClientExportFilter({"client_inn": "1234"}, queryset=Client.objects.all())

        self.assertEqual(f.qs.count(), 1)
        self.assertEqual(f.qs.first(), self.client1)

    def test_insurer_filter_by_name(self):
        """Тест фильтрации страховщиков по имени"""
        from .filters import InsurerExportFilter
        from apps.insurers.models import Insurer

        f = InsurerExportFilter(
            {"insurer_name": "Тестовая"}, queryset=Insurer.objects.all()
        )

        self.assertEqual(f.qs.count(), 1)
        self.assertEqual(f.qs.first(), self.insurer)


class CustomExporterTest(TestCase):
    """Тесты CustomExporter"""

    def setUp(self):
        from apps.clients.models import Client
        from apps.insurers.models import Insurer, Branch, InsuranceType
        from apps.policies.models import Policy, PaymentSchedule
        from datetime import date
        from decimal import Decimal

        # Создаем тестовые данные
        self.client = Client.objects.create(
            client_name="Тестовый клиент", client_inn="1234567890"
        )

        self.insurer = Insurer.objects.create(insurer_name="Тестовая СК")
        self.branch = Branch.objects.create(branch_name="Тестовый филиал")
        self.insurance_type = InsuranceType.objects.create(name="КАСКО")

        self.policy = Policy.objects.create(
            policy_number="POL-001",
            dfa_number="DFA-001",
            client=self.client,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            property_description="Тестовое имущество",
            start_date=date(2024, 1, 15),
            end_date=date(2024, 12, 31),
            policy_active=True,
        )
        PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=1,
            due_date=date(2024, 1, 20),
            insurance_sum=Decimal("1000000.00"),
            amount=Decimal("100000.00"),
            kv_rub=Decimal("0.00"),
        )

    def test_custom_exporter_get_headers(self):
        """Тест получения заголовков CustomExporter"""
        from .exporters import CustomExporter
        from apps.policies.models import Policy

        fields = ["policy_number", "client__client_name", "premium_total"]
        exporter = CustomExporter(Policy.objects.all(), fields, "policies")

        headers = exporter.get_headers()

        self.assertEqual(len(headers), 3)
        self.assertEqual(headers[0], "Номер полиса")
        self.assertEqual(headers[1], "Лизингополучатель")
        self.assertEqual(headers[2], "Общая премия")

    def test_custom_exporter_get_field_value(self):
        """Тест получения значения поля с вложенными полями"""
        from .exporters import CustomExporter
        from apps.policies.models import Policy

        exporter = CustomExporter(Policy.objects.all(), [], "policies")

        # Простое поле
        value = exporter.get_field_value(self.policy, "policy_number")
        self.assertEqual(value, "POL-001")

        # Вложенное поле
        value = exporter.get_field_value(self.policy, "client__client_name")
        self.assertEqual(value, "Тестовый клиент")

        # Двойное вложение
        value = exporter.get_field_value(self.policy, "insurer__insurer_name")
        self.assertEqual(value, "Тестовая СК")

    def test_custom_exporter_format_date(self):
        """Тест форматирования даты в CustomExporter"""
        from datetime import date
        from .exporters import CustomExporter
        from apps.policies.models import Policy

        exporter = CustomExporter(Policy.objects.all(), [], "policies")

        value = exporter.get_field_value(self.policy, "start_date")
        self.assertEqual(value, date(2024, 1, 15))

    def test_custom_exporter_format_decimal(self):
        """Тест форматирования Decimal в CustomExporter"""
        from .exporters import CustomExporter
        from apps.policies.models import Policy

        exporter = CustomExporter(Policy.objects.all(), [], "policies")

        value = exporter.get_field_value(self.policy, "premium_total")
        self.assertEqual(value, 100000.0)
        self.assertIsInstance(value, float)

    def test_custom_exporter_format_bool(self):
        """Тест форматирования boolean в CustomExporter"""
        from .exporters import CustomExporter
        from apps.policies.models import Policy

        exporter = CustomExporter(Policy.objects.all(), [], "policies")

        value = exporter.get_field_value(self.policy, "policy_active")
        self.assertEqual(value, "Да")

    def test_custom_exporter_get_row_data(self):
        """Тест получения данных строки"""
        from .exporters import CustomExporter
        from apps.policies.models import Policy

        fields = ["policy_number", "client__client_name", "policy_active"]
        exporter = CustomExporter(Policy.objects.all(), fields, "policies")

        row = exporter.get_row_data(self.policy)

        self.assertEqual(len(row), 3)
        self.assertEqual(row[0], "POL-001")
        self.assertEqual(row[1], "Тестовый клиент")
        self.assertEqual(row[2], "Да")

    def test_custom_exporter_filename(self):
        """Тест генерации имени файла"""
        from .exporters import CustomExporter
        from apps.policies.models import Policy

        exporter = CustomExporter(Policy.objects.all(), [], "policies")
        filename = exporter.get_filename()

        self.assertEqual(filename, "custom_export_policies")


class CustomExportViewTest(TestCase):
    """Тесты представления CustomExportView"""

    def setUp(self):
        from apps.clients.models import Client
        from apps.insurers.models import Insurer, Branch, InsuranceType
        from apps.policies.models import Policy
        from datetime import date

        self.user = User.objects.create_user(
            username="testuser", password="testpass123"
        )

        # Создаем тестовые данные
        self.client = Client.objects.create(
            client_name="Тестовый клиент", client_inn="1234567890"
        )

        self.insurer = Insurer.objects.create(insurer_name="Тестовая СК")
        self.branch = Branch.objects.create(branch_name="Тестовый филиал")
        self.insurance_type = InsuranceType.objects.create(name="КАСКО")

        self.policy = Policy.objects.create(
            policy_number="POL-001",
            dfa_number="DFA-001",
            client=self.client,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            property_description="Тестовое имущество",
            start_date=date(2024, 1, 15),
            end_date=date(2024, 12, 31),
            policy_active=True,
        )

    def test_custom_export_view_requires_login(self):
        """Тест: CustomExportView требует авторизации"""
        from django.test import Client as TestClient

        client = TestClient()
        response = client.get("/reports/custom/")

        # Должен перенаправить на страницу входа
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_custom_export_view_get(self):
        """Тест GET запроса к CustomExportView"""
        from django.test import Client as TestClient

        client = TestClient()
        client.login(username="testuser", password="testpass123")

        response = client.get("/reports/custom/")

        self.assertEqual(response.status_code, 200)
        self.assertIn("available_fields", response.context)
        self.assertIn("templates", response.context)

    def test_save_template(self):
        """Тест сохранения шаблона"""
        from django.test import Client as TestClient

        client = TestClient()
        client.login(username="testuser", password="testpass123")

        response = client.post(
            "/reports/custom/",
            {
                "action": "save_template",
                "template_name": "Мой шаблон",
                "data_source": "policies",
                "fields": ["policy_number", "client__client_name"],
            },
        )

        # Проверяем что шаблон создан
        template = CustomExportTemplate.objects.filter(
            user=self.user, name="Мой шаблон"
        ).first()

        self.assertIsNotNone(template)
        self.assertEqual(template.data_source, "policies")
        self.assertEqual(
            template.config["fields"], ["policy_number", "client__client_name"]
        )

    def test_export_without_fields(self):
        """Тест экспорта без выбранных полей"""
        from django.test import Client as TestClient

        client = TestClient()
        client.login(username="testuser", password="testpass123")

        response = client.post(
            "/reports/custom/",
            {"action": "export", "data_source": "policies", "fields": []},
        )

        # Должен перенаправить обратно с сообщением об ошибке
        self.assertEqual(response.status_code, 302)

    def test_export_with_fields(self):
        """Тест экспорта с выбранными полями"""
        from django.test import Client as TestClient

        client = TestClient()
        client.login(username="testuser", password="testpass123")

        response = client.post(
            "/reports/custom/",
            {
                "action": "export",
                "data_source": "policies",
                "fields": ["policy_number", "client__client_name"],
            },
        )

        # Должен вернуть Excel файл
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_optimize_queryset_uses_deep_select_related_for_payment_fields(self):
        """Тест: optimize_queryset добавляет deep select_related для вложенных payment-полей"""
        from datetime import date
        from decimal import Decimal
        from apps.policies.models import PaymentSchedule
        from .views import CustomExportView
        from .exporters import CustomExporter

        PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=1,
            due_date=date(2024, 2, 15),
            insurance_sum=Decimal("100000.00"),
            amount=Decimal("10000.00"),
            kv_rub=Decimal("1000.00"),
        )

        view = CustomExportView()
        fields = [
            "policy__client__client_name",
            "policy__insurer__insurer_name",
        ]
        queryset = view.optimize_queryset(PaymentSchedule.objects.all(), fields)
        exporter = CustomExporter(queryset, fields, "payments")

        # Один запрос на выборку payment + joins, без N+1 на policy/client/insurer
        with self.assertNumQueries(1):
            for payment in queryset:
                exporter.get_row_data(payment)

    def test_delete_template(self):
        """Тест удаления шаблона"""
        from django.test import Client as TestClient

        # Делаем пользователя staff для прохождения middleware
        self.user.is_staff = True
        self.user.save()

        # Создаем шаблон
        template = CustomExportTemplate.objects.create(
            user=self.user,
            name="Шаблон для удаления",
            data_source="policies",
            config={"fields": ["policy_number"]},
        )

        client = TestClient()
        client.login(username="testuser", password="testpass123")

        response = client.post(f"/reports/custom/template/{template.id}/delete/")

        # Проверяем что шаблон удален
        self.assertFalse(CustomExportTemplate.objects.filter(id=template.id).exists())


class PolicyExporterTest(TestCase):
    """Тесты PolicyExporter"""

    def setUp(self):
        """Подготовка тестовых данных"""
        from apps.clients.models import Client
        from apps.insurers.models import Insurer, Branch, InsuranceType
        from apps.policies.models import Policy
        from datetime import date
        from decimal import Decimal

        # Создаем клиента
        self.client = Client.objects.create(
            client_name="Тестовый клиент", client_inn="1234567890"
        )

        # Создаем страховщика
        self.insurer = Insurer.objects.create(insurer_name="Тестовый страховщик")

        # Создаем филиал
        self.branch = Branch.objects.create(branch_name="Главный офис")

        # Создаем вид страхования
        self.insurance_type = InsuranceType.objects.create(name="КАСКО")

        # Создаем полис
        self.policy = Policy.objects.create(
            policy_number="TEST-001",
            dfa_number="DFA-001",
            client=self.client,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            property_description="Тестовое имущество",
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
            policy_active=True,
        )

    def test_policy_exporter_get_headers(self):
        """Тест получения заголовков PolicyExporter"""
        from apps.policies.models import Policy
        from apps.reports.exporters import PolicyExporter

        queryset = Policy.objects.all()
        exporter = PolicyExporter(queryset, [])
        headers = exporter.get_headers()

        self.assertEqual(len(headers), 14)
        self.assertIn("Номер полиса", headers)
        self.assertIn("Клиент", headers)
        self.assertIn("Страховщик", headers)

    def test_policy_exporter_get_row_data(self):
        """Тест получения данных строки для полиса"""
        from datetime import date
        from apps.reports.exporters import PolicyExporter

        exporter = PolicyExporter([], [])
        row = exporter.get_row_data(self.policy)

        self.assertEqual(len(row), 14)
        self.assertEqual(row[0], "TEST-001")  # policy_number
        self.assertEqual(row[1], "DFA-001")  # dfa_number
        self.assertEqual(row[2], "Тестовый клиент")  # client
        self.assertEqual(row[3], "Тестовый страховщик")  # insurer
        self.assertEqual(row[6], date(2024, 1, 1))  # start_date
        self.assertEqual(row[10], "Активен")  # policy_active

    def test_policy_exporter_filename(self):
        """Тест генерации имени файла"""
        from apps.policies.models import Policy
        from apps.reports.exporters import PolicyExporter

        queryset = Policy.objects.all()
        exporter = PolicyExporter(queryset, [])

        self.assertEqual(exporter.get_filename(), "policies")

    def test_policy_exporter_export(self):
        """Тест полного экспорта полисов"""
        from apps.policies.models import Policy
        from apps.reports.exporters import PolicyExporter

        queryset = Policy.objects.select_related(
            "client", "insurer", "branch", "insurance_type"
        ).all()
        exporter = PolicyExporter(queryset, [])
        response = exporter.export()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("policies_", response["Content-Disposition"])


class PaymentExporterTest(TestCase):
    """Тесты PaymentExporter"""

    def setUp(self):
        """Подготовка тестовых данных"""
        from apps.clients.models import Client
        from apps.insurers.models import Insurer, Branch, InsuranceType, CommissionRate
        from apps.policies.models import Policy, PaymentSchedule
        from datetime import date
        from decimal import Decimal

        # Создаем клиента
        self.client = Client.objects.create(
            client_name="Тестовый клиент", client_inn="1234567890"
        )

        # Создаем страховщика
        self.insurer = Insurer.objects.create(insurer_name="Тестовый страховщик")

        # Создаем филиал
        self.branch = Branch.objects.create(branch_name="Главный офис")

        # Создаем вид страхования
        self.insurance_type = InsuranceType.objects.create(name="КАСКО")

        # Создаем ставку комиссии
        self.commission_rate = CommissionRate.objects.create(
            insurer=self.insurer,
            insurance_type=self.insurance_type,
            kv_percent=Decimal("15.00"),
        )

        # Создаем полис
        self.policy = Policy.objects.create(
            policy_number="TEST-001",
            dfa_number="DFA-001",
            client=self.client,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            property_description="Тестовое имущество",
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
            policy_active=True,
        )

        # Создаем платеж (дата в будущем, чтобы не был просрочен)
        from datetime import timedelta
        from django.utils import timezone

        future_date = timezone.now().date() + timedelta(days=30)

        # Создаем платеж без commission_rate, чтобы избежать автоматического пересчета
        self.payment = PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=1,
            due_date=future_date,
            amount=Decimal("50000.00"),
            insurance_sum=Decimal("1000000.00"),
            kv_rub=Decimal("7500.00"),
        )

    def test_payment_exporter_get_headers(self):
        """Тест получения заголовков PaymentExporter"""
        from apps.policies.models import PaymentSchedule
        from apps.reports.exporters import PaymentExporter

        queryset = PaymentSchedule.objects.all()
        exporter = PaymentExporter(queryset, [])
        headers = exporter.get_headers()

        self.assertEqual(len(headers), 12)
        self.assertIn("Номер полиса", headers)
        self.assertIn("Клиент", headers)
        self.assertIn("Статус", headers)

    def test_payment_exporter_get_row_data(self):
        """Тест получения данных строки для платежа"""
        from apps.reports.exporters import PaymentExporter

        exporter = PaymentExporter([], [])
        row = exporter.get_row_data(self.payment)

        self.assertEqual(len(row), 12)
        self.assertEqual(row[0], "TEST-001")  # policy_number
        self.assertEqual(row[1], "Тестовый клиент")  # client
        self.assertEqual(row[2], 1)  # year_number
        self.assertEqual(row[3], 1)  # installment_number
        # Не проверяем точную дату, так как она динамическая
        self.assertEqual(row[11], "Ожидается")  # status

    def test_payment_exporter_status_paid(self):
        """Тест статуса 'Оплачен'"""
        from apps.reports.exporters import PaymentExporter
        from datetime import date

        # Обновляем только поле paid_date без вызова save() чтобы избежать пересчета kv_rub
        from apps.policies.models import PaymentSchedule

        PaymentSchedule.objects.filter(id=self.payment.id).update(
            paid_date=date(2024, 2, 1)
        )
        self.payment.refresh_from_db()

        exporter = PaymentExporter([], [])
        row = exporter.get_row_data(self.payment)

        self.assertEqual(row[11], "Оплачен")

    def test_payment_exporter_status_cancelled(self):
        """Тест статуса 'Отменен'"""
        from apps.reports.exporters import PaymentExporter
        from datetime import date

        # Делаем полис неактивным и устанавливаем дату расторжения
        self.policy.policy_active = False
        self.policy.termination_date = date(2024, 1, 15)
        self.policy.save()

        exporter = PaymentExporter([], [])
        row = exporter.get_row_data(self.payment)

        self.assertEqual(row[11], "Отменен")

    def test_payment_exporter_filename(self):
        """Тест генерации имени файла"""
        from apps.policies.models import PaymentSchedule
        from apps.reports.exporters import PaymentExporter

        queryset = PaymentSchedule.objects.all()
        exporter = PaymentExporter(queryset, [])

        self.assertEqual(exporter.get_filename(), "payments")

    def test_payment_exporter_export(self):
        """Тест полного экспорта платежей"""
        from apps.policies.models import PaymentSchedule
        from apps.reports.exporters import PaymentExporter

        queryset = PaymentSchedule.objects.select_related(
            "policy", "policy__client", "policy__insurer", "commission_rate"
        ).all()
        exporter = PaymentExporter(queryset, [])
        response = exporter.export()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("payments_", response["Content-Disposition"])


class MonthlyKVExportViewTest(TestCase):
    """Тесты экспорта «КВ за месяц»."""

    def setUp(self):
        from datetime import date
        from decimal import Decimal
        from apps.clients.models import Client
        from apps.insurers.models import Insurer, Branch, InsuranceType
        from apps.policies.models import Policy, PaymentSchedule

        self.user = User.objects.create_user(
            username="monthly_kv_user", password="testpass123"
        )
        self.admin = User.objects.create_user(
            username="monthly_kv_admin", password="testpass123", is_staff=True
        )

        self.client_obj = Client.objects.create(
            client_name="Лизингополучатель", client_inn="7700000001"
        )
        self.policyholder = Client.objects.create(
            client_name="Страхователь", client_inn="7700000002"
        )
        self.insurer = Insurer.objects.create(insurer_name="СК Тест")
        self.branch = Branch.objects.create(branch_name="Москва")
        self.insurance_type = InsuranceType.objects.create(name="КАСКО")

        self.policy = Policy.objects.create(
            policy_number="POL-KV-001",
            dfa_number="DFA-KV-001",
            client=self.client_obj,
            policyholder=self.policyholder,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            property_description="Автомобиль",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 12, 31),
            policy_active=True,
        )

        PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=1,
            due_date=date(2025, 2, 10),
            amount=Decimal("10000.00"),
            insurance_sum=Decimal("500000.00"),
            kv_rub=Decimal("1500.00"),
            paid_date=date(2025, 2, 15),
            insurer_date=date(2025, 2, 20),
        )

        PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=2,
            due_date=date(2025, 3, 10),
            amount=Decimal("12000.00"),
            insurance_sum=Decimal("500000.00"),
            kv_rub=Decimal("1800.00"),
            paid_date=date(2025, 3, 15),
            insurer_date=date(2025, 3, 20),
        )

    def test_non_admin_cannot_export_monthly_kv(self):
        """Тест: обычный пользователь не может выгрузить КВ за месяц."""
        self.client.login(username="monthly_kv_user", password="testpass123")

        response = self.client.get(
            reverse("reports:export_monthly_kv_report"),
            {"kv_month": 2, "kv_year": 2025},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("reports:index"))

    def test_admin_can_export_monthly_kv(self):
        """Тест: админ получает Excel с нужными колонками и данными."""
        from datetime import date
        from io import BytesIO
        from openpyxl import load_workbook

        self.client.login(username="monthly_kv_admin", password="testpass123")

        response = self.client.get(
            reverse("reports:export_monthly_kv_report"),
            {"kv_month": 2, "kv_year": 2025},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("kv_for_month_2025_02_", response["Content-Disposition"])

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(sheet["A1"].value, "КВ ЗА МЕСЯЦ - 02.2025")
        self.assertEqual(
            [cell.value for cell in sheet[3]],
            [
                "Страховщик",
                "Номер полиса",
                "Номер ДФА",
                "Страхователь",
                "КВ в %",
                "Страховая премия",
                "Дата фактической оплаты",
                "КВ (в руб)",
                "Филиал",
            ],
        )

        data_rows = [
            row
            for row in sheet.iter_rows(min_row=5, max_col=9, values_only=True)
            if any(value is not None and value != "" for value in row)
        ]

        self.assertEqual(len(data_rows), 1)
        row = data_rows[0]

        self.assertEqual(row[0], "СК Тест")
        self.assertEqual(row[1], "POL-KV-001")
        self.assertEqual(row[2], "DFA-KV-001")
        self.assertEqual(row[3], "Страхователь")
        self.assertAlmostEqual(float(row[4]), 15.0)
        self.assertAlmostEqual(float(row[5]), 10000.0)
        exported_paid_date = row[6].date() if hasattr(row[6], "date") else row[6]
        self.assertEqual(exported_paid_date, date(2025, 2, 15))
        self.assertAlmostEqual(float(row[7]), 1500.0)
        self.assertEqual(row[8], "Москва")

    def test_admin_gets_redirect_when_no_data(self):
        """Тест: при отсутствии данных админ возвращается на страницу экспортов."""
        self.client.login(username="monthly_kv_admin", password="testpass123")

        response = self.client.get(
            reverse("reports:export_monthly_kv_report"),
            {"kv_month": 1, "kv_year": 2025},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("reports:index"))


class ThreePercentReportExportViewTest(TestCase):
    """Тесты экспорта «Отчет по 3%»."""

    def setUp(self):
        from datetime import date
        from decimal import Decimal
        from apps.clients.models import Client
        from apps.insurers.models import Branch, CommissionRate, InsuranceType, Insurer
        from apps.policies.models import Policy, PaymentSchedule

        self.user = User.objects.create_user(
            username="three_percent_user", password="testpass123"
        )
        self.admin = User.objects.create_user(
            username="three_percent_admin", password="testpass123", is_staff=True
        )

        self.client_obj = Client.objects.create(
            client_name="Лизингополучатель 3%", client_inn="7700000011"
        )
        self.policyholder = Client.objects.create(
            client_name="Страхователь 3%", client_inn="7700000012"
        )
        self.insurer = Insurer.objects.create(insurer_name="СК 3%")
        self.branch = Branch.objects.create(branch_name="Санкт-Петербург")
        self.insurance_type = InsuranceType.objects.create(name="ОСАГО")
        self.commission_rate = CommissionRate.objects.create(
            insurer=self.insurer,
            insurance_type=self.insurance_type,
            kv_percent=Decimal("15.00"),
        )

        self.policy = Policy.objects.create(
            policy_number="POL-3P-001",
            dfa_number="DFA-3P-001",
            client=self.client_obj,
            policyholder=self.policyholder,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            property_description="Тестовое имущество для 3%",
            start_date=date(2025, 1, 1),
            end_date=date(2025, 12, 31),
            policy_active=True,
            info3="Тестовое значение Инфо 3",
        )

        PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=1,
            due_date=date(2025, 2, 10),
            amount=Decimal("10000.00"),
            insurance_sum=Decimal("500000.00"),
            commission_rate=self.commission_rate,
            kv_rub=Decimal("1500.00"),
            paid_date=date(2025, 2, 15),
            insurer_date=date(2025, 2, 20),
        )

        PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=2,
            due_date=date(2025, 3, 10),
            amount=Decimal("12000.00"),
            insurance_sum=Decimal("500000.00"),
            commission_rate=self.commission_rate,
            kv_rub=Decimal("1800.00"),
            paid_date=date(2025, 3, 15),
            insurer_date=None,
        )

        PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=3,
            due_date=date(2025, 4, 10),
            amount=Decimal("14000.00"),
            insurance_sum=Decimal("500000.00"),
            commission_rate=self.commission_rate,
            kv_rub=Decimal("2100.00"),
            paid_date=date(2025, 4, 15),
            insurer_date=date(2025, 4, 20),
        )

    def test_non_admin_cannot_export_three_percent_report(self):
        """Тест: обычный пользователь не может выгрузить отчет по 3%."""
        self.client.login(username="three_percent_user", password="testpass123")

        response = self.client.get(
            reverse("reports:export_three_percent_report"),
            {"three_percent_quarter": 1, "three_percent_year": 2025},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("reports:index"))

    def test_admin_can_export_three_percent_report_for_quarter(self):
        """Тест: админ получает Excel только с согласованными платежами квартала."""
        from datetime import date
        from io import BytesIO
        from openpyxl import load_workbook

        self.client.login(username="three_percent_admin", password="testpass123")

        response = self.client.get(
            reverse("reports:export_three_percent_report"),
            {"three_percent_quarter": 1, "three_percent_year": 2025},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("three_percent_report_2025_q1_", response["Content-Disposition"])

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active

        self.assertEqual(
            sheet["A1"].value,
            (
                "Расчет комиссионного вознаграждения ЗАО «Альянс-Лизинг» "
                "за первый квартал 2025 года"
            ),
        )
        self.assertEqual(
            [cell.value for cell in sheet[3]],
            [
                "Номер ДФА",
                "Номер полиса",
                "Лизингополучатель",
                "Страхователь",
                "Страховая премия",
                "Дата фактической оплаты",
                "Ставка вознаграждения в процентах",
                "Сумма вознаграждения в рублях",
                "Страховщик",
                "Филиал",
                "КВ в процентах",
                "КВ в рублях",
                "Дата согласования акта с СК",
                "Инфо 3",
            ],
        )

        data_rows = [
            row
            for row in sheet.iter_rows(min_row=5, max_col=14, values_only=True)
            if any(value is not None and value != "" for value in row)
        ]

        self.assertEqual(len(data_rows), 1)
        row = data_rows[0]

        self.assertEqual(row[0], "DFA-3P-001")
        self.assertEqual(row[1], "POL-3P-001")
        self.assertEqual(row[2], "Лизингополучатель 3%")
        self.assertEqual(row[3], "Страхователь 3%")
        self.assertAlmostEqual(float(row[4]), 10000.0)
        exported_paid_date = row[5].date() if hasattr(row[5], "date") else row[5]
        self.assertEqual(exported_paid_date, date(2025, 2, 15))
        self.assertEqual(row[6], 3)
        self.assertAlmostEqual(float(row[7]), 300.0)
        self.assertEqual(row[8], "СК 3%")
        self.assertEqual(row[9], "Санкт-Петербург")
        self.assertAlmostEqual(float(row[10]), 15.0)
        self.assertAlmostEqual(float(row[11]), 1500.0)
        exported_insurer_date = row[12].date() if hasattr(row[12], "date") else row[12]
        self.assertEqual(exported_insurer_date, date(2025, 2, 20))
        self.assertEqual(row[13], "Тестовое значение Инфо 3")

    def test_three_percent_report_highlights_nonstandard_kv_rows(self):
        """Тест: строки с КВ не по базовой ставке выделяются в Excel."""
        from datetime import date
        from decimal import Decimal
        from io import BytesIO
        from openpyxl import load_workbook
        from apps.policies.models import PaymentSchedule

        PaymentSchedule.objects.filter(
            policy=self.policy,
            year_number=1,
            installment_number=2,
        ).update(
            insurer_date=date(2025, 3, 20),
            kv_rub=Decimal("1000.00"),
        )

        self.client.login(username="three_percent_admin", password="testpass123")

        response = self.client.get(
            reverse("reports:export_three_percent_report"),
            {"three_percent_quarter": 1, "three_percent_year": 2025},
        )

        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(filename=BytesIO(response.content))
        sheet = workbook.active

        standard_cell = sheet.cell(row=5, column=1)
        nonstandard_cell = sheet.cell(row=6, column=1)

        standard_fill_rgb = (
            standard_cell.fill.start_color.rgb
            if standard_cell.fill and standard_cell.fill.start_color
            else ""
        )
        nonstandard_fill_rgb = (
            nonstandard_cell.fill.start_color.rgb
            if nonstandard_cell.fill and nonstandard_cell.fill.start_color
            else ""
        )
        nonstandard_font_rgb = (
            nonstandard_cell.font.color.rgb
            if nonstandard_cell.font and nonstandard_cell.font.color
            else ""
        )

        self.assertEqual(sheet.cell(row=6, column=1).value, "DFA-3P-001")
        self.assertAlmostEqual(float(sheet.cell(row=6, column=11).value), 15.0)
        self.assertAlmostEqual(float(sheet.cell(row=6, column=12).value), 1000.0)
        self.assertFalse(str(standard_fill_rgb).upper().endswith("7A1F2B"))
        self.assertTrue(str(nonstandard_fill_rgb).upper().endswith("7A1F2B"))
        self.assertTrue(str(nonstandard_font_rgb).upper().endswith("FFFFFF"))

    def test_admin_gets_redirect_when_no_three_percent_data(self):
        """Тест: при отсутствии согласованных платежей за квартал админ возвращается."""
        self.client.login(username="three_percent_admin", password="testpass123")

        response = self.client.get(
            reverse("reports:export_three_percent_report"),
            {"three_percent_quarter": 3, "three_percent_year": 2025},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("reports:index"))


class BackupExportViewTest(TestCase):
    """Тесты скачивания backup-файлов на странице экспортов."""

    def setUp(self):
        self.user = User.objects.create_user(
            username="regular_user", password="testpass123"
        )
        self.admin = User.objects.create_user(
            username="admin_user", password="testpass123", is_staff=True
        )

    def _prepare_backup_files(self, root_dir):
        db_dir = Path(root_dir) / "database"
        media_dir = Path(root_dir) / "media"
        db_dir.mkdir(parents=True, exist_ok=True)
        media_dir.mkdir(parents=True, exist_ok=True)

        db_backup = db_dir / "latest_backup.sql.gz"
        media_backup = media_dir / "latest_backup.tar.gz"
        db_backup.write_bytes(b"db-backup-content")
        media_backup.write_bytes(b"media-backup-content")

        return db_backup, media_backup

    def test_backup_download_requires_login(self):
        """Тест: скачивание backup требует авторизации."""
        with TemporaryDirectory() as temp_dir:
            self._prepare_backup_files(temp_dir)
            with override_settings(BACKUP_BASE_DIR=temp_dir):
                response = self.client.get(reverse("reports:export_database_backup"))

        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_non_admin_cannot_download_backup(self):
        """Тест: обычный пользователь не может скачать backup."""
        self.client.login(username="regular_user", password="testpass123")
        with TemporaryDirectory() as temp_dir:
            self._prepare_backup_files(temp_dir)
            with override_settings(BACKUP_BASE_DIR=temp_dir):
                response = self.client.get(reverse("reports:export_database_backup"))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("reports:index"))

    def test_admin_can_download_database_backup(self):
        """Тест: админ может скачать backup базы данных."""
        self.client.login(username="admin_user", password="testpass123")
        with TemporaryDirectory() as temp_dir:
            db_backup, _ = self._prepare_backup_files(temp_dir)
            with override_settings(BACKUP_BASE_DIR=temp_dir):
                response = self.client.get(reverse("reports:export_database_backup"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment;", response["Content-Disposition"])
        self.assertIn(db_backup.name, response["Content-Disposition"])
        self.assertEqual(b"".join(response.streaming_content), b"db-backup-content")

    def test_admin_can_download_media_backup(self):
        """Тест: админ может скачать backup media."""
        self.client.login(username="admin_user", password="testpass123")
        with TemporaryDirectory() as temp_dir:
            _, media_backup = self._prepare_backup_files(temp_dir)
            with override_settings(BACKUP_BASE_DIR=temp_dir):
                response = self.client.get(reverse("reports:export_media_backup"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment;", response["Content-Disposition"])
        self.assertIn(media_backup.name, response["Content-Disposition"])
        self.assertEqual(b"".join(response.streaming_content), b"media-backup-content")

    def test_admin_exports_page_contains_backup_context(self):
        """Тест: для админа в контексте страницы экспортов есть данные по backup."""
        self.client.login(username="admin_user", password="testpass123")
        with TemporaryDirectory() as temp_dir:
            self._prepare_backup_files(temp_dir)
            with override_settings(BACKUP_BASE_DIR=temp_dir):
                response = self.client.get(reverse("reports:index"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("database_backup_info", response.context)
        self.assertIn("media_backup_info", response.context)
        self.assertTrue(response.context["database_backup_info"]["available"])
        self.assertTrue(response.context["media_backup_info"]["available"])
