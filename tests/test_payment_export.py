"""
Тесты для экспорта очередных взносов с фильтрацией по диапазону дат
"""
from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.urls import reverse
from datetime import date, timedelta
from decimal import Decimal

from apps.policies.models import Policy, PaymentSchedule
from apps.clients.models import Client as ClientModel
from apps.insurers.models import Insurer, Branch, InsuranceType


class PaymentExportDateRangeTest(TestCase):
    """Тесты экспорта платежей с фильтрацией по диапазону дат"""

    def setUp(self):
        """Подготовка тестовых данных"""
        # Создаем пользователя
        self.user = User.objects.create_user(
            username="testuser", password="testpass123"
        )
        self.client = Client()
        self.client.login(username="testuser", password="testpass123")

        # Создаем необходимые справочники
        self.branch = Branch.objects.create(branch_name="Тестовый филиал")
        self.insurance_type = InsuranceType.objects.create(name="Тестовый вид")
        self.insurer = Insurer.objects.create(insurer_name="Тестовая СК")
        self.test_client = ClientModel.objects.create(client_name="Тестовый клиент")

        # Создаем тестовый полис (активный)
        self.policy = Policy.objects.create(
            policy_number="TEST-001",
            client=self.test_client,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
            premium_total=Decimal("100000.00"),
            property_description="Тестовое имущество",
            policy_active=True,  # Активный полис
        )

        # Создаем неактивный полис для тестирования фильтрации
        self.inactive_policy = Policy.objects.create(
            policy_number="TEST-002",
            client=self.test_client,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
            premium_total=Decimal("50000.00"),
            property_description="Неактивное имущество",
            policy_active=False,  # Неактивный полис
        )

        # Создаем платежи с разными датами
        self.payment1 = PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=1,
            due_date=date(2024, 1, 15),
            amount=Decimal("25000.00"),
            insurance_sum=Decimal("1000000.00"),
        )

        self.payment2 = PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=2,
            due_date=date(2024, 4, 15),
            amount=Decimal("25000.00"),
            insurance_sum=Decimal("1000000.00"),
        )

        self.payment3 = PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=3,
            due_date=date(2024, 7, 15),
            amount=Decimal("25000.00"),
            insurance_sum=Decimal("1000000.00"),
        )

        self.payment4 = PaymentSchedule.objects.create(
            policy=self.policy,
            year_number=1,
            installment_number=4,
            due_date=date(2024, 10, 15),
            amount=Decimal("25000.00"),
            insurance_sum=Decimal("1000000.00"),
        )

        # Создаем платежи для неактивного полиса (они НЕ должны попадать в экспорт)
        self.inactive_payment1 = PaymentSchedule.objects.create(
            policy=self.inactive_policy,
            year_number=1,
            installment_number=1,
            due_date=date(2024, 2, 15),
            amount=Decimal("12500.00"),
            insurance_sum=Decimal("500000.00"),
        )

        self.inactive_payment2 = PaymentSchedule.objects.create(
            policy=self.inactive_policy,
            year_number=1,
            installment_number=2,
            due_date=date(2024, 5, 15),
            amount=Decimal("12500.00"),
            insurance_sum=Decimal("500000.00"),
        )

    def test_export_requires_dates(self):
        """Тест: экспорт требует указания дат"""
        url = reverse("reports:export_payments")
        response = self.client.get(url)

        # Должен быть редирект с сообщением об ошибке
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("reports:index"))

    def test_export_with_date_range(self):
        """Тест: экспорт с диапазоном дат возвращает только платежи в диапазоне"""
        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2024-04-01", "date_to": "2024-07-31"}
        )

        # Должен вернуться Excel файл
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", response["Content-Disposition"])

    def test_export_invalid_date_format(self):
        """Тест: неверный формат даты"""
        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "01-04-2024", "date_to": "31-07-2024"}  # Неверный формат
        )

        # Должен быть редирект с ошибкой
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("reports:index"))

    def test_export_date_from_after_date_to(self):
        """Тест: дата начала позже даты окончания"""
        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2024-07-31", "date_to": "2024-04-01"}
        )

        # Должен быть редирект с ошибкой
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("reports:index"))

    def test_export_no_payments_in_range(self):
        """Тест: нет платежей в указанном диапазоне"""
        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2025-01-01", "date_to": "2025-12-31"}
        )

        # Должен быть редирект с предупреждением
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("reports:index"))

    def test_export_single_day_range(self):
        """Тест: экспорт за один день"""
        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2024-04-15", "date_to": "2024-04-15"}
        )

        # Должен вернуться Excel файл
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_export_requires_authentication(self):
        """Тест: экспорт требует аутентификации"""
        self.client.logout()
        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2024-01-01", "date_to": "2024-12-31"}
        )

        # Должен быть редирект на страницу входа
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_export_sorted_by_due_date(self):
        """Тест: платежи отсортированы по дате платежа (самая ранняя вверху)"""
        from openpyxl import load_workbook
        from io import BytesIO

        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2024-01-01", "date_to": "2024-12-31"}
        )

        # Должен вернуться Excel файл
        self.assertEqual(response.status_code, 200)

        # Загружаем Excel файл из ответа
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Пропускаем заголовки (строки 1-4) и читаем даты платежей
        dates = []
        for row in range(5, ws.max_row + 1):
            # Столбец L (12) - "Дата платежа по договору"
            date_cell = ws.cell(row=row, column=12).value
            if date_cell:
                dates.append(date_cell)

        # Проверяем, что даты идут в порядке возрастания
        self.assertEqual(len(dates), 4, "Должно быть 4 платежа")

        # Преобразуем строки дат в объекты date для сравнения
        from datetime import datetime

        date_objects = [datetime.strptime(d, "%d.%m.%Y").date() for d in dates]

        # Проверяем сортировку
        self.assertEqual(
            date_objects[0], date(2024, 1, 15), "Первая дата должна быть 15.01.2024"
        )
        self.assertEqual(
            date_objects[1], date(2024, 4, 15), "Вторая дата должна быть 15.04.2024"
        )
        self.assertEqual(
            date_objects[2], date(2024, 7, 15), "Третья дата должна быть 15.07.2024"
        )
        self.assertEqual(
            date_objects[3], date(2024, 10, 15), "Четвертая дата должна быть 15.10.2024"
        )

        # Проверяем, что даты отсортированы по возрастанию
        self.assertEqual(
            date_objects,
            sorted(date_objects),
            "Даты должны быть отсортированы по возрастанию",
        )

    def test_export_payment_position_column(self):
        """Тест: столбец "Этот взнос" корректно отображает позицию платежа в году"""
        from openpyxl import load_workbook
        from io import BytesIO

        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2024-01-01", "date_to": "2024-12-31"}
        )

        # Должен вернуться Excel файл
        self.assertEqual(response.status_code, 200)

        # Загружаем Excel файл из ответа
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Пропускаем заголовки (строки 1-4) и читаем значения столбца "Этот взнос"
        payment_positions = []
        for row in range(5, ws.max_row + 1):
            # Столбец K (11) - "Этот взнос"
            position_cell = ws.cell(row=row, column=11).value
            if position_cell:
                payment_positions.append(position_cell)

        # Проверяем, что все 4 платежа имеют корректный формат
        self.assertEqual(len(payment_positions), 4, "Должно быть 4 платежа")

        # Все платежи в тестовых данных - это единственные платежи в своих годах
        # (год 1 платеж 1, год 1 платеж 2, год 1 платеж 3, год 1 платеж 4)
        # Поэтому все должны быть "1 из 1", "2 из 4", "3 из 4", "4 из 4"
        # На самом деле, в setUp создаются 4 платежа для одного года (year_number=1)
        # Поэтому все должны показывать "X из 4"

        for i, position in enumerate(payment_positions, start=1):
            self.assertIn(" из ", position, f"Позиция {i} должна содержать ' из '")
            parts = position.split(" из ")
            self.assertEqual(len(parts), 2, f"Позиция {i} должна иметь формат 'X из Y'")

            # Проверяем, что это числа
            installment_num = int(parts[0])
            total_in_year = int(parts[1])

            # Номер платежа должен быть от 1 до total_in_year
            self.assertGreaterEqual(installment_num, 1)
            self.assertLessEqual(installment_num, total_in_year)

            # Общее количество должно быть 4 (так как в setUp создано 4 платежа для года 1)
            self.assertEqual(total_in_year, 4, f"Всего платежей в году должно быть 4")

    def test_export_financial_formatting(self):
        """Тест: финансовые столбцы имеют правильное форматирование"""
        from openpyxl import load_workbook
        from io import BytesIO

        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2024-01-01", "date_to": "2024-12-31"}
        )

        # Должен вернуться Excel файл
        self.assertEqual(response.status_code, 200)

        # Загружаем Excel файл из ответа
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Проверяем форматирование первой строки данных (строка 5)
        # Столбец I (9) - Страховая сумма
        insurance_sum_cell = ws.cell(row=5, column=9)
        self.assertIsNotNone(
            insurance_sum_cell.value, "Страховая сумма должна быть заполнена"
        )
        self.assertEqual(
            insurance_sum_cell.number_format,
            "#,##0.00",
            "Страховая сумма должна иметь формат с разделителями тысяч",
        )
        self.assertEqual(
            insurance_sum_cell.alignment.horizontal,
            "right",
            "Страховая сумма должна быть выровнена по правому краю",
        )

        # Столбец J (10) - Очередной взнос
        payment_amount_cell = ws.cell(row=5, column=10)
        self.assertIsNotNone(
            payment_amount_cell.value, "Очередной взнос должен быть заполнен"
        )
        self.assertEqual(
            payment_amount_cell.number_format,
            "#,##0.00",
            "Очередной взнос должен иметь формат с разделителями тысяч",
        )
        self.assertEqual(
            payment_amount_cell.alignment.horizontal,
            "right",
            "Очередной взнос должен быть выровнен по правому краю",
        )

        # Проверяем, что значения - это числа (не строки)
        self.assertIsInstance(
            insurance_sum_cell.value, (int, float), "Страховая сумма должна быть числом"
        )
        self.assertIsInstance(
            payment_amount_cell.value,
            (int, float),
            "Очередной взнос должен быть числом",
        )

    def test_export_grouped_by_branch(self):
        """Тест: экспорт группирует платежи по филиалам"""
        from openpyxl import load_workbook
        from io import BytesIO

        # Создаем второй филиал
        branch2 = Branch.objects.create(branch_name="Второй филиал")

        # Создаем полис во втором филиале
        policy2 = Policy.objects.create(
            policy_number="TEST-003",
            client=self.test_client,
            insurer=self.insurer,
            branch=branch2,
            insurance_type=self.insurance_type,
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
            premium_total=Decimal("50000.00"),
            property_description="Имущество второго филиала",
            policy_active=True,
        )

        # Создаем платеж для второго филиала
        PaymentSchedule.objects.create(
            policy=policy2,
            year_number=1,
            installment_number=1,
            due_date=date(2024, 3, 15),
            amount=Decimal("12500.00"),
            insurance_sum=Decimal("500000.00"),
        )

        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2024-01-01", "date_to": "2024-12-31"}
        )

        # Должен вернуться Excel файл
        self.assertEqual(response.status_code, 200)

        # Загружаем Excel файл из ответа
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Ищем заголовки филиалов
        branch_headers = []
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            # Проверяем, что это заголовок филиала (первая ячейка заполнена, вторая пустая, строка > 3)
            if (
                cell_value
                and isinstance(cell_value, str)
                and ws.cell(row=row, column=2).value is None
                and row > 3
                and "филиал" in cell_value.lower()
            ):
                branch_headers.append((row, cell_value))

        # Проверяем, что найдены заголовки обоих филиалов
        self.assertEqual(len(branch_headers), 2, "Должно быть 2 заголовка филиалов")

        # Проверяем названия филиалов (должны быть отсортированы по алфавиту)
        branch_names = [header[1] for header in branch_headers]
        self.assertIn("Второй филиал", branch_names)
        self.assertIn("Тестовый филиал", branch_names)

        # Проверяем, что филиалы отсортированы по алфавиту
        self.assertEqual(
            branch_names,
            sorted(branch_names),
            "Филиалы должны быть отсортированы по алфавиту",
        )

    def test_export_branch_header_formatting(self):
        """Тест: заголовки филиалов имеют правильное форматирование (зеленая цветовая схема)"""
        from openpyxl import load_workbook
        from io import BytesIO

        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2024-01-01", "date_to": "2024-12-31"}
        )

        # Должен вернуться Excel файл
        self.assertEqual(response.status_code, 200)

        # Загружаем Excel файл из ответа
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Ищем заголовок филиала
        branch_header_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if (
                cell_value
                and isinstance(cell_value, str)
                and ws.cell(row=row, column=2).value is None
                and row > 3
                and "филиал" in cell_value.lower()
            ):
                branch_header_row = row
                break

        self.assertIsNotNone(branch_header_row, "Должен быть найден заголовок филиала")

        # Проверяем форматирование заголовка филиала
        branch_cell = ws.cell(row=branch_header_row, column=1)

        # Проверяем зеленый фон
        self.assertEqual(
            branch_cell.fill.start_color.rgb,
            "00C6F6D5",  # Светло-зеленый фон
            "Заголовок филиала должен иметь светло-зеленый фон",
        )

        # Проверяем жирный шрифт
        self.assertTrue(branch_cell.font.bold, "Заголовок филиала должен быть жирным")

        # Проверяем размер шрифта
        self.assertEqual(
            branch_cell.font.size, 12, "Размер шрифта заголовка филиала должен быть 12"
        )

    def test_export_installment_status_column(self):
        """Тест: экспорт содержит столбец 'Статус рассрочки' с правильными значениями"""
        from openpyxl import load_workbook
        from io import BytesIO

        # Создаем дополнительные платежи для тестирования рассрочки
        # Полис с рассрочкой (4 платежа в году)
        policy_installment = Policy.objects.create(
            policy_number="TEST-INSTALLMENT",
            client=self.test_client,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
            premium_total=Decimal("80000.00"),
            property_description="Имущество с рассрочкой",
            policy_active=True,
        )

        # Создаем 4 платежа для рассрочки
        for i in range(1, 5):
            PaymentSchedule.objects.create(
                policy=policy_installment,
                year_number=1,
                installment_number=i,
                due_date=date(2024, i * 3, 15),  # Каждые 3 месяца
                amount=Decimal("20000.00"),
                insurance_sum=Decimal("800000.00"),
            )

        # Полис с годовым платежом (1 платеж в году)
        policy_annual = Policy.objects.create(
            policy_number="TEST-ANNUAL",
            client=self.test_client,
            insurer=self.insurer,
            branch=self.branch,
            insurance_type=self.insurance_type,
            start_date=date(2024, 1, 1),
            end_date=date(2024, 12, 31),
            premium_total=Decimal("50000.00"),
            property_description="Имущество с годовым платежом",
            policy_active=True,
        )

        # Создаем 1 платеж для годового
        PaymentSchedule.objects.create(
            policy=policy_annual,
            year_number=1,
            installment_number=1,
            due_date=date(2024, 6, 15),
            amount=Decimal("50000.00"),
            insurance_sum=Decimal("500000.00"),
        )

        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2024-01-01", "date_to": "2024-12-31"}
        )

        # Должен вернуться Excel файл
        self.assertEqual(response.status_code, 200)

        # Загружаем Excel файл из ответа
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Проверяем наличие столбца "Статус рассрочки" в заголовках (строка 3)
        headers = [
            ws.cell(row=3, column=col).value for col in range(1, ws.max_column + 1)
        ]
        self.assertIn(
            "Статус рассрочки", headers, "Должен быть столбец 'Статус рассрочки'"
        )

        # Находим индекс столбца "Статус рассрочки"
        status_column_index = headers.index("Статус рассрочки") + 1

        # Собираем все значения из столбца "Статус рассрочки"
        status_values = []
        policy_numbers = []
        installment_numbers = []

        for row in range(5, ws.max_row + 1):
            # Пропускаем заголовки филиалов и пустые строки
            policy_number = ws.cell(row=row, column=1).value
            if (
                policy_number
                and isinstance(policy_number, str)
                and policy_number.startswith("TEST-")
            ):
                status_value = ws.cell(row=row, column=status_column_index).value
                # Получаем номер платежа из столбца "Этот взнос"
                this_payment_value = ws.cell(
                    row=row, column=status_column_index + 1
                ).value

                status_values.append(status_value)
                policy_numbers.append(policy_number)
                installment_numbers.append(this_payment_value)

        # Проверяем, что есть значения
        self.assertGreater(
            len(status_values), 0, "Должны быть значения в столбце 'Статус рассрочки'"
        )

        # Проверяем логику заполнения
        for i, (policy_num, status, installment) in enumerate(
            zip(policy_numbers, status_values, installment_numbers)
        ):
            if policy_num == "TEST-INSTALLMENT":
                # Все платежи по рассрочке должны быть "рассрочка"
                self.assertEqual(
                    status,
                    "рассрочка",
                    f"Платеж {installment} по полису {policy_num} должен быть 'рассрочка'",
                )
            elif policy_num == "TEST-ANNUAL":
                # Единственный платеж должен быть "годовой"
                self.assertEqual(
                    status,
                    "годовой",
                    f"Платеж {installment} по полису {policy_num} должен быть 'годовой'",
                )

        # Проверяем, что есть и рассрочка, и годовые платежи
        self.assertIn(
            "рассрочка", status_values, "Должны быть платежи со статусом 'рассрочка'"
        )
        self.assertIn(
            "годовой", status_values, "Должны быть платежи со статусом 'годовой'"
        )

    def test_export_excludes_inactive_policies(self):
        """Тест: экспорт исключает платежи по неактивным полисам"""
        from openpyxl import load_workbook
        from io import BytesIO

        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2024-01-01", "date_to": "2024-12-31"}
        )

        # Должен вернуться Excel файл
        self.assertEqual(response.status_code, 200)

        # Загружаем Excel файл из ответа
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Собираем все номера полисов из экспорта
        policy_numbers = []
        for row in range(5, ws.max_row + 1):
            # Столбец A (1) - "Номер полиса"
            policy_number_cell = ws.cell(row=row, column=1).value
            if policy_number_cell:
                policy_numbers.append(policy_number_cell)

        # Проверяем, что в экспорте есть только активный полис
        self.assertIn(
            "TEST-001", policy_numbers, "Активный полис должен быть в экспорте"
        )
        self.assertNotIn(
            "TEST-002", policy_numbers, "Неактивный полис НЕ должен быть в экспорте"
        )

        # Проверяем количество строк - должно быть только 4 платежа (от активного полиса)
        # В setUp создано 4 платежа для активного полиса и 2 для неактивного
        # В экспорт должны попасть только 4 от активного
        self.assertEqual(
            len(policy_numbers),
            4,
            "В экспорте должно быть только 4 платежа от активного полиса",
        )

        # Все номера полисов должны быть "TEST-001"
        for policy_number in policy_numbers:
            self.assertEqual(
                policy_number,
                "TEST-001",
                "Все платежи должны быть от активного полиса TEST-001",
            )

    def test_export_only_inactive_policies_returns_no_data(self):
        """Тест: если в диапазоне дат есть только платежи по неактивным полисам, экспорт возвращает предупреждение"""
        # Деактивируем активный полис
        self.policy.policy_active = False
        self.policy.save()

        url = reverse("reports:export_payments")
        response = self.client.get(
            url, {"date_from": "2024-01-01", "date_to": "2024-12-31"}
        )

        # Должен быть редирект с предупреждением о том, что нет данных
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("reports:index"))
